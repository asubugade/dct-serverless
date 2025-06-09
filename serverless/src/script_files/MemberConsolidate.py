import shutil
from pathlib import Path
import sys
import json
import pandas
from openpyxl.reader.excel import load_workbook
import openpyxl
import os
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors
from openpyxl import styles
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import xlsxwriter
import pydash
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode
from anytree.importer import JsonImporter
from anytree.importer import DictImporter
from anytree.exporter import DotExporter
from anytree import AnyNode, RenderTree

from MessageHandling import MessageHandling
import CommonValidations
import datetime
import xml.etree.cElementTree as ET
from collections import defaultdict

from datetime import datetime
oNow = datetime.now()

import re
from pyconfig import loadConfig
from pyconfig.model_common import model_common_Cls
from pyconfig import awsCloudFileManager

from bson import ObjectId
from pyconfig.LogService import LogService
from shutil import copyfile
import glob
import pandas.api.types as pdt
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

sys.path.insert(0, os.path.abspath(__file__))

# Explore Parameteres
# oPostParameters = json.loads(sys.argv[1])

# with open(r'c:\vartemp\tempmembercosolidate.txt','w') as f:
#     f.write(str(sys.argv[1]))
# with open(r'c:\vartemp\tempmembercosolidate.txt','r') as f:
#     oPostParameters = json.loads(f.read())
model_common = model_common_Cls()

# id_schedule =str(sys.argv[1])
id_schedule = '683d9046ee80995c70bd1c9f'
sch_data =  model_common.get_LaneScheduleById(id_schedule)
if not sch_data :
    LogService.log('No Scheule id found , system exit')
    sys.exit()
oPostParameters = sch_data.get('jDump')
MessageHandling = MessageHandling()
iTemplateID = oPostParameters['iTemplateID']
oDataFrameList = model_common.FunDCT_GetTemplateLog('gen_templates',iTemplateID)
cTemplateType = oDataFrameList[0]['cTemplateType']
cFileDir = oPostParameters['cDirFile']
_iTemplateMemberSubmissionID = oPostParameters['_iTemplateMemberSubmissionID']
cFileType = '.xlsx'
cScacReuested = oPostParameters['SCAC']
cSamplfilepath = str(cFileDir) + cScacReuested + '_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,oDataFrameList[0]['cTemplateSampleFile'],cSamplfilepath)
oWorkbookTemplate = load_workbook(cSamplfilepath)
oSheetTemplate = oWorkbookTemplate.active
iMaxRowTemplate = oSheetTemplate.max_row
iMaxColumnTemplate = oSheetTemplate.max_column
aUnMappedColumnLabels = defaultdict(dict)
aUnMappedColumnRange = defaultdict(dict)
aUnMappedColumnRangeDetails = defaultdict(dict)
aRangeStart = []
aRangeEnd = []
_cCompanyname =cScacReuested
startProcessRequestDetails = {
                "bProcesslock": "Y",
                "tProcessStart": datetime.now(),
                "tUpdated": datetime.now()
            }
model_common.FunDCT_UpdateConsolidateTemplateStartOrEndprocessingAndProcesslock(_iTemplateMemberSubmissionID, startProcessRequestDetails)
iMaxDepthHeaders = model_common.FunDCT_GetMaxDepthHeader(iTemplateID)
model_common.FunDCT_UpdateMemberSubmissionlogsStartprocessingAndProcesslock(_iTemplateMemberSubmissionID)
def FunDCT_MemberAllFile():
    try:
        oConsolidationTemplateLog = model_common.FunDCT_GetConsolidationData(iTemplateID)
        cFilePath =  str(cFileDir) + cScacReuested + '_Submission' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'

        oJaonDataS3 = oConsolidationTemplateLog[0]['aConsolidationData'][cScacReuested][0]['aConsolidationData']

        oJaonData = FunDCT_LoadJsonByObjectKey(oJaonDataS3)
        if oJaonData:
            if str(oPostParameters['preFillData']).lower() == 'no':
                jsondDta = json.loads(oJaonData)
                exlWb = openpyxl.Workbook()
                del exlWb['Sheet']
                bookobjWriter = pandas.ExcelWriter(cFilePath, engine='openpyxl')
                bookobjWriter.book = exlWb
                for eahobj  in jsondDta:
                    oDataFrame = pandas.read_json(jsondDta[eahobj])
                    for x in oDataFrame:
                        if pdt.is_datetime64_any_dtype(oDataFrame[x]):
                            oDataFrame[x] = oDataFrame[x].dt.strftime('%Y-%m-%d %M:%H:%S')
                    oDataFrame.to_excel(bookobjWriter,sheet_name=eahobj, index=False)
                bookobjWriter.save()
                bookobjWriter.close()
            else:
                oDataFrame = pandas.read_json(oJaonData)
                oDataFrame.to_excel(cFilePath,index=False,engine="openpyxl",)
                oBook = openpyxl.load_workbook(cFilePath)
                oBookSheet = oBook.active
                for col in range(1,oBookSheet.max_column + 1):
                    try:
                            if "#DCT#" in oBookSheet.cell(iMaxDepthHeaders,col).value:
                                oBookSheet.cell(iMaxDepthHeaders,col).value = ''
                    except:
                        pass
            #######################################################
            #######Onlly header Mapping #####################
            aKeys = []
            bColumnHeaderNotInRange = False
            bInUnmappedRange = False
            if len(aUnMappedColumnLabels) > 0:
                aUnMappedColumnRangeList = [(aRangeStart[iRangeIndex], aRangeEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeStart))]
                bInUnmappedRange = True
                bColumnHeaderNotInRange = True
                aKeys = aUnMappedColumnLabels[_cCompanyname]['_cCompanyname'][0].keys()
            
            iIndex = 0
            for iRow in range(1, iMaxRowTemplate+1):
                iSubsctractCols = 0
                for jColumn in range(1, iMaxColumnTemplate+1):
                    oSampleCell = oSheetTemplate.cell(row=iRow, column=jColumn)
                    oSampleCellVal = oSampleCell.value
                    if len(aKeys) > 0:
                        bColumnHeaderNotInRange = True if str(oSampleCellVal) in aKeys else False
                        bInUnmappedRange = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aUnMappedColumnRangeList)
                    
                    iMergedCellLen = FunDCT_GetMergeCellRangeLen(oSheetTemplate, iRow, jColumn)
                    if (oSampleCellVal is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen > 0:
                        iSubsctractCols += iMergedCellLen
                        continue
                    elif (oSampleCellVal is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen <= 0:
                        iSubsctractCols += 1
                        continue
                    # elif (oSampleCellVal is None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True):
                    #     iSubsctractCols += 1
                    #     continue

                    # iBackwardCol = int(jColumn) - int(iSubsctractCols)
                    # oSampleCell = oSheetTemplate.cell(row=iRow, column=iBackwardCol)
                    # oSampleCellVal = oSampleCell.value
                    

                    if oSampleCellVal is not None:
                        iBackwardCol = int(jColumn) - int(iSubsctractCols)
                        oUploadedCell = oBookSheet.cell(row=iRow, column=iBackwardCol)
                        oUploadedCellVal = str(oUploadedCell.value)
                        bMatchHeaderValue = True if str(oSampleCellVal).strip() == str(oUploadedCellVal) else False

                        # if not bMatchHeaderValue:
                        #     cErrorMessage = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE','TEMPLATE_VALIDATE_HEADER_MISMATCH').replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$', str(oSampleCellVal)).replace('#$DCTVARIABLE_INCORRECT_CELL_VAL#$', str(oUploadedCellVal) )
                        #     FunDCT_PrepareTemplateStatusFile(iIndex, oUploadedCell.row, oUploadedCell.column, str(oUploadedCell.coordinate), cErrorMessage)
                        #     iIndex += 1
                        # else:
                        oBookSheet.cell(iRow, iBackwardCol).value = oSampleCell.value                       
                        oBookSheet.cell(iRow, iBackwardCol).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cTextColor = FunDCT_GetTextColor(oSampleCell)
                        cBackColor = FunDCT_GetBackColor(oSampleCell,cTextColor)
                        thin_border = Side(style="thin")
                        oBookSheet.cell(row=iRow, column=iBackwardCol).border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                        oBookSheet.cell(row=iRow, column=iBackwardCol).font = Font(color=cTextColor)
                        oBookSheet.cell(row=iRow, column=iBackwardCol).fill = PatternFill(fgColor=cBackColor,fill_type = "solid")
                        if FunDCT_CellMerged(oSheetTemplate, iRow, iBackwardCol):
                            if iMergedCellLen > 0:
                                ranges = f"{get_column_letter(iBackwardCol)}{iRow}:{get_column_letter((iBackwardCol - 1)+iMergedCellLen)}{iRow}"
                                oBookSheet.merge_cells(ranges)
                ################## Heading End
                
            oBook.save(cFilePath)
                
            cS3DirUploadFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_CONSOLIDATION) + '/' + cScacReuested + '/'
            cS3UrlUploadedOrg = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirUploadFile),cFilePath)
            if os.path.exists(cFilePath):
                    os.remove(cFilePath)
            if os.path.exists(cSamplfilepath):
                    os.remove(cSamplfilepath)
                
            oStatusDetails = model_common.statusActive[0]
            iStatusID = oStatusDetails.get('_id')
            existing_data = model_common.getTmplMembersubmissionReqDetails(iStatusID,_iTemplateMemberSubmissionID)
            is_consolidation = existing_data.get("IsConsolidationRequested", {})
            IsConsolidationRequested = {
                "aMemberScacCode": is_consolidation.get("aMemberScacCode", ""),
                "iEnteredby": is_consolidation.get("iEnteredby", 0),
                "iTemplateID": is_consolidation.get("iTemplateID", 0),
                "IsAlreadyRequested": False
            }
            aRequestDetails = {
                "cTemplateStatusFile": cS3UrlUploadedOrg,
                "IsConsolidationRequested": IsConsolidationRequested,
                "bProcesslock": "N",
                "bProcessed": "Y",
                "tProcessEnd": datetime.utcnow(),
                "tUpdated": datetime.utcnow()
            }
            model_common.FunDCT_UpdateMemberSubmissionStartOrEndprocessingAndProcesslock(_iTemplateMemberSubmissionID,aRequestDetails)
            oTemplateDetails = model_common.FunDCT_GetTemplateDetails(oPostParameters['iTemplateID'])
            model_common.FunDCT_SendMemberSubmissionFile(cS3UrlUploadedOrg, oTemplateDetails, oPostParameters)
            
            # return MessageHandling.FunDCT_MessageHandling('Success', cS3UrlUploadedOrg)
        else:
            oStatusDetails = model_common.statusActive[0]
            iStatusID = oStatusDetails.get('_id')
            existing_data = model_common.getTmplMembersubmissionReqDetails(iStatusID,_iTemplateMemberSubmissionID)
            is_consolidation = existing_data.get("IsConsolidationRequested", {})
            IsConsolidationRequested = {
                "aMemberScacCode": is_consolidation.get("aMemberScacCode", ""),
                "iEnteredby": is_consolidation.get("iEnteredby", 0),
                "iTemplateID": is_consolidation.get("iTemplateID", 0),
                "IsAlreadyRequested": False
            }
            aRequestDetails = {
                "cTemplateStatusFile": cS3UrlUploadedOrg,
                "IsConsolidationRequested": IsConsolidationRequested,
                "bProcesslock": "N",
                "bProcessed": "Y",
                "tProcessEnd": datetime.utcnow(),
                "tUpdated": datetime.utcnow()
            }
            model_common.FunDCT_UpdateMemberSubmissionStartOrEndprocessingAndProcesslock(_iTemplateMemberSubmissionID,aRequestDetails)

            # return MessageHandling.FunDCT_MessageHandling('Error', 'No Data')
    
    except Exception as e:
        oStatusDetails = model_common.statusActive[0]
        iStatusID = oStatusDetails.get('_id')
        existing_data = model_common.getTmplMembersubmissionReqDetails(iStatusID,_iTemplateMemberSubmissionID)
        is_consolidation = existing_data.get("IsConsolidationRequested", {})
        IsConsolidationRequested = {
                "aMemberScacCode": is_consolidation.get("aMemberScacCode", ""),
                "iEnteredby": is_consolidation.get("iEnteredby", 0),
                "iTemplateID": is_consolidation.get("iTemplateID", 0),
                "IsAlreadyRequested": False
            }
        aRequestDetails = {
                "cTemplateStatusFile": cS3UrlUploadedOrg,
                "IsConsolidationRequested": IsConsolidationRequested,
                "bProcesslock": "N",
                "bProcessed": "Y",
                "tProcessEnd": datetime.utcnow(),
                "tUpdated": datetime.utcnow()
            }
        model_common.FunDCT_UpdateMemberSubmissionStartOrEndprocessingAndProcesslock(_iTemplateMemberSubmissionID,aRequestDetails)

        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
    # return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_MergeAllTemplateFile Error while parsing file due to ' + str(e) + 'Line no - '+ str(line))


def FunDCT_CellMerged(oSheet, iRow, jColumn):
    try:
        oCell = oSheet.cell(iRow, jColumn)
        for mergedCell in oSheet.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                return True
        return False
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CellMerged Error while parsing file due to ' + str(e))

def FunDCT_GetMergeCellRangeLen(oSheetTemplate, iRow, jColumn):
    try:
        iTotalLen = 0
        oCell = oSheetTemplate.cell(iRow, jColumn)
        for mergedCell in oSheetTemplate.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                cCellRangeStart, cCellRangeEnd = str(mergedCell).split(':')
                iCellCol = CommonValidations.FunDCT_GetColumnNumber(
                    ''.join([i for i in cCellRangeStart if not i.isdigit()]))
                iTotalLen = int(CommonValidations.FunDCT_GetColumnNumber(
                    ''.join([i for i in cCellRangeEnd if not i.isdigit()]))) - (int(iCellCol)-1)
        return iTotalLen
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetMergeCellRangeLen Error while parsing file due to ' + str(e))

def FunDCT_LoadJsonByObjectKey(s3objKey):
    if s3objKey  == '':
        return ''
    try:
        filetype = s3objKey.split('.')[-1]
        if filetype.startswith('xls'):
            cTempfilePath = str(cFileDir) +'Tempary_memberconsolidaiton_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
            awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,s3objKey,cTempfilePath)
            loadwb = load_workbook(cTempfilePath,read_only=True)
            odataframe = pandas.ExcelFile(loadwb,engine="openpyxl").parse()
            # for x in odataframe:
            #     if pdt.is_datetime64_any_dtype(odataframe[x]):
            #             odataframe[x] = odataframe[x].dt.strftime('%Y-%m-%d %M:%H:%S')
            iIndexDataFrame = iter(range(1, len(odataframe.columns) + 1))
            # for col in odataframe.columns:
            #     if col.startswith('Unnamed'):
            #         odataframe.drop(columns=[col],inplace=True)
            odataframe = odataframe.dropna(axis=1, how='all')
        
            if iMaxRowTemplate > 1 :
                col = ''
                ii = 0
                for x in odataframe.iloc[0]:

                    if x:
                        if 'date' in str(x) or 'Date' in str(x):
                            col = list(odataframe.columns)[ii]                        
                            odataframe[col] = odataframe[col].apply(dateconc)
                            # odataframe[col] = odataframe[col].dt.strftime('%Y-%m-%d %M:%H:%S')

                    ii+=1
            else:
                for x in odataframe:
                    if pdt.is_datetime64_any_dtype(odataframe[x]):
                            odataframe[x] = odataframe[x].dt.strftime('%Y-%m-%d %M:%H:%S')
            odataframe.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in odataframe.columns]
            # jData = odataframe.to_json(date_format="epoch")
            # jData = odataframe.to_json(date_format="iso")
            jData = odataframe.to_json()
            if os.path.exists(cTempfilePath): os.remove(cTempfilePath)
        else:
            oData = awsCloudFileManager.download_data_from_bucket(loadConfig.AWS_BUCKET,s3objKey)
            oData =str(oData).replace('null','\\"\\"')
            jsoLod = json.loads(oData)
            odataframe = pandas.DataFrame(eval(jsoLod),dtype=str,)
            for x in odataframe:
                if pdt.is_datetime64_any_dtype(odataframe[x]):
                        odataframe[x] = odataframe[x].dt.strftime('%Y-%m-%d %M:%H:%S')

            iIndexDataFrame = iter(range(1, len(odataframe.columns) + 1))
            odataframe.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in odataframe.columns]
            df = odataframe.applymap(lambda x: x.replace("\\", "") if isinstance(x, str) else x)
            jData = df.to_json()
        return jData
    except Exception as E:
        return ''
def dateconc(val):
    try:
        return val.strftime('%Y-%m-%d %H:%M:%S')
    except:
        return val
def FunDCT_GetUnMappednColumnDetails(_cCompanyname):
    try:
        aGetUnMappedColumnLabels = model_common.FunDCT_GetUnMappednColumns(iTemplateID)
        for cMemberColumnKey, aMemberColumnDetails in enumerate(aGetUnMappedColumnLabels):
            for iIndex in range(len(aMemberColumnDetails['aUnMappedColumns'])):
                if _cCompanyname in aMemberColumnDetails['aUnMappedColumns'].keys():
                    for cKey, cVal in aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0].items():
                        iRangeStart = int(aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0][cKey]['iActualCellCol'])
                        iTotalLen = int(aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0][cKey]['iTotalLen'])
                        iRangeEnd = int(iRangeStart) + int(iTotalLen-1)
                        
                        aRangeStart.append(iRangeStart)
                        aRangeEnd.append(int(iRangeStart) + int(iTotalLen)-1)
                        aUnMappedColumnRange[iRangeStart] = {
                            'iRangeStart': iRangeStart,
                            'iRangeEnd': iRangeEnd,
                            'iTotalLen': iTotalLen,
                        }
                        aUnMappedColumnRangeDetails[str(iRangeStart)+"-"+str(iRangeEnd)] = {
                            'iRangeStart': iRangeStart,
                            'iRangeEnd': iRangeEnd,
                            'iTotalLen': iTotalLen,
                        }
                        aUnMappedColumnLabels[str(_cCompanyname)] = {
                            '_cCompanyname': aMemberColumnDetails['aUnMappedColumns'][_cCompanyname],
                        }
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetUnMappednColumnDetails Error while parsing file due to ' + str(e))
FunDCT_GetUnMappednColumnDetails(_cCompanyname)
def FunDCT_GetTextColor(oCell):
    try:
        cTextColor = '000000'
        if oCell.font.color:
            if 'rgb' in oCell.font.color.__dict__:
                cTextColor = oCell.font.color.rgb[2:]          

        return cTextColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))

def FunDCT_GetBackColor(oCell, cTextColor):
    try:
        cBackColor = 'ffffff'
        
        if oCell.fill.start_color != '' and type(oCell.fill.start_color.index) != int:                    
            cBackColor = oCell.fill.start_color.index[2:]
            if cTextColor == '000000' and cBackColor == '000000':
                cBackColor = 'ffffff'
        elif oCell.fill.start_color != '' and type(oCell.fill.start_color.index) == int:
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index][2:])
        elif isinstance(oCell.fill, PatternFill) and oCell.fill.patternType == 'solid':
            if oCell.fill.fgColor.type =='rgb':
                cBackColor = oCell.fill.fgColor.rgb[2:]
            elif oCell.fill.fgColor.type =='indexed':
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index -1][2:])
        
        return cBackColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))

FunDCT_MemberAllFile()
