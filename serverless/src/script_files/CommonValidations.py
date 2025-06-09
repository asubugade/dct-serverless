from copy import copy
from pathlib import Path
import sys
import inspect
import os
from pyconfig.MemoryService import MemoryService
from pyconfig.LogService import LogService

path = str(Path(Path(__file__).parent.absolute()).parent.absolute())
sys.path.insert(0, path)
file_dir = os.path.dirname(__file__)
sys.path.append(file_dir)
cmd_folder = os.path.realpath(os.path.abspath(os.path.split(inspect.getfile( inspect.currentframe() ))[0]))
if cmd_folder not in sys.path:
    sys.path.insert(0, cmd_folder)

import json
import pandas
import numpy as np
from openpyxl.reader.excel import load_workbook
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors, Protection,protection
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import xlsxwriter
import pydash
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode
from anytree.importer import JsonImporter
from anytree.importer import DictImporter
from anytree.exporter import DotExporter
from anytree import AnyNode, RenderTree
from script_files import MessageHandling
import datetime
import xml.etree.cElementTree as ET
from collections import defaultdict
from openpyxl.comments import Comment
from openpyxl.utils import units
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

import shutil
from shutil import copyfile
from pathlib import Path
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

import re
from pyconfig import loadConfig
from pyconfig.model_common import model_common_Cls
from datetime import datetime
from openpyxl.utils import get_column_letter
oNow = datetime.now()

aDistributionMembers = defaultdict(dict)
aNSrow = list()
###############################
# COMMON VALIDATIONS FUNCTIONS
###############################
MessageHandling = MessageHandling.MessageHandling()
model_common = model_common_Cls()
def FunDCT_CommonValidations(oPostParameters):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'MANDATORY')
        if oPostParameters['_cUploadType'] != 'DISTRIBUTE':
            if oPostParameters['cValidationString'] == 'MANDATORY':
                bMatch = FunDCT_MandatoryVal(oPostParameters['oCellVal'])
                return bMatch
            else: 
                return True
        else:
            return True
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CommonValidations Error while parsing file due to ' + str(e))


def FunDCT_MandatoryVal(cCellVal):
    try:
        # bMatch = False
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'MANDATORY')
        if cCellVal is not None:
            bMatch = True
        #Below else added by spirgonde for jira WNT2023-1649    
        # else:
        #     bMatch = False   
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_MandatoryVal Error while parsing file due to ' + str(e))

###############################
# RANGE VALIDATIONS FUNCTIONS
###############################

def FunDCT_RangeValidations(oPostParameters):
    try:
        aExplRange = ''
        if oPostParameters['cValidationString'].startswith('TEXT_RANGE('):
            bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'INVALID_NUMERIC_RANGE')
            aExplRange = oPostParameters['cValidationString'].replace(
                ')', '').split('TEXT_RANGE(')[1]
            bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'INVALID_NUMERIC_RANGE').replace('#$valrange$#',aExplRange)
            [iRangeFrom, iRangeTo] = [int(iRangeFrom) for iRangeFrom in aExplRange.split(
                loadConfig.PY_RANGE_VALUE_SEPARATOR)]
            if oPostParameters['oCellVal'] != '' and oPostParameters['oCellVal'] is not None:
                bMatch = FunDCT_TextRange(oPostParameters['oCellVal'], iRangeFrom, iRangeTo)            
        if oPostParameters['cValidationString'].startswith('NUMERIC_RANGE('):
            bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'INVALID_NUMERIC_RANGE')
            aExplRange = oPostParameters['cValidationString'].replace(
                ')', '').split('NUMERIC_RANGE(')[1]
            bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'INVALID_NUMERIC_RANGE').replace('#$valrange$#',aExplRange)
            [iRangeFrom, iRangeTo] = [float(iRangeFrom) for iRangeFrom in aExplRange.split(
                loadConfig.PY_RANGE_VALUE_SEPARATOR)]
            # Added Asif jira no #1974
            try:
                oPostParameters['oCellVal'] = int(oPostParameters['oCellVal'])
            except:
                try:
                    oPostParameters['oCellVal'] = float(oPostParameters['oCellVal'])
                except:
                    pass
            # if oPostParameters['NS'] and str(oPostParameters['oCellVal']).strip().upper() != 'NS':
            #     bMatch = FunDCT_NS(oPostParameters['oSheetUploaded'].cell(row=oPostParameters['iRow'], column=oPostParameters['iCol']))
                
            if (isinstance(oPostParameters['oCellVal'], int) or isinstance(oPostParameters['oCellVal'], float)) and (iRangeFrom == 0 and iRangeTo == 0):
                bMatch = True
            
            elif isinstance(oPostParameters['oCellVal'], int):
                bMatch = FunDCT_NumericIntRange(oPostParameters['oCellVal'], iRangeFrom, iRangeTo)
            elif isinstance(oPostParameters['oCellVal'], float):
                bMatch = FunDCT_NumericFloatRange(
                    oPostParameters['oCellVal'], iRangeFrom, iRangeTo)
            
            if bMatch is not True:
                bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'INVALID_NUMERIC_RANGE').replace('#$valrange$#',aExplRange)
            
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_RangeValidations Error while parsing file due to ' + str(e))

def FunDCT_NS(oCell):
    bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'NO_SERVICE')
    try:
        if str(oCell.value).lower() == 'ns':
            bMatch = True
        
        return bMatch
    except Exception as err:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_NumericFloatRange Error while parsing file due to ' + str(err))


def FunDCT_NumericFloatRange(cCellVal, iRangeFrom, iRangeTo):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_NUMERIC_RANGE')
        if iRangeFrom != 0 and iRangeTo == 0:
            if iRangeFrom < float(cCellVal) < 1000000000.0:
                bMatch = True
        elif iRangeFrom != 0 and iRangeTo != 0:
            if iRangeFrom <= float(cCellVal) <= iRangeTo:
                bMatch = True
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_NumericFloatRange Error while parsing file due to ' + str(e))


def FunDCT_NumericIntRange(cCellVal, iRangeFrom, iRangeTo):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_NUMERIC_RANGE')
        if iRangeFrom != 0 and iRangeTo == 0:
            if int(iRangeFrom) <= int(cCellVal):
                bMatch = True
        elif iRangeFrom != 0 and iRangeTo != 0:
            if int(iRangeFrom) <= int(cCellVal) <= int(iRangeTo):
                bMatch = True
        # elif iRangeFrom == 0 and iRangeTo != 0:
        #     if int(iRangeFrom) >= int(cCellVal):
        #         bMatch = True
        # elif str(iRangeFrom) !='' and str(iRangeTo) != '':
        #     if str(iRangeFrom) <= str(cCellVal) <= str(iRangeTo):
        #         bMatch = True
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_NumericIntRange Error while parsing file due to ' + str(e))

def FunDCT_TextRange(cCellVal, iRangeFrom, iRangeTo):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_TEXT_RANGE')
        if iRangeFrom == 0 and iRangeTo == 0:
            bMatch = True
        elif iRangeFrom != 0 and iRangeTo == 0:
            iCount = len(str(cCellVal))
            if iCount in range(iRangeFrom, iRangeTo):
                bMatch = True
        elif iRangeFrom != 0 and iRangeTo != 0:
            iCount = len(str(cCellVal))
            ## 'iRangeTo' changed to 'iRangeTo + 1'
            if iCount in range(iRangeFrom, iRangeTo + 1):
                bMatch = True

        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_TextRange Error while parsing file due to ' + str(e))


def FunDCT_AllowedTextOnly(cCellVal, aAllowedWords):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_ALLOWED_TEXT')
        if(str(cCellVal).isalpha()):
         
        ##code added by spirgonde to remove white spaces rom allowedwords WNT2023-1649 START...   
         if isinstance(aAllowedWords, list):
          aAllowedWordsNew = []
          for i in aAllowedWords:
           aAllowedWordsNew.append(str(i).strip())  
         
         else:
            aAllowedWordsNew = aAllowedWords   
        ##code added by spirgonde to remove white spaces rom allowedwords WNT2023-1649 END.  
          
         if cCellVal in aAllowedWordsNew: ## changed aAllowedWords to aAllowedWordsNew for jira WNT2023-1649
            bMatch = True

        return bMatch
    except Exception as e:
        ## print('cCellVal',cCellVal,'aAllowedWords',aAllowedWords)
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_AllowedTextOnly Error while parsing file due to ' + str(e))

###############################
# RESTRICTED TEXT VALIDATIONS FUNCTIONS
###############################


def FunDCT_RestrictedTextValidations(oPostParameters):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_RESTRICTED_TEXT')
        if oPostParameters['cValidationString'].startswith('RESTRICTED_TEXT('):
            aExplRange = oPostParameters['cValidationString'].replace(
                ')', '').split('RESTRICTED_TEXT(')[1]
            bMatch = FunDCT_RestrictedText(oPostParameters['oCellVal'], aExplRange)
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_RestrictedTextValidations Error while parsing file due to ' + str(e))
def FunDCT_DATAValidations(oPostParameters):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_RESTRICTED_TEXT')
        if oPostParameters['cValidationString'].startswith('DATA_VALIDATION('):
            aExplRange = oPostParameters['cValidationString'].replace(
                ')', '').split('DATA_VALIDATION(')[1]
            bMatch = FunDCT_RestrictedText(oPostParameters['oCellVal'], aExplRange)
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_RestrictedTextValidations Error while parsing file due to ' + str(e))


def FunDCT_RestrictedText(cCellVal, aRestrictedWords):
    try:
        bMatch = True
        ## if cCellVal is not None and .strip() added below by spirgonde for validation issue for jira WNT2023-1649
        if cCellVal is not None:
            try:
                if not str(cCellVal).strip() in str(aRestrictedWords).split(','):   
                    bMatch = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE', 'INVALID_RESTRICTED_TEXT').replace(
                    '#$DCTVARIABLE_INCORRECT_CELL_VAL#$', str(cCellVal)).replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$', aRestrictedWords)
            except:
                return MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE', 'INVALID_RESTRICTED_TEXT').replace(
                    '#$DCTVARIABLE_INCORRECT_CELL_VAL#$', str(cCellVal)).replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$', aRestrictedWords)
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_RestrictedText Error while parsing file due to ' + str(e))

###############################
# DATE VALIDATIONS FUNCTIONS
###############################

#Code added by spirgonde for new Date Format dd-mmm-yyyy start ...
def FunDCT_ConvertAlphanumericDateToNumeric(cCellVal):
    try:
         cDay, cMonth,cYear  = cCellVal.split('-')
         month = datetime.strptime(cMonth,'%b')
         cMonth = month.strftime("%m")
         dateTuple = (cYear,cMonth,cDay)
         cCellVal= "-".join(dateTuple)
         return cCellVal
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DateRangeValidations Error while parsing file due to ' + str(e) +' Line No - ' + str(line))

#Code added by spirgonde for new Date Format dd-mmm-yyyy start end .


def FunDCT_DateRangeValidations(oPostParameters):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_DATERANGE')
        if oPostParameters['cValidationString'].startswith('DATE_RANGE('):
            aExplRange = oPostParameters['cValidationString'].replace(
                ')', '').split('DATE_RANGE(')[1]
            [iRangeFrom, iRangeTo] = aExplRange.split(
                loadConfig.PY_RANGE_VALUE_SEPARATOR)
            
            #Code added by spirgonde for new Date Format dd-mmm-yyyy start
            dateFormat = "ymd"
            if "-" in str(iRangeFrom):
             cDay, cMonth,cYear  = iRangeFrom.split('-') 
             length = len(str(cMonth))
             if length == 3 :
              dateFormat = "mby"
           
            try:
             if  isinstance(oPostParameters['oCellVal'], datetime):
                 aCellVal = oPostParameters['oCellVal'].strftime('%Y-%m-%d')

             else:    
                x = datetime.strptime(oPostParameters['oCellVal'],'%Y-%m-%d %H:%M:%S')
                aCellVal = str(oPostParameters['oCellVal']).split(' ')[0]
            except Exception:
                bMatch = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE', 'INVALID_DATERANGE').replace(
                    '#$from#$', str(iRangeFrom)).replace('#$to#$', str(iRangeTo))
                return(bMatch)
                 #Code added by spirgonde for new Date Format dd-mmm-yyyy end 
            bMatch = FunDCT_DateRange(aCellVal, iRangeFrom, iRangeTo,dateFormat) #dateFormat added by spirgonde for new Date Format dd-mmm-yyyy 
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DateRangeValidations Error while parsing file due to ' + str(e))


def FunDCT_DateValidator(cCellVal):
    bIsValidDate = True
    try:
            cYear, cMonth, cDay = cCellVal.split('-')
            datetime(int(cYear), int(cMonth),int(cDay)) #changed datetime.datetime(int(cYear), int(cMonth),int(cDay)) to datetime(int(cYear),int(cMonth),int(cDay)) by spirgonde for new Date Format dd-mmm-yyyy start
    except ValueError as e: 
        bIsValidDate = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_DATERANGE') 
    return bIsValidDate


def FunDCT_DateRangeValidator(cCellVal, cStartDate, cEndDate): 
    try:
        bIsValidDateRange = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_DATERANGE')
        tInputDate = datetime.strptime(cCellVal, "%Y-%m-%d") #changed datetime.datetime.strptime(cCellVal, "%Y-%m-%d") to datetime.strptime(cCellVal, "%Y-%m-%d") by spirgonde for new Date Format dd-mmm-yyyy start
        tStartDate = datetime.strptime(cStartDate, "%Y-%m-%d") #changed datetime.datetime.strptime(cCellVal, "%Y-%m-%d") to datetime.strptime(cCellVal, "%Y-%m-%d") by spirgonde for new Date Format dd-mmm-yyyy start
        tEndDate = datetime.strptime(cEndDate, "%Y-%m-%d")  #changed datetime.datetime.strptime(cCellVal, "%Y-%m-%d") to datetime.strptime(cCellVal, "%Y-%m-%d") by spirgonde for new Date Format dd-mmm-yyyy start
        if tStartDate <= tInputDate <= tEndDate:
            bIsValidDateRange = True
      
        return bIsValidDateRange
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DateRangeValidator Error while parsing file due to ' + str(e))


def FunDCT_CompareDate(tDateOne, tDateTwo, cCompareString):  
    try:
       
        bIsValidDateRange = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_DATERANGE')
        if cCompareString == '>':
            if tDateOne > tDateTwo:
                bIsValidDateRange = True
        elif cCompareString == '>=':
            if tDateOne >= tDateTwo:
                bIsValidDateRange = True
        elif cCompareString == '<':
            if tDateOne < tDateTwo:
                bIsValidDateRange = True
        elif cCompareString == '<=':
            if tDateOne <= tDateTwo:
                bIsValidDateRange = True
        return bIsValidDateRange
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CompareDate Error while parsing file due to ' + str(e))


def FunDCT_DateRange(tCellVal, tStartDate, tEndDate,dateFormat): #dateFormat added by spirgonde for new Date Format dd-mmm-yyyy 
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_DATERANGE')
        if tStartDate == '0' or tEndDate == '0':
            bMatch = FunDCT_DateValidator(tCellVal)
        elif tEndDate == '>' or tEndDate == '>=' or tEndDate == '<' or tEndDate == '<=':
            bValidateCellVal = FunDCT_DateValidator(tCellVal)
            if dateFormat =="mby":                                            #  if dateFormat =="mby": added by spirgonde for new Date Format dd-mmm-yyyy   
             tStartDate = FunDCT_ConvertAlphanumericDateToNumeric(tStartDate) # tStartDate = FunDCT_ConvertAlphanumericDateToNumeric(tStartDate) added by spirgonde for new Date Format dd-mmm-yyyy 

            bValidateStartDate = FunDCT_DateValidator(tStartDate)
            if bValidateCellVal and bValidateStartDate:
                bMatch = FunDCT_CompareDate(tCellVal, tStartDate, tEndDate) ## additional parameter dateFormat added by spirgonde for new Date Format dd-mmm-yyyy   
              
        else:
            bValidateCellVal = FunDCT_DateValidator(tCellVal)
            if dateFormat =="mby":                                      #  if dateFormat =="mby": added by spirgonde for new Date Format dd-mmm-yyyy   
             tStartDate = FunDCT_ConvertAlphanumericDateToNumeric(tStartDate) #tStartDate = FunDCT_ConvertAlphanumericDateToNumeric(tStartDate) added by spirgonde for new Date Format dd-mmm-yyyy 
             tEndDate = FunDCT_ConvertAlphanumericDateToNumeric(tEndDate) # tEndDate = FunDCT_ConvertAlphanumericDateToNumeric(tEndDate) added by spirgonde for new Date Format dd-mmm-yyyy  

            bValidateStartDate = FunDCT_DateValidator(tStartDate)
            bValidateEndDate = FunDCT_DateValidator(tEndDate) 
        
            if bValidateCellVal and bValidateStartDate and bValidateEndDate:
                bMatch = FunDCT_DateRangeValidator(
                    tCellVal, tStartDate, tEndDate) #additional parameter dateFormat added by spirongde for new Date Format dd-mmm-yyyy  
        if bMatch is not True:
            bMatch = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE', 'INVALID_DATERANGE').replace(
                    '#$from#$', str(tStartDate)).replace('#$to#$', tEndDate)
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DateRange Error while parsing file due to ' + str(e)) 

###############################
# DEFAULT VALIDATIONS FUNCTIONS
###############################

def FunDCT_DefaultValidations(oPostParameters):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
            'APPLICATION', 'SERVER_ERROR')
        
        if oPostParameters['cValidationString'].startswith('DEFAULT_LOCATION'):
            bMatch = FunDCT_Default_Location(str(oPostParameters['oCellVal']).strip())
        elif oPostParameters['cValidationString'].startswith('DEFAULT_COUNTRY'):
            bMatch = FunDCT_Default_Country(str(oPostParameters['oCellVal']).strip())
        elif oPostParameters['cValidationString'].startswith('DEFAULT_ORIGIN_SCAC'):
            bMatch = FunDCT_Default_Scac(oPostParameters)
        elif oPostParameters['cValidationString'].startswith('DEFAULT_OCEAN_SCAC'):
            bMatch = FunDCT_Default_Scac(oPostParameters)
        elif oPostParameters['cValidationString'].startswith('DEFAULT_DESTINATION_SCAC'):
            bMatch = FunDCT_Default_Scac(oPostParameters)    
        elif oPostParameters['_cUploadType'] != 'DISTRIBUTE' and oPostParameters['cValidationString'].startswith('DEFAULT_CURRENCY'):
            bMatch = FunDCT_Default_Currency(str(oPostParameters['oCellVal']).strip())
        elif oPostParameters['_cUploadType'] != 'DISTRIBUTE' and oPostParameters['cValidationString'].startswith('DEFAULT_BASIS'):
            bMatch = FunDCT_Default_Basis(str(oPostParameters['oCellVal']).strip())
        elif oPostParameters['_cUploadType'] != 'DISTRIBUTE' and oPostParameters['cValidationString'].startswith('DEFAULT_CHARGECODE'):
            bMatch = FunDCT_Default_ChargeCode(str(oPostParameters['oCellVal']).strip())
        elif oPostParameters['cValidationString'].startswith('DEFAULT_REGION'):
            bMatch = FunDCT_Default_Region(str(oPostParameters['oCellVal']).strip())
        else:
            bMatch = True


        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DefaultValidations Error while parsing file due to ' + str(e))

def FunDCT_Default_Region(cCellVal):
    try:
        bValidate = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_REGION')
        # oValidRegion = model_common.FunDCT_GetRegions()
        # for i, d in enumerate(oValidRegion):
        #     if d['cCode'] == cCellVal:
        #         bValidate = True
        
        if model_common.FunDCT_checkRegions(cCellVal):
            bValidate= True
        return bValidate
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_Default_Location Error while parsing file due to ' + str(e))


def FunDCT_Default_Location(cCellVal):
    try:
        bValidate = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_LOCATION')
        # oValidLocation = model_common.FunDCT_GetLocations()
        # for i, d in enumerate(oValidLocation):
        #     if d['cCode'] == cCellVal:
        #         bValidate = True
         
        if model_common.FunDCT_CheckLocations(cCellVal):
            bValidate =True
        return bValidate
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_Default_Location Error while parsing file due to ' + str(e))


def FunDCT_Default_Country(cCellVal):
    try:
        bValidate = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_COUNTRY')
        # oValidCountry = model_common.FunDCT_GetCountries()
        # for i, d in enumerate(oValidCountry):
        #     if d['cCode'] == cCellVal:
        #         bValidate = True
        if model_common.FunDCT_CheckCountries(cCellVal):
            bValidate = True

        return bValidate
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_Default_Country Error while parsing file due to ' + str(e))

#FOR Member Scac
def FunDCT_ScacExist(aDistributionMembers, cCellVal):
    try:
        bMatch = False
        for cKey, cDistScaccode in aDistributionMembers.items():
            if cCellVal in aDistributionMembers[cKey]['cDistScaccode']:
                bMatch = True
        return bMatch
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_ScacExist Error while parsing file due to ' + str(e))

def FunDCT_GetHeaderValueForColumn(oSheetUploaded, jColumn):
    try:
        cCellVal = None
        for iIndexRow in range(1, 100):
            try:
                cCellVal = str(oSheetUploaded.iat[iIndexRow-1, jColumn-1])
            except:
                pass
            if cCellVal :
                return cCellVal
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetHeaderValueForColumn Error while parsing file due to ' + str(e))

def FunDCT_GetValidationForColumn(validations,iCol):
    try:
        return validations[iCol]

    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetValidationForColumn Error while parsing file due to ' + str(e))
def FunDCT_CheckAvailableInOtherCharge(oPostParameters):
    try:
        iMaxRowUploaded ,iMaxColumnUploaded = oPostParameters['oSheetUploaded'].shape

        iIndex = 0
        for iIndexRow in range(1, iMaxRowUploaded+1):
            # for jIndexCol in range(1, iMaxColumnUploaded + 1):
            for jIndexCol in model_common.FunDCT_GetSCACHeadersIndex(oPostParameters['oValidationsDetails']):
                
                oCell = oPostParameters['oSheetUploaded'].iat[iIndexRow-1, jIndexCol-1]
                if oCell :
                    cCellVal = str(oCell).strip()
                    if cCellVal == str(oPostParameters['oCellVal']).strip():
                        cHeaderVal = FunDCT_GetHeaderValueForColumn(oPostParameters['oSheetUploaded'], jIndexCol)
                        # below code is commented due to scaccolumn identifying based on values not on validations
                        # if cHeaderVal == 'OFR' or cHeaderVal == 'FOB' or cHeaderVal == 'PLC':
                        #     oPostParameters['aScacAvailableInColumn'][cHeaderVal] = cHeaderVal
                        if 'DEFAULT_SCAC' in oPostParameters['oValidationsDetails'][jIndexCol-1]['cValidations']:
                            oPostParameters['aScacAvailableInColumn'][cHeaderVal] = cHeaderVal
                        elif 'DEFAULT_ORIGIN_SCAC' in oPostParameters['oValidationsDetails'][jIndexCol-1]['cValidations']:
                            oPostParameters['aScacAvailableInColumn'][cHeaderVal] = cHeaderVal
                        elif 'DEFAULT_DESTINATION_SCAC' in oPostParameters['oValidationsDetails'][jIndexCol-1]['cValidations']:
                            oPostParameters['aScacAvailableInColumn'][cHeaderVal] = cHeaderVal
                        elif 'DEFAULT_OCEAN_SCAC' in oPostParameters['oValidationsDetails'][jIndexCol-1]['cValidations']:
                            oPostParameters['aScacAvailableInColumn'][cHeaderVal] = cHeaderVal
        return oPostParameters['aScacAvailableInColumn']
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CheckAvailableInOtherCharge Error while parsing file due to ' + str(e))
def getAllScacHeaders(oVal,subValidation=None):
    tlist = []
    for ov in oVal:
        if subValidation:
            if subValidation in ov['cValidations']:
                return ov['cHeaderLabel']
        else:
            if 'DEFAULT_SCAC' in ov['cValidations']:
                tlist.append(ov['cHeaderLabel'])
            elif 'DEFAULT_ORIGIN_SCAC' in ov['cValidations']:
                tlist.append(ov['cHeaderLabel'])
            elif 'DEFAULT_DESTINATION_SCAC' in ov['cValidations']:
                tlist.append(ov['cHeaderLabel'])
            elif 'DEFAULT_OCEAN_SCAC' in ov['cValidations']:
                tlist.append(ov['cHeaderLabel'])
    return tlist
def getSCACValiations(oVal,dVal):
    tlist = []
    for ov in oVal:
        for do in dVal:
            if do == ov['cHeaderLabel']:
                for k in ov['cValidations']:
                    if 'SCAC' in k:
                        tlist.append(k)
    return tlist
def getSCACValiationsDictByHeaderValidation(oVal,dVal):
    datadic = defaultdict(dict)
    for ov in oVal:
        # for do in dVal:
            # if do == ov['cHeaderLabel']:
            for k in ov['cValidations']:
                if 'SCAC' in k:
                    datadic[k] = {'cHeaderLabel':ov['cHeaderLabel'],'cParentHeader':ov['cParentHeader']}
    return datadic

def FunDCT_SetUnMappendColumns(oPostParameters):
    try:
        listofHeaders = getAllScacHeaders(oPostParameters['oValidationsDetails'])
        ovald = getSCACValiations(oPostParameters['oValidationsDetails'],oPostParameters['aScacAvailableInColumn'])
        dictBaseHead = getSCACValiationsDictByHeaderValidation(oPostParameters['oValidationsDetails'],oPostParameters['aScacAvailableInColumn'])
        oPostParameters['aUnMappedScacColumns'] = defaultdict(dict)
        dDATA = getSCAC_columnMapping(oPostParameters['oValidationsDetails'])
       
        if int(oPostParameters['iMaxDepthHeaders']) == 1:
            if all(key in oPostParameters['aScacAvailableInColumn'] for key in listofHeaders):
                oPostParameters['aUnMappedScacColumns']['all'] = oPostParameters['aScacAvailableInColumn']
            else:
                for eo in ovald:
                    COLN = ''
                    if 'DEFAULT_ORIGIN_SCAC' == eo:
                        listofcolheaders = dDATA.get('DESTINATION_COLUMN')
                        if len(listofcolheaders) > 0:
                            for oLi in listofcolheaders:
                                vald = FunDCT_GetValidationForColumn(oPostParameters['oValidationsDetails'],int(oLi))
                                oPostParameters['aUnMappedScacColumns'][vald['cHeaderLabel']] = vald['cHeaderLabel']
                        listofcolheaders = dDATA.get('OCEAN_COLUMN')
                        if len(listofcolheaders) > 0:
                            for oLi in listofcolheaders:
                                vald = FunDCT_GetValidationForColumn(oPostParameters['oValidationsDetails'],int(oLi))
                                oPostParameters['aUnMappedScacColumns'][vald['cHeaderLabel']] = vald['cHeaderLabel']
                    elif 'DEFAULT_DESTINATION_SCAC' == eo:
                        listofcolheaders = dDATA.get('ORIGIN_COLUMN')
                        if len(listofcolheaders) > 0:
                            for oLi in listofcolheaders:
                                vald = FunDCT_GetValidationForColumn(oPostParameters['oValidationsDetails'],int(oLi))
                                oPostParameters['aUnMappedScacColumns'][vald['cHeaderLabel']] = vald['cHeaderLabel']
                        listofcolheaders = dDATA.get('OCEAN_COLUMN')
                        if len(listofcolheaders) > 0:
                            for oLi in listofcolheaders:
                                vald = FunDCT_GetValidationForColumn(oPostParameters['oValidationsDetails'],int(oLi))
                                oPostParameters['aUnMappedScacColumns'][vald['cHeaderLabel']] = vald['cHeaderLabel']
                    elif 'DEFAULT_OCEAN_SCAC' == eo:
                        listofcolheaders = dDATA.get('DESTINATION_COLUMN')
                        if len(listofcolheaders) > 0:
                            for oLi in listofcolheaders:
                                vald = FunDCT_GetValidationForColumn(oPostParameters['oValidationsDetails'],int(oLi))
                                oPostParameters['aUnMappedScacColumns'][vald['cHeaderLabel']] = vald['cHeaderLabel']
                        listofcolheaders = dDATA.get('ORIGIN_COLUMN')
                        if len(listofcolheaders) > 0:
                            for oLi in listofcolheaders:
                                vald = FunDCT_GetValidationForColumn(oPostParameters['oValidationsDetails'],int(oLi))
                                oPostParameters['aUnMappedScacColumns'][vald['cHeaderLabel']] = vald['cHeaderLabel']
        if int(oPostParameters['iMaxDepthHeaders']) == 2:
            if all(key in oPostParameters['aScacAvailableInColumn'] for key in listofHeaders):
                oPostParameters['aUnMappedScacColumns']['all'] = oPostParameters['aScacAvailableInColumn']
            elif all(key in ovald for key in {"DEFAULT_OCEAN_SCAC","DEFAULT_ORIGIN_SCAC"}):
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'DESTINATION_COLUMN' in conditinalValidation:
                        oPostParameters['aUnMappedScacColumns'][dictBaseHead['DEFAULT_DESTINATION_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                        if eachValidation['cParentHeader'] !='':
                            oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
            elif all(key in ovald for key in {"DEFAULT_OCEAN_SCAC","DEFAULT_DESTINATION_SCAC"}):
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'ORIGIN_COLUMN' in conditinalValidation:
                        oPostParameters['aUnMappedScacColumns'][dictBaseHead['DEFAULT_ORIGIN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                        if eachValidation['cParentHeader'] !='':
                            oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
            elif all(key in ovald for key in {"DEFAULT_ORIGIN_SCAC","DEFAULT_DESTINATION_SCAC"}):
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'OCEAN_COLUMN' in conditinalValidation:
                        oPostParameters['aUnMappedScacColumns'][dictBaseHead['DEFAULT_OCEAN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                        if eachValidation['cParentHeader'] !='':
                            oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
            elif all(key in ovald for key in {"DEFAULT_DESTINATION_SCAC"}):
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'OCEAN_COLUMN' in conditinalValidation :
                        oPostParameters['aUnMappedScacColumns'][dictBaseHead['DEFAULT_OCEAN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']         
                        if eachValidation['cParentHeader'] !='':
                            oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                    elif 'ORIGIN_COLUMN' in conditinalValidation:
                        oPostParameters['aUnMappedScacColumns'][dictBaseHead['DEFAULT_ORIGIN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                        if eachValidation['cParentHeader'] !='':
                            oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
            
            elif all(key in ovald for key in {"DEFAULT_ORIGIN_SCAC"}):
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'OCEAN_COLUMN' in conditinalValidation :
                        oPostParameters['aUnMappedScacColumns'][dictBaseHead['DEFAULT_OCEAN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                        if eachValidation['cParentHeader'] !='':
                            oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                    elif 'DESTINATION_COLUMN' in conditinalValidation:
                        oPostParameters['aUnMappedScacColumns'][dictBaseHead['DEFAULT_DESTINATION_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                        if eachValidation['cParentHeader'] !='':
                            oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
            elif all(key in ovald for key in {"DEFAULT_OCEAN_SCAC"}):
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'DESTINATION_COLUMN' in conditinalValidation:
                        oPostParameters['aUnMappedScacColumns'][dictBaseHead['DEFAULT_DESTINATION_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                        if eachValidation['cParentHeader'] !='':
                            oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                    elif 'ORIGIN_COLUMN' in conditinalValidation:
                        oPostParameters['aUnMappedScacColumns'][dictBaseHead['DEFAULT_ORIGIN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                        if eachValidation['cParentHeader'] !='':
                            oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
        
       # elif all(key in oPostParameters['aScacAvailableInColumn'] for key in {"OFR","PLC"}):
        #     oPostParameters['aUnMappedScacColumns']['FOB'] = 'ORIGIN CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['ORIGIN CHARGES'] = 'ORIGIN CHARGES'
        # elif all(key in oPostParameters['aScacAvailableInColumn'] for key in {"FOB","PLC"}):
        #     oPostParameters['aUnMappedScacColumns']['OFR'] = 'OCEAN CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['OCEAN CHARGES'] = 'OCEAN CHARGES'
        # elif all(key in oPostParameters['aScacAvailableInColumn'] for key in {"OFR"}):
        #     oPostParameters['aUnMappedScacColumns']['FOB'] = 'ORIGIN CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['ORIGIN CHARGES'] = 'ORIGIN CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['PLC'] = 'DESTINATION CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['DESTINATION CHARGES'] = 'DESTINATION CHARGES'
        # elif all(key in oPostParameters['aScacAvailableInColumn'] for key in {"FOB"}):
        #     oPostParameters['aUnMappedScacColumns']['OFR'] = 'OCEAN CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['OCEAN CHARGES'] = 'OCEAN CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['PLC'] = 'DESTINATION CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['DESTINATION CHARGES'] = 'DESTINATION CHARGES'
        # elif all(key in oPostParameters['aScacAvailableInColumn'] for key in {"PLC"}):
        #     oPostParameters['aUnMappedScacColumns']['OFR'] = 'OCEAN CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['FOB'] = 'ORIGIN CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['OCEAN CHARGES'] = 'OCEAN CHARGES'
        #     oPostParameters['aUnMappedScacColumns']['ORIGIN CHARGES'] = 'ORIGIN CHARGES'
        # elif str(oPostParameters['cValidationString']).upper() =='DEFAULT_ORIGIN_SCAC':
        #     for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
        #         validations = eachValidation['cValidations']
        #         conditinalValidation = eachValidation['cValidationsCondition']
        #         if 'DESTINATION_COLUMN' in conditinalValidation:
        #             cHeaderVal = FunDCT_GetHeaderValueForColumn(oPostParameters['oSheetUploaded'], iColStart)
        #             oPostParameters['aUnMappedScacColumns'][cHeaderVal] = cHeaderVal
        #             if eachValidation['cParentHeader'] !='':
        #                 oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
        # elif str(oPostParameters['cValidationString']).upper() =='DEFAULT_DESTINATION_SCAC':
        #     for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
        #         validations = eachValidation['cValidations']
        #         conditinalValidation = eachValidation['cValidationsCondition']
        #         if 'ORIGIN_COLUMN' in conditinalValidation:
        #             cHeaderVal = FunDCT_GetHeaderValueForColumn(oPostParameters['oSheetUploaded'], iColStart)
        #             oPostParameters['aUnMappedScacColumns'][cHeaderVal] = cHeaderVal
        #             if eachValidation['cParentHeader'] !='':
        #                 oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
        # elif str(oPostParameters['cValidationString']).upper() =='DEFAULT_OCEAN_SCAC':
        #     for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
        #         validations = eachValidation['cValidations']
        #         conditinalValidation = eachValidation['cValidationsCondition']
        #         if 'OCEAN_COLUMN' in conditinalValidation:
        #             cHeaderVal = FunDCT_GetHeaderValueForColumn(oPostParameters['oSheetUploaded'], iColStart)
        #             oPostParameters['aUnMappedScacColumns'][cHeaderVal] = cHeaderVal
        #             if eachValidation['cParentHeader'] !='':
        #                 oPostParameters['aUnMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']

         



        return oPostParameters['aUnMappedScacColumns']
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_SetUnMappendColumns Error while parsing file due to ' + str(e))

def FunDCT_SetMappendColumns(oPostParameters):
    try:
        oPostParameters['aMappedScacColumns'] = defaultdict(dict)
        listofHeaders = getAllScacHeaders(oPostParameters['oValidationsDetails'])
        ovald = getSCACValiations(oPostParameters['oValidationsDetails'],oPostParameters['aScacAvailableInColumn'])
        dictBaseHead = getSCACValiationsDictByHeaderValidation(oPostParameters['oValidationsDetails'],oPostParameters['aScacAvailableInColumn'])
        dDATA = getSCAC_columnMapping(oPostParameters['oValidationsDetails'])
        if int(oPostParameters['iMaxDepthHeaders']) == 1:
            for eo in ovald:
                COLN = ''
                if 'DEFAULT_ORIGIN_SCAC' == eo:
                    COLN = 'ORIGIN_COLUMN'
                elif 'DEFAULT_DESTINATION_SCAC' == eo:
                       COLN = 'DESTINATION_COLUMN'
                elif 'DEFAULT_OCEAN_SCAC' == eo:
                       COLN = 'OCEAN_COLUMN'
                if COLN !='':
                    listofcolheaders = dDATA.get(COLN)
                    for oLi in listofcolheaders:
                        vald = FunDCT_GetValidationForColumn(oPostParameters['oValidationsDetails'],int(oLi))
                        oPostParameters['aMappedScacColumns'][vald['cHeaderLabel']] = vald['cHeaderLabel']
                    



        if int(oPostParameters['iMaxDepthHeaders']) == 2:
            if all(key in oPostParameters['aScacAvailableInColumn'] for key in listofHeaders):
                # for k in dictBaseHead:
                    # oPostParameters['aMappedScacColumns'][k] = dictBaseHead[k]['cParentHeader']
                    # oPostParameters['aMappedScacColumns'][dictBaseHead[k]['cParentHeader']] = dictBaseHead[k]['cParentHeader']    
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'OCEAN_COLUMN' in conditinalValidation :
                        if dictBaseHead.get('DEFAULT_OCEAN_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_OCEAN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                    elif 'ORIGIN_COLUMN' in conditinalValidation:
                        if dictBaseHead.get('DEFAULT_ORIGIN_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_ORIGIN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                        
                    elif 'DESTINATION_COLUMN' in conditinalValidation:
                        if dictBaseHead.get('DEFAULT_DESTINATION_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_DESTINATION_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                
                # oPostParameters['aMappedScacColumns']['OFR'] = 'OCEAN CHARGES'
                # oPostParameters['aMappedScacColumns']['OCEAN CHARGES'] = 'OCEAN CHARGES'
                # oPostParameters['aMappedScacColumns']['FOB'] = 'ORIGIN CHARGES'
                # oPostParameters['aMappedScacColumns']['ORIGIN CHARGES'] = 'ORIGIN CHARGES'
                # oPostParameters['aMappedScacColumns']['PLC'] = 'DESTINATION CHARGES'
                # oPostParameters['aMappedScacColumns']['DESTINATION CHARGES'] = 'DESTINATION CHARGES'
            elif all(key in ovald for key in {"DEFAULT_OCEAN_SCAC","DEFAULT_ORIGIN_SCAC"}):
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'OCEAN_COLUMN' in conditinalValidation :
                        if dictBaseHead.get('DEFAULT_OCEAN_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_OCEAN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                    elif 'ORIGIN_COLUMN' in conditinalValidation:
                        if dictBaseHead.get('DEFAULT_ORIGIN_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_ORIGIN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                    
                
            elif all(key in ovald for key in {"DEFAULT_DESTINATION_SCAC","DEFAULT_ORIGIN_SCAC"}):
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'ORIGIN_COLUMN' in conditinalValidation:
                        if dictBaseHead.get('DEFAULT_ORIGIN_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_ORIGIN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                    elif 'DESTINATION_COLUMN' in conditinalValidation:
                        if dictBaseHead.get('DEFAULT_DESTINATION_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_DESTINATION_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']

            elif all(key in ovald for key in {"DEFAULT_DESTINATION_SCAC","DEFAULT_OCEAN_SCAC"}):
                
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'OCEAN_COLUMN' in conditinalValidation :
                        if dictBaseHead.get('DEFAULT_OCEAN_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_OCEAN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                    
                    elif 'DESTINATION_COLUMN' in conditinalValidation:
                        if dictBaseHead.get('DEFAULT_DESTINATION_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_DESTINATION_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']

            elif all(key in ovald for key in {"DEFAULT_OCEAN_SCAC"}):
                
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']
                    if 'OCEAN_COLUMN' in conditinalValidation :
                        if dictBaseHead.get('DEFAULT_OCEAN_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_OCEAN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']
                        
                    
            elif all(key in ovald for key in {"DEFAULT_ORIGIN_SCAC"}):
                
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']

                    if 'ORIGIN_COLUMN' in conditinalValidation:
                        if dictBaseHead.get('DEFAULT_ORIGIN_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_ORIGIN_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']

            elif all(key in ovald for key in {"DEFAULT_DESTINATION_SCAC"}):
                for iColStart, eachValidation in enumerate(oPostParameters['oValidationsDetails']):
                    conditinalValidation = eachValidation['cValidationsCondition']

                    if 'DESTINATION_COLUMN' in conditinalValidation:
                        if dictBaseHead.get('DEFAULT_DESTINATION_SCAC'):
                            oPostParameters['aMappedScacColumns'][dictBaseHead['DEFAULT_DESTINATION_SCAC']['cHeaderLabel']] = eachValidation['cParentHeader']
                            if eachValidation['cParentHeader'] !='':
                                oPostParameters['aMappedScacColumns'][eachValidation['cParentHeader']] = eachValidation['cParentHeader']

            
      
        return oPostParameters['aMappedScacColumns']
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_SetMappendColumns Error while parsing file due to ' + str(e))

# e.g., A = 1, Z = 26
def FunDCT_GetColumnNumber(cLetter):
    iNum = 0
    for cLetterVal in cLetter:
        iNum = iNum * 26 + 1 + ord(cLetterVal) - ord('A')
    return int(iNum)

# def FunDCT_GetColumnIndex(oSheetUploaded, cHeaderLabel):
#     try:
#         iMaxRowUploaded = oSheetUploaded.shape[0]
#         iMaxColumnUploaded = oSheetUploaded.shape[1]
#         # iMaxColumnUploaded = oSheetUploaded.max_column

#         for iIndexRow in range(1, iMaxRowUploaded + 1):
#             for jIndexCol in range(1, iMaxColumnUploaded + 1):
#                 oMemberCellVal = oSheetUploaded.cell(
#                     row=iIndexRow, column=jIndexCol)
#                 if oMemberCellVal.value == cHeaderLabel:
#                     return jIndexCol
#     except Exception as e:
#         return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetColumnIndex Error while parsing file due to ' + str(e))

def FunDCT_GetColumnIndex(oSheetUploaded, cHeaderLabel):
    for key, row in oSheetUploaded.iterrows():
        for col_idx, item in enumerate(row):
            if item == cHeaderLabel:
                return col_idx + 1
    return 0
def FunDCT_GetMappigColumnRangesForMaxDepept1(oValidationsDetails):
    try:
        aMergedCellRanges = defaultdict(dict)
        oDataSCACMapped = getSCAC_columnMapping(oValidationsDetails)

        for colum in ['ORIGIN_COLUMN', 'DESTINATION_COLUMN', 'OCEAN_COLUMN']:
            if oDataSCACMapped.get(colum) is not None:
                listofH = oDataSCACMapped.get(colum)
                for ol in listofH:
                    vald = FunDCT_GetValidationForColumn(oValidationsDetails,int(ol))
                    aMergedCellRanges[vald['cHeaderLabel']] = {
                    'cellRange': '',
                    'iCellRow': 1,
                    'iCellCol': int(ol)+1,
                    'iTotalLen': 1,
                    'cCellRangeStart': int(ol)+1,
                    'cCellRangeEnd': int(ol) +1,
                    'cCellVal': vald['cHeaderLabel'],
                }
        return aMergedCellRanges
                
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetMappigColumnRangesForMaxDepept1 Error while parsing file due to ' + str(e))

    
def FunDCT_GetMergedCellRanges(oSheetUploaded, aMergedCellRanges):
    try:
        aMergedCellRanges = defaultdict(dict)
        for mergedCell in oSheetUploaded.merged_cells.ranges:
            cCellRangeStart, cCellRangeEnd = str(mergedCell).split(':')
            iCellRow = ''.join(filter(str.isdigit, cCellRangeStart))
            iCellCol = FunDCT_GetColumnNumber(
                ''.join([i for i in cCellRangeStart if not i.isdigit()]))
            iTotalLen = int(FunDCT_GetColumnNumber(
                ''.join([i for i in cCellRangeEnd if not i.isdigit()]))) - (int(iCellCol)-1)
            aMergedCellRanges[str(oSheetUploaded.cell(int(iCellRow), int(iCellCol)).value)] = {
                'cellRange': str(mergedCell),
                'iCellRow': int(iCellRow),
                'iCellCol': int(iCellCol),
                'iTotalLen': int(iTotalLen),
                'cCellRangeStart': cCellRangeStart,
                'cCellRangeEnd': cCellRangeEnd,
                'cCellVal': oSheetUploaded.cell(int(iCellRow), int(iCellCol)).value,
            }
        return aMergedCellRanges
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetMergedCellRanges Error while parsing file due to ' + str(e))



def FunDCT_RemoveUnmappedChargesColumn(oPostParameters):
    try:
        # oWorkBookWriteMember = openpyxl.load_workbook(
        #     oPostParameters['cMemberDistributionFilePath'])
        
        
        # oWorkSheetMember = oWorkBookWriteMember.active
        

        oPostParameters['aScacAvailableInColumn'] = defaultdict(dict)
        oPostParameters['aScacAvailableInColumn'] = FunDCT_CheckAvailableInOtherCharge(oPostParameters)
        oPostParameters['aUnMappedScacColumns'] = defaultdict(dict)
        oPostParameters['aUnMappedScacColumns'] = FunDCT_SetUnMappendColumns(oPostParameters)
        oPostParameters['aMappedScacColumns'] = defaultdict(dict)
        oPostParameters['aMappedScacColumns'] = FunDCT_SetMappendColumns(oPostParameters)
        oPostParameters['aMergedCellRanges'] = defaultdict(dict)
        oPostParameters['aMergedCellRanges'] = FunDCT_GetMergedCellRanges(oPostParameters['oSheetTemplate'], oPostParameters['aMergedCellRanges'])
        if int(oPostParameters['iMaxDepthHeaders']) == 1:
            oPostParameters['aMergedCellRanges'] = FunDCT_GetMappigColumnRangesForMaxDepept1(oPostParameters['oValidationsDetails'])
        # oWorkBookWriteMember.save(oPostParameters['cMemberDistributionFilePath'])
        # oWorkBookWriteMember.close()
        # oWorkBookWriteMember = openpyxl.load_workbook(oPostParameters['cMemberDistributionFilePath'])
        oWorkBookWriteMember = oPostParameters['oWorkBookWriteMember']
         
        
        
        
        oWorkSheetMember = oWorkBookWriteMember.active
        iMaxRow = oWorkSheetMember.max_row
        if len(oPostParameters['aUnMappedScacColumns']) > 0:
            oPostParameters['aDeleteMemberRows'] = defaultdict(dict)
            oPostParameters['aConsolidationData'] = defaultdict(dict)
            oPostParameters['aMemberCutoffs'] = defaultdict(dict)
            iSkipRows = 0
            for cKey, cDistScaccode in oPostParameters['aUnMappedScacColumns'].items():
                if all(key in oPostParameters['aMergedCellRanges'] for key in {cKey}):
                    cChargesHeaderLabel = oPostParameters['aMergedCellRanges'][cKey]['cCellVal']
                    iTotalLen = int(int(oPostParameters['aMergedCellRanges'][cChargesHeaderLabel]['iTotalLen']))
                    iActualCellCol = int(oPostParameters['aMergedCellRanges'][cChargesHeaderLabel]['iCellCol'])
                    iCellCol = int(oPostParameters['aMergedCellRanges'][cChargesHeaderLabel]['iCellCol'])
                    # if iSkipRows > 0:
                    #     iCellCol = (iCellCol - int(iSkipRows-iTotalLen))
                    iStartCol = iActualCellCol
                    if iSkipRows > 0:
                        iCellCol = (iCellCol - int(iSkipRows-iTotalLen))
                        iStartCol = (int(iActualCellCol)-int(iSkipRows))
                    iSkipRows += iTotalLen
                    oPostParameters['aDeleteMemberRows'][str(str(oPostParameters['oCellVal']).strip())][str(cChargesHeaderLabel)] = { 
                        'cChargesHeaderLabel': cChargesHeaderLabel,
                        'iActualCellCol': iActualCellCol,
                        'iCellCol': iCellCol,
                        'iTotalLen': iTotalLen,
                        'iSkipRows': iSkipRows
                    }
                    oPostParameters['aConsolidationData'][str(str(oPostParameters['oCellVal']).strip())] = { 
                        'aConsolidationData': ''
                    }
                    oPostParameters['aMemberCutoffs'][str(str(oPostParameters['oCellVal']).strip())] = { 
                        'tCuttoffdate': str(oPostParameters['tCuttoffdate']),
                    }
                    oWorkSheetMember.delete_cols(iStartCol, iTotalLen)
                    imaxrowss = oWorkSheetMember.max_row
                    iMaxCols = oWorkSheetMember.max_column


                    for cRow in oWorkSheetMember.iter_rows(min_row=int(oPostParameters['iMaxDepthHeaders']), max_row=imaxrowss,max_col=iMaxCols, values_only=False):
                        for cCel in cRow:
                            if cCel:
                                if cCel.data_type == 'f' and cCel.value.startswith('='):
                                    fval = cCel.value
                                    fgroups = re.findall(r"([A-Z]+[0-9]+)\b",fval)
                                    for eachGrp in fgroups:
                                        cFGr = re.search(r"([A-Z]+)",eachGrp)
                                        colIndexinformula = column_index_from_string(cFGr.group(1))
                                        if colIndexinformula > cCel.column:
                                            newcolIndexinformula = colIndexinformula - iTotalLen
                                            colletter= get_column_letter(newcolIndexinformula)
                                            fval = fval.replace(cFGr.group(1),colletter)
                                            cCel.value = fval


                    # oWorkBookWriteMember.save(str(oPostParameters['cMemberDistributionFilePath']))
                
                else:
                    oPostParameters['aMemberCutoffs'][str(str(oPostParameters['oCellVal']).strip())] = { 
                        'tCuttoffdate': str(oPostParameters['tCuttoffdate']),
                    }
            oPostParameters['aMappedScacColumnsDetails'] = defaultdict(dict)
            for cKey, cDistScaccode in oPostParameters['aMappedScacColumns'].items():
                if all(key in oPostParameters['aMergedCellRanges'] for key in {cKey}):
                    cChargesHeaderLabel = oPostParameters['aMergedCellRanges'][cKey]['cCellVal']
                    iTotalLen = int(int(oPostParameters['aMergedCellRanges'][cChargesHeaderLabel]['iTotalLen']))
                    iActualCellCol = int(oPostParameters['aMergedCellRanges'][cChargesHeaderLabel]['iCellCol'])
                    iCellCol = int(oPostParameters['aMergedCellRanges'][cChargesHeaderLabel]['iCellCol'])

                    if iSkipRows > 0:
                        iCellCol = (iCellCol - int(abs(iSkipRows-iTotalLen)))
                    iSkipRows += iTotalLen
                    oPostParameters['aMappedScacColumnsDetails'][str(str(oPostParameters['oCellVal']).strip())][str(cChargesHeaderLabel)] = { 
                        'cChargesHeaderLabel': cChargesHeaderLabel,
                        'iActualCellCol': iActualCellCol,
                        'iCellCol': iCellCol,
                        'iTotalLen': iTotalLen,
                        'iSkipRows': iSkipRows
                    }
                
            if oPostParameters['_cUploadType'] == 'DISTRIBUTE':
                model_common.FunDCT_UpdateMappingColumnDetails(oPostParameters['iTemplateID'], str(oPostParameters['oCellVal']).strip(), oPostParameters['aDeleteMemberRows'], "aUnMappedColumns")
                model_common.FunDCT_UpdateMappingColumnDetails(oPostParameters['iTemplateID'], str(oPostParameters['oCellVal']).strip(), oPostParameters['aMappedScacColumnsDetails'], "aMappedColumns")
                model_common.FunDCT_UpdateMappingColumnDetails(oPostParameters['iTemplateID'], str(oPostParameters['oCellVal']).strip(), oPostParameters['aConsolidationData'], "aConsolidationData")
                model_common.FunDCT_UpdateMemberCuttoff(oPostParameters['iTemplateID'], str(oPostParameters['oCellVal']).strip(), oPostParameters['aMemberCutoffs'], "aMemberCutoffs")
        if oPostParameters['_cUploadType'] == 'DISTRIBUTE' and oPostParameters['cTemplateType'] != 'RFQTool' and oPostParameters.get('aMemberCutoffs') is None:
                oPostParameters['aConsolidationData'] = defaultdict(dict)
                oPostParameters['aMemberCutoffs'] = defaultdict(dict)
                oPostParameters['aConsolidationData'][str(str(oPostParameters['oCellVal']).strip())] = {'aConsolidationData': ''}
                oPostParameters['aMemberCutoffs'][str(str(oPostParameters['oCellVal']).strip())] = {'tCuttoffdate': str(oPostParameters['tCuttoffdate'])}
                model_common.FunDCT_UpdateMappingColumnDetails(oPostParameters['iTemplateID'], str(oPostParameters['oCellVal']).strip(), oPostParameters['aConsolidationData'], "aConsolidationData")
                model_common.FunDCT_UpdateMemberCuttoff(oPostParameters['iTemplateID'], str(oPostParameters['oCellVal']).strip(), oPostParameters['aMemberCutoffs'], "aMemberCutoffs")
        # merging level 0 cells Added by Asif
        if int(oPostParameters['iMaxDepthHeaders']) > 1:
            iii = 1
            jjj = 1
            cColumnLetter1 = f"{get_column_letter(2)}1:{get_column_letter(oWorkSheetMember.max_column + 1)}1" 
            iMCOL = oWorkSheetMember.max_column 
            for eachCell in oWorkSheetMember[cColumnLetter1]:
                
                for nCel in eachCell:
                    jjj+=1
                    if nCel.value:
                        ranges = f"{get_column_letter(iii)}1:{get_column_letter(jjj-1)}1"
                        oWorkSheetMember.merge_cells(ranges)
                        # nCel.alignment =Alignment(horizontal="center", vertical="center",wrap_text=True)
                        iii=jjj
                    elif iMCOL == jjj:
                        ranges = f"{get_column_letter(iii)}1:{get_column_letter(jjj)}1"
                        oWorkSheetMember.merge_cells(ranges)
                        # nCel.alignment =Alignment(horizontal="center", vertical="center",wrap_text=True)
                        
        
        data_ValidationsConditionsE = getDataValidationAndCordinates(oValidationDetails=oPostParameters['oValidationsDetails'])
        n =1
        for ro in oWorkSheetMember.iter_rows(min_row=int(oPostParameters['iMaxDepthHeaders']), max_row=int(oPostParameters['iMaxDepthHeaders']),max_col=oWorkSheetMember.max_column, values_only=True):
            for c in ro:
                if data_ValidationsConditionsE.get(c):
                    dataString = data_ValidationsConditionsE.get(c)
                    if dataString:
                        dataValidation = DataValidation(type="list", formula1=f'"{",".join(dataString)}"', showDropDown=False, allow_blank=True,showErrorMessage=True,) # adding drop dwon in excel file
                        oWorkSheetMember.add_data_validation(dataValidation)
                        inCl = int(oPostParameters['iMaxDepthHeaders'])+1
                        cRef = f'{get_column_letter(n)}{inCl}:{get_column_letter(n)}{iMaxRow}'
                        dataValidation.add(cRef)
                n +=1

        oWorkBookWriteMember.save(str(oPostParameters['cMemberDistributionFilePath']))

    except Exception as e:
        LogService.log('Error :FunDCT_RemoveUnmappedChargesColumn Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_RemoveUnmappedChargesColumn Error while parsing file due to ' + str(e))


def getSCAC_columnMapping(oValidation):
    tDict = {'ORIGIN_COLUMN':[],'OCEAN_COLUMN':[],'DESTINATION_COLUMN':[],'DEFAULT_ORIGIN_SCAC':'','DEFAULT_DESTINATION_SCAC':'','DEFAULT_OCEAN_SCAC':'','SERVICEINFO_COLUMN':[],'SERVICEINFO_HEADER':''}
    originCol = []
    oceanCol = []
    destinationCol = []
    serviceCol = []
    unlockedIfCol = []
    for inx,oVal, in enumerate(oValidation,0):
        for eachVal in oVal['cValidations']:
            if 'SCAC' in eachVal:
                tDict[eachVal] = inx
            if 'SERVICEINFO_HEADER' in eachVal:
                tDict[eachVal] = inx
            if 'ORIGIN_COLUMN' in eachVal:
                 originCol.append(inx)
            if 'DESTINATION_COLUMN' in eachVal:
                 destinationCol.append(inx)
            if 'OCEAN_COLUMN' in eachVal:
                 oceanCol.append(inx)
            if 'SERVICEINFO_COLUMN' in eachVal:
                 serviceCol.append(inx)
            if 'UNLOCKEDIF' in eachVal:
                 unlockedIfCol.append(inx)
    tDict['ORIGIN_COLUMN'] = originCol
    tDict['OCEAN_COLUMN'] = oceanCol
    tDict['SERVICEINFO_COLUMN'] = serviceCol
    tDict['UNLOCKEDIF'] = unlockedIfCol
    tDict['DESTINATION_COLUMN'] = destinationCol
    return tDict



def FunDCT_Remove_unused_columns(wb:openpyxl):
    #   wb = openpyxl.load_workbook(file_path)
    try:
        sheet = wb.active
        # Reverse Iteration to avoid shifting on column 
        for col in reversed(range(1, sheet.max_column + 1)):
        # Check if all Empty cell
            
             
            for cell in sheet.iter_cols(min_col=col, max_col=col):
                is_empty = True
                for ec in cell:
                    if ec.value is not None:
                        is_empty = False
                        break
            if is_empty:
                sheet.delete_cols(col)
        return wb
    except Exception as e:
        LogService.log('Error :FunDCT_Remove_unused_columns Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_Remove_unused_columns Error while parsing file due to ' + str(e))
 
def FunDCT_CopyOrDeleteIrrelevantRows(oPostParameters, oWorkSheetMember):
    try:
        iMaxRowUploaded,iMaxColumnUploaded = oPostParameters['oSheetUploaded'].shape
        iMaxDepthHeaders = model_common.FunDCT_GetMaxDepthHeader(oPostParameters['iTemplateID'])

        
        oWorkBookSampleFilesCopyComment = openpyxl.load_workbook(oPostParameters['cSampleFullFilePath'])
        oWorkSheetSampoleFileCopyComment = oWorkBookSampleFilesCopyComment.active
        
        oWorkSheetMember.protection.password = loadConfig.EXL_PASS
        oWorkSheetMember.protection.sheet = True
        oWorkSheetMember.protection.autoFilter = False
        # oWorkSheetMember.protection.selectUnlockedCells = True
        # oWorkSheetMember.protection.selectLockedCells = False
        oWorkSheetMember.protection.formatCells = False
        oWorkSheetMember.auto_filter.ref = f'A{iMaxDepthHeaders}:{get_column_letter(iMaxColumnUploaded)}{iMaxDepthHeaders}'
        oWorkSheetMember.auto_filter.enable = True

        aMemberScacColIndex = []
        aMemberScacColLabel = defaultdict(dict)
        aGetMemberColumns = model_common.FunDCT_GetMemberScacValidationCol(oPostParameters['iTemplateID'])
        for cMemberColumnKey, aMemberColumnDetails in enumerate(aGetMemberColumns):
            # if aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel'] == 'SERVICE INFO':
            if aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel'] == 'SERVICE INFO':
                continue
            # iColumnIndex = FunDCT_GetColumnIndex(oPostParameters['oSheetUploaded'], aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel'])
            iColumnIndex = list(oPostParameters['oSheetUploaded'].loc[iMaxDepthHeaders-1].values).index(aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel'])
            aMemberScacColIndex.append(iColumnIndex)
            aMemberScacColLabel[iColumnIndex]['cHeaderLabel'] = str(aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel']).strip()
            oPostParameters['aCountMappedColumn'][str(oPostParameters['oCellVal']).strip()][aMemberScacColLabel[iColumnIndex]['cHeaderLabel']] = {
                'iTotalLanes': 0,
                'iSubmittedLanes': 0,
                'iRemainingLanes': 0,
            }
        iMaxRow = iMaxDepthHeaders + 1
        mappingSCAC =  getSCAC_columnMapping(oPostParameters['oValidationsDetails'])
        dataValidationDict = dict() 
        sheetname = oWorkSheetMember.title
        sheetCo_RelatedDataColWise = oPostParameters['sheetCo_RelatedData']['1']
        for iIndexRow in range(1, iMaxRowUploaded + 1):
        # for iIndexRow, row in enumerate(oPostParameters['oSheetUploaded'], 1):
            bCheck = False
            listofmemberscaccolumn = [str(oPostParameters['oSheetUploaded'].iat[iIndexRow-1,x]).strip() for x  in aMemberScacColIndex ]
            for jIndexCol in range(1, iMaxColumnUploaded + 1):
            # for jIndexCol,oMemberCellVal in enumerate(row,1 ):
                oDetails = oPostParameters['oValidationsDetails'][jIndexCol-1]
                conditinalValidation = oDetails['cValidationsCondition']
                validations = oDetails['cValidations']
                
                oMemberCellVal = str(oPostParameters['oSheetUploaded'].iat[iIndexRow-1, jIndexCol-1])
                # vAligmentCell = oMemberCellVal.alignment.vertical
                vAligmentCell = 'center'
                # hAligmentCell = oMemberCellVal.alignment.horizontal
                hAligmentCell = 'left'
                # bwrap = True if oMemberCellVal.alignment.wrapText else False
                bwrap = True
                if iIndexRow <= iMaxDepthHeaders:
                    oWorkSheetMember.cell(row=iIndexRow, column=jIndexCol).value = oMemberCellVal
                    cCommentStr = oWorkSheetSampoleFileCopyComment.cell(row=iIndexRow, column=jIndexCol).comment
                    # oDetailsHeader = headerValidaton[jIndexCol-1]
                    # cTextColor = oDetailsHeader['cTextColorRow']['hex']
                    # cTextBGColor = oDetailsHeader['cBackColorRow']['hex']
                    
                    
                    cTextColor = FunDCT_GetTextColor(oWorkSheetSampoleFileCopyComment.cell(row=iIndexRow, column=jIndexCol))
                    cTextBGColor = FunDCT_GetBackColor(oWorkSheetSampoleFileCopyComment.cell(row=iIndexRow, column=jIndexCol),cTextColor)
                    oWorkSheetMember.cell(row=iIndexRow, column=jIndexCol).font = Font(color=cTextColor)
                    oWorkSheetMember.cell(row=iIndexRow, column=jIndexCol).fill = PatternFill(fgColor=cTextBGColor,fill_type = "solid")
                    thin_border = Side(style="thin")
                    oWorkSheetMember.cell(row=iIndexRow, column=jIndexCol).border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                    oWorkSheetMember.cell(row=iIndexRow, column=jIndexCol).alignment = Alignment(horizontal=hAligmentCell, vertical=vAligmentCell, wrap_text=bwrap)
                    if str(cCommentStr).strip() != '' and cCommentStr is not None:
                            oWorkSheetMember.cell(row=iIndexRow, column=jIndexCol).comment = cCommentStr
                        
                    oWorkSheetMember.cell(row=iIndexRow, column=jIndexCol).value = oMemberCellVal
                if str(oPostParameters['oCellVal']).strip() in listofmemberscaccolumn:
                    bCheck = True
                    rowTextColor = oDetails['cTextColorRow']['hex']
                    rowBGColor = oDetails['cBackColorRow']['hex']
                    # if str(oMemberCellVal).startswith('='):
                    if sheetCo_RelatedDataColWise.get(str(jIndexCol-1)).get('formula'):
                        # oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).value = oMemberCellVal
                        formula = sheetCo_RelatedDataColWise.get(str(jIndexCol-1)).get('formula')
                        matches = re.findall(r"([A-Z]+[0-9]+)\b",formula)
                        for mtch in matches:
                            rownumber = re.search(r'([0-9]+)\b',str(mtch)).group(1)
                            newR = mtch.replace(rownumber,str(iMaxRow))
                            formula = formula.replace(mtch,newR)
                        oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).value = formula
                        oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).font = Font(color=rowTextColor)
                        oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).fill = PatternFill(fgColor=rowBGColor,fill_type = "solid")
                        oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).alignment = Alignment(horizontal=hAligmentCell, vertical=vAligmentCell)      
                    else:
                        formate = sheetCo_RelatedDataColWise.get(str(jIndexCol-1)).get('number_format')
                        if '%' in str(formate):
                            try:
                                oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).value = f"{round(float(oMemberCellVal) * 100, 2)}%"
                            except:
                                oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).value = oMemberCellVal
                                
                        else:
                            oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).value = oMemberCellVal
                        
                        oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).font = Font(color=rowTextColor)
                        oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).fill = PatternFill(fgColor=rowBGColor,fill_type = "solid")
                        oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).alignment = Alignment(horizontal=hAligmentCell, vertical=vAligmentCell)
                        sSCACname = ''
                        if 'ORIGIN_COLUMN' in conditinalValidation:
                            sSCACname = 'DEFAULT_ORIGIN_SCAC' 
                        elif 'OCEAN_COLUMN' in conditinalValidation:
                            sSCACname = 'DEFAULT_OCEAN_SCAC' 
                        elif 'DESTINATION_COLUMN' in conditinalValidation:
                            sSCACname = 'DEFAULT_DESTINATION_SCAC' 
                        elif 'SERVICEINFO_COLUMN' in conditinalValidation:
                            sSCACname ='SERVICEINFO_HEADER'
                        if mappingSCAC.get(sSCACname) is not None:
                            try:
                                if mappingSCAC.get(sSCACname):
                                    scacint = int(mappingSCAC.get(sSCACname))
                                    cValueCatch = oPostParameters['oSheetUploaded'].iat[iIndexRow-1, scacint]
                                    if cValueCatch:
                                        if str(oPostParameters['oCellVal']).strip() == str(cValueCatch).strip():
                                            oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).protection = Protection(locked=False)
                            except:
                                pass
                        # bVCase =  any([True if 'DATA_VALIDATION' in x else False for x in conditinalValidation])
                        # if  bVCase:
                        #     print('')
                        if 'UNLOCKEDIF' or 'UNLOCKED' in validations:
                            for eachVal in conditinalValidation:
                                # if 'DATA_VALIDATION' in eachVal: 
                                #     conditionalString = eachVal.replace('(','').replace(')','').replace('DATA_VALIDATION','')
                                #     listofdataValidation =conditionalString.split(',')
                                #     if not dataValidationDict.get(jIndexCol):
                                #         dataValidation = DataValidation(type="list", formula1=f'"{",".join(listofdataValidation)}"', showDropDown=False, allow_blank=True) # adding drop dwon in excel file
                                #         oWorkSheetMember.add_data_validation(dataValidation)
                                #         dataValidationDict[jIndexCol]= dataValidation
                                #     if dataValidationDict.get(jIndexCol):
                                #         dataValidation = dataValidationDict.get(jIndexCol)
                                #         dataValidation.add(oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).coordinate)
                            
                                if 'UNLOCKED' in eachVal:
                                    oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).protection = Protection(locked=False)

                                

                                if 'UNLOCKEDIF' in eachVal:
                                    conditionalString = eachVal.replace('(','').replace(')','').replace('UNLOCKEDIF','')
                                    # iValidationsDetails = oPostParameters['oValidationsDetails']
                                    # matchingIndex = next(
                                    #     (index for index, item in enumerate(iValidationsDetails) if item.get('cHeaderLabel') == conditionalString), 
                                    #     None
                                    # )
                                    matchingIndex = pydash.find_index(oPostParameters['oValidationsDetails'],lambda x : x.get('cHeaderLabel') == conditionalString)
                                    if matchingIndex is not None:
                                        cColumnLetterofRelatedColmn = get_column_letter(int(matchingIndex + 1))  # +1 if using 1-based indexing

                                        oRCellVal = str(oPostParameters['oSheetUploaded'].iat[iIndexRow-1, int(matchingIndex)])
                                        if oRCellVal in str(oPostParameters['cMemberDistributionFilePath']) :
                                            oWorkSheetMember.cell(row=iMaxRow, column=jIndexCol).protection = Protection(locked=False)
            
            #     if (len(list(filter (lambda x : x == jIndexCol, aMemberScacColIndex))) > 0) :
            #         if bCheck == True and str(oPostParameters['oCellVal']).strip() == str(oMemberCellVal.value).strip():
            #             bCheck = False
            # if bCheck == True and iMaxDepthHeaders < iIndexRow:
            #     oWorkSheetMember.delete_rows(iIndexRow,1)
            if bCheck == True:
                iMaxRow  +=1
        return oPostParameters['aCountMappedColumn'],oWorkSheetMember
    except Exception as e:
        LogService.log('Error :FunDCT_CopyOrDeleteIrrelevantRows Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CopyOrDeleteIrrelevantRows Error while parsing file due to ' + str(e))
def FunDCT_DeleteEmptyRowsAndShiftUp(oWorkSheetMember):
    try:
        aRowDetails = []
        for iIndex in range(1, oWorkSheetMember.max_row):
            if oWorkSheetMember.cell(iIndex, 1).value is None:
                aRowDetails.append(iIndex)
        for iRowDel in range(len(aRowDetails)):
            oWorkSheetMember.delete_rows(idx=aRowDetails[iRowDel], amount=1)
            # aRowDetails = list(map(lambda cKey: cKey - 1, aRowDetails))
            aRowDetails = [ckey -1 for ckey in aRowDetails]
            
    except Exception as e:
        LogService.log('Error :FunDCT_DeleteEmptyRowsAndShiftUp Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DeleteEmptyRowsAndShiftUp Error while parsing file due to ' + str(e))

def FunDCT_TrimAllColumnsDataFrame(oDataFrame):
    try:
        cStrings = lambda cCellString: cCellString.strip() if isinstance(cCellString, str) else cCellString
        return oDataFrame.applymap(cStrings)
    except Exception as e:
        LogService.log('Error :FunDCT_TrimAllColumnsDataFrame Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_TrimAllColumnsDataFrame Error while parsing file due to ' + str(e))

def FunDCT_PreparedUploadedStat(oPostParameters):
    try:
        oDataFrameAsRecords= dict()   
        for cKey, cDistScaccode in oPostParameters['aCountMappedColumn'].items():
            for cColKey, cColVal in cDistScaccode.items():
                # with open('temp.txt','w') as f:
                #     f.writelines(str(oPostParameters['cMemberDistributionFilePath']))
                #     f.close()
                
                if not oDataFrameAsRecords:
                    ndf = pandas.ExcelFile(oPostParameters['oWorkBookWriteMember'],'openpyxl')
                    oDataFrame = ndf.parse(na_values = "Mssing", na_filter=True, skiprows = int(int(oPostParameters['iMaxDepthHeaders']) - 1))
                    oDataFrame = oDataFrame.applymap(str)
                # oDataFrame = pandas.read_excel(str(oPostParameters['cMemberDistributionFilePath']), na_values = "Mssing", na_filter=True, skiprows = int(int(oPostParameters['iMaxDepthHeaders']) - 1),engine='openpyxl')
                # oDataFrame = FunDCT_TrimAllColumnsDataFrame(oDataFrame) commenting this as on start we triming all data         
                    oDataFrameAsRecords = oDataFrame.to_dict('records')    
                iTotalLanesCount = sum(x[cColKey].strip() == str(oPostParameters['oCellVal']).strip() for x in oDataFrameAsRecords)
                oPostParameters['aCountMappedColumn'][str(oPostParameters['oCellVal']).strip()][cColKey] = {
                    'iTotalLanes': iTotalLanesCount,
                    'iSubmittedLanes': 0,
                    'iRemainingLanes': 0,
                }
        model_common.FunDCT_UpdateMappingColumnDetails(oPostParameters['iTemplateID'], str(oPostParameters['oCellVal']).strip(), oPostParameters['aCountMappedColumn'], "aMappedScacColumnsStat")
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_PreparedUploadedStat Error while parsing file due to ' + str(e) + '----' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_PreparedUploadedStat Error while parsing file due to ' + str(e) + '----' + str(line))

def FunDCT_CreateAndCopyHeader(oPostParameters):
    try:
        
        # tdf = oPostParameters['oSheetUploaded']
        # tdf = tdf[(tdf[oPostParameters['lScac']] == oPostParameters['_cCompanyname']).any(axis=1)]
        # tdf.to_excel(oPostParameters['cMemberDistributionFilePath'],na_rep='',header=False,index=False,index_label=None,startrow=int(oPostParameters['iMaxDepthHeaders'])+1)
        oWorkBookWriteMember = openpyxl.load_workbook(
            oPostParameters['cMemberDistributionFilePath'])
        oWorkSheetMember = oWorkBookWriteMember.active
        
        # Column width set
        for col in range(1, 101):  # 1 to 50 (A to AX)`
            col_letter = openpyxl.utils.get_column_letter(col)
            oWorkSheetMember.column_dimensions[col_letter].width = 25
        
        oPostParameters['aCountMappedColumn'] = defaultdict(dict)
        oPostParameters['aCountMappedColumn'],oWorkSheetMember = FunDCT_CopyOrDeleteIrrelevantRows(oPostParameters, oWorkSheetMember)
        # FunDCT_DeleteEmptyRowsAndShiftUp(oWorkSheetMember)
        
      
        oPostParameters['oWorkBookWriteMember'] = oWorkBookWriteMember
        oWorkBookWriteMember.save(str(oPostParameters['cMemberDistributionFilePath']))
        

        FunDCT_PreparedUploadedStat(oPostParameters)

        return FunDCT_RemoveUnmappedChargesColumn(oPostParameters)
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_CreateAndCopyHeader Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CreateAndCopyHeader Error while parsing file due to ' + str(e))

def FunDCT_Default_Scac(oPostParameters):
    try:
        bValidate = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_MEMBERSCAC')
        # oValidMemberScac = model_common.FunDCT_GetMember()
        # for cMemberKey, aMemberDetails in enumerate(oValidMemberScac):
        #     if aMemberDetails['cScac'] == str(oPostParameters['oCellVal']).strip():
        #         bValidate = True
        aGetSelectedMembers = model_common.FunDCT_GetMembersForDistribution(oPostParameters['_iTemplateUploadLogID'], oPostParameters['_cUploadType'])
        if model_common.FunDCT_CheckMemberbySCAC(str(oPostParameters['oCellVal']).strip()):
            bValidate = True
        bCheckMember = True
        if  oPostParameters['cValidationString'].startswith('SERVICEINFO_HEADER'):
            bCheckMember = False
        if oPostParameters['_cUploadType'] != 'DISTRIBUTE':
            bCheckMember = False
        if len(aGetSelectedMembers) > 0 :
            if  oPostParameters['oCellVal'] not in aGetSelectedMembers:
                bCheckMember = False
        
        if str(oPostParameters['oCellVal']).strip() in oPostParameters['aRestrictedScacCode']:
            return 'Restricted scac code - '

        # Create Member Wise File For Template Distribution
        if bValidate == True and bCheckMember == True:
            if FunDCT_ScacExist(aDistributionMembers, str(oPostParameters['oCellVal']).strip()) == False:
                oPostParameters['cMemberDistributionFilePath'] = (
                    oPostParameters['cDirValidateTemplateStatusFile'] +
                    str(oPostParameters['oCellVal']).strip() + '_' +
                    str(oPostParameters['iTemplateID']) + '_' +
                    str(oNow.strftime("%Y-%m-%d-%H-%M-%S")) +
                    '.xlsx'
                )
                copyfile(oPostParameters['cDirValidateTemplateStatusFile'] + oPostParameters['_cValidateTmplateStatusFile'], oPostParameters['cMemberDistributionFilePath'])
                
                FunDCT_CreateAndCopyHeader(oPostParameters)
                Path(oPostParameters['cDirDistribution']+str(oPostParameters['iTemplateID'])).mkdir(exist_ok=True)
                cDistributeMemberFile =  str(oPostParameters['cDirDistribution']) + str(oPostParameters['iTemplateID']) + '/' + str(oPostParameters['oCellVal']).strip() + '_' + str(oPostParameters['iTemplateID']) + '.xlsx'
                os.remove(cDistributeMemberFile) if os.path.exists(cDistributeMemberFile) else None
                copyfile(oPostParameters['cMemberDistributionFilePath'], cDistributeMemberFile)
                aDistributionMembers[len(aDistributionMembers)+1] = {
                    'cDistScaccode': str(oPostParameters['oCellVal']).strip(),
                    'cMemberDistributionFilePath': cDistributeMemberFile,
                    'cMemberEmail': model_common.FunDCT_GetEmailFromCompanyName(str(oPostParameters['oCellVal']).strip()),
                }
        return bValidate
    except Exception as e:
        LogService.log('Error :FunDCT_Default_Scac Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_Default_Scac Error while parsing file due to ' + str(e))

def FunDCT_Default_Currency(cCellVal):
    try:
        bValidate = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_CURRENCY')
        # oValidCurrency = model_common.FunDCT_GetCurrency()
        # for i, d in enumerate(oValidCurrency):
        #     if d['cCode'] == cCellVal:
        #         bValidate = True
        if not cCellVal:
            return True
        if model_common.FunDCT_CheckCurrency(cCellVal):
            bValidate = True


        return bValidate
    except Exception as e:
        LogService.log('Error :FunDCT_Default_Currency Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_Default_Currency Error while parsing file due to ' + str(e))

def FunDCT_Default_Basis(cCellVal):
    try:
        bValidate = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_BASIS')
        # oValidBasis = model_common.FunDCT_GetBasis()
        # for i, d in enumerate(oValidBasis):
        #     if d['cCode'] == cCellVal:
        #         bValidate = True
        if not cCellVal:
            return True
        if model_common.FunDCT_CheckBasis(cCellVal):
            bValidate = True

        return bValidate
    except Exception as e:
        LogService.log('Error :FunDCT_Default_Basis Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_Default_Basis Error while parsing file due to ' + str(e))

def FunDCT_Default_ChargeCode(cCellVal):
    try:
        bValidate = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_CHARGECODE')
        # oValidChargeCode = model_common.FunDCT_GetChargecode()
        # for i, d in enumerate(oValidChargeCode):
        #     if d['cCode'] == cCellVal:
        #         bValidate = True
        if model_common.FunDCT_CheckChargecode(cCellVal):
            bValidate = True


        return bValidate
    except Exception as e:
        LogService.log('Error :FunDCT_Default_ChargeCode Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_Default_ChargeCode Error while parsing file due to ' + str(e))

###############################
# MIXED DATA VALIDATIONS FUNCTIONS
###############################

def FunDCT_MixedData(oPostParameters):
    try:
        bValidate = MessageHandling.FunDCT_GetValidationDescription(
            'UPLOAD_TEMPLATE', 'INVALID_MIXEDDATA')
        bMatchRange = False
        bMatchAllowedText = False
        bValidateDefault = False
        
        if loadConfig.PY_BRACKET_VALIDATION_SEPARATOR in oPostParameters['cValidationString']:
            aPrimaryValidators = oPostParameters['cValidationString'].split(
                loadConfig.PY_BRACKET_VALIDATION_SEPARATOR)
            for cValidationString in aPrimaryValidators:
                if cValidationString.startswith('MIXED_DATA(DEFAULT_'):
                    oPostParameters['cValidationString'] = oPostParameters['cValidationString'].replace(
                        ')', '').split('MIXED_DATA(')[1]
                    bValidateDefault = FunDCT_DefaultValidations(oPostParameters)
                    continue
                elif 'MIXED_DATA(' in cValidationString:
                    oPostParameters['cValidationString'] = cValidationString.replace(
                        ')', '').split('MIXED_DATA(')[1]
                else:
                    oPostParameters['cValidationString'] = cValidationString.replace(')', '')
                if loadConfig.PY_RANGE_VALUE_SEPARATOR in oPostParameters['cValidationString']:
                    aSplittedVal = oPostParameters['cValidationString'].split(
                        loadConfig.PY_RANGE_VALUE_SEPARATOR)

                    iCount = len(aSplittedVal)
                    bIsDigit = all(map(str.isdigit, aSplittedVal))
                    bIsAlpha = all(map(str.isalpha, aSplittedVal))
                  ## 'added str(oPostParameters['oCellVal']).isdigit' in below condition for validation issue               
                    if bIsDigit and iCount == 2 and str(oPostParameters['oCellVal']).isdigit:
                        [iRangeFrom, iRangeTo] = [int(iRanges) for iRanges in oPostParameters['cValidationString'].split(
                            loadConfig.PY_RANGE_VALUE_SEPARATOR)]
                        if (isinstance(oPostParameters['oCellVal'], int) or isinstance(oPostParameters['oCellVal'], float)) and (iRangeFrom == 0 and iRangeTo == 0):
                            bMatchRange = True
                        elif isinstance(oPostParameters['oCellVal'], int):
                            bMatchRange = FunDCT_NumericIntRange(
                                oPostParameters['oCellVal'], iRangeFrom, iRangeTo)
                        elif isinstance(oPostParameters['oCellVal'], float):
                            bMatchRange = FunDCT_NumericFloatRange(
                                oPostParameters['oCellVal'], iRangeFrom, iRangeTo)
                    else:
                        bMatchAllowedText = FunDCT_AllowedTextOnly(
                            oPostParameters['oCellVal'], aSplittedVal)
                else:
                    #added 'if(bMatchAllowedText == True):break' below for validation issue 
                    bMatchAllowedText = FunDCT_AllowedTextOnly(
                        oPostParameters['oCellVal'], oPostParameters['cValidationString'])
                    if(bMatchAllowedText == True):
                        break
            
            # below added == True by spirognde for validation issue    
            if bMatchRange ==True or bMatchAllowedText==True or bValidateDefault==True:
                bValidate = True
        else:
            if loadConfig.PY_RANGE_VALUE_SEPARATOR in oPostParameters['cValidationString']:
                if 'MIXED_DATA(' in oPostParameters['cValidationString']:
                    oPostParameters['cValidationString'] = oPostParameters['cValidationString'].replace(
                        ')', '').split('MIXED_DATA(')[1]
                if loadConfig.PY_RANGE_VALUE_SEPARATOR in oPostParameters['cValidationString']:
                    aSplittedVal = oPostParameters['cValidationString'].split(
                        loadConfig.PY_RANGE_VALUE_SEPARATOR)

                iCount = len(aSplittedVal)
                bIsDigit = all(map(str.isdigit, aSplittedVal))
                bIsAlpha = all(map(str.isalpha, aSplittedVal))
                if bIsDigit and iCount == 2:
                    [iRangeFrom, iRangeTo] = [int(iRangeFrom) for iRangeFrom in oPostParameters['cValidationString'].split(
                        loadConfig.PY_RANGE_VALUE_SEPARATOR)]
                    if (isinstance(oPostParameters['oCellVal'], int) or isinstance(oPostParameters['oCellVal'], float)) and (iRangeFrom == 0 and iRangeTo == 0):
                        bMatchRange = True
                    elif isinstance(oPostParameters['oCellVal'], int):
                        bMatchRange = FunDCT_NumericIntRange(
                            oPostParameters['oCellVal'], iRangeFrom, iRangeTo)
                    elif isinstance(oPostParameters['oCellVal'], float):
                        bMatchRange = FunDCT_NumericFloatRange(
                            oPostParameters['oCellVal'], iRangeFrom, iRangeTo)
                elif bIsAlpha:
                    bMatchAllowedText = FunDCT_AllowedTextOnly(
                        oPostParameters['oCellVal'], aSplittedVal)
            elif oPostParameters['cValidationString'].startswith('MIXED_DATA(DEFAULT_'):
                oPostParameters['cValidationString'] = oPostParameters['cValidationString'].replace(
                    ')', '').split('MIXED_DATA(')[1]
                bValidate = FunDCT_DefaultValidations(oPostParameters)
            
            # below added == True by spirognde  for validation issue      
            if bMatchRange==True or bMatchAllowedText == True:
                bValidate = True
        
        return bValidate
    except Exception as e:
        LogService.log('Error :FunDCT_MixedData Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_MixedData Error while parsing file due to ' + str(e))



# #########################################
    ######### NOT ACCEPTED COLUMN WITH SPECIFIED COLUMNS
    # Written by Asif
# #############################################
    
def FunDCT_NOT_ACCEPTEDIF(oPostParameters):
    try:
        bValidate = False
    
        if 'NOT_ACCEPTED' in oPostParameters['cValidationString']:
            primeValidation = oPostParameters['cValidationString']
            primeValidation = primeValidation.replace('NOT_ACCEPTEDIF(', '')
            primeValidation = primeValidation.replace(')','')
            columnNames = primeValidation.split(loadConfig.PY_RANGE_VALUE_SEPARATOR)
            for eachCol in columnNames:
                colName = get_column_letter(int(eachCol))
                cGetDependentColumnVal = FunDCT_GetDependentColumnValue(oPostParameters['oSheetUploaded'], colName, oPostParameters['iRow'])
                bValidate = FunDCT_NOT_ACCEPTED(cGetDependentColumnVal,oPostParameters['oCellVal'],colName)
        return bValidate
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_NotAcceptedIF Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_NotAcceptedIF Error while parsing file due to ' + str(e))


def FunDCT_NOT_ACCEPTED(cGetDependentColumnVal, oCellVal,colName):
    try:
        bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'NOT_ACCEPTEDIF')
        bMatch =  MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE', 'NOT_ACCEPTED').replace(
                    '#$DCTVARIABLE_INCORRECT_CELL_VAL#$', str(cGetDependentColumnVal)).replace('$#colRef$#',str(colName))
        if type(cGetDependentColumnVal) == int or type(oCellVal) == int and type(cGetDependentColumnVal) == float or type(oCellVal) == float:
            if cGetDependentColumnVal > 0 and oCellVal == 0 :
                bMatch = True
            elif cGetDependentColumnVal == 0 and oCellVal > 0 :
                bMatch = True
            
        return bMatch
    except Exception as e:
        LogService.log('Error :FunDCT_NotAccepted Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_NotAccepted Error while parsing file due to ' + str(e))









###############################
# DEPENDENT DATA VALIDATIONS FUNCTIONS
###############################

def FunDCT_GetDependentColumnValue(oSheetUploaded, cCoordinate, jColumn):
    try:
        oCell = oSheetUploaded.iat[int(cCoordinate),int(jColumn)]
        return str(oCell)
    except Exception as e:
        LogService.log('Error :FunDCT_GetDependentColumnValue Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetDependentColumnValue Error while parsing file due to ' + str(e))

def FunDCT_DependentData(oPostParameters):
    try:
        bValidate = False
        bMatchAllowedText = False
        if loadConfig.PY_BRACKET_VALIDATION_SEPARATOR in oPostParameters['cValidationString']:
            aPrimaryValidators = oPostParameters['cValidationString'].split(
                loadConfig.PY_BRACKET_VALIDATION_SEPARATOR)
            iIndexDependentValidator = 1

            for cValidationString in aPrimaryValidators:
                if iIndexDependentValidator <= 1:
                    aExplRange = aPrimaryValidators[0].split('DEPENDENT(')[1]
                    cGetDependentColumnVal = FunDCT_GetDependentColumnValue(
                        oPostParameters['oSheetUploaded'],  oPostParameters['iRow'],aExplRange)

                    if len(aPrimaryValidators) >= 3:
                        if aPrimaryValidators[1].startswith('MIXED_DATA(DEFAULT_'):
                            aSplittedVal = aPrimaryValidators[2].replace('))', '').split(
                                loadConfig.PY_RANGE_VALUE_SEPARATOR)
                            bMatchAllowedText = FunDCT_AllowedTextOnly(
                                cGetDependentColumnVal, aSplittedVal)
                            if bMatchAllowedText == True:
                                bValidate = FunDCT_AllowedTextOnly(
                                    oPostParameters['oCellVal'], aSplittedVal)
                                break
                            else:
                                if aPrimaryValidators[1].startswith('MIXED_DATA(DEFAULT_'):
                                    oPostParameters['cValidationString'] = aPrimaryValidators[1].split(
                                        'MIXED_DATA(')[1]
                                    bValidate = FunDCT_DefaultValidations(oPostParameters)
                                    break
                        
                        elif aPrimaryValidators[1].startswith('TEXT_RANGE(') or aPrimaryValidators[1].startswith('NUMERIC_RANGE('):
                            aSplittedVal = aPrimaryValidators[2].replace(')', '').split(
                                loadConfig.PY_RANGE_VALUE_SEPARATOR)
                            bMatchAllowedText = FunDCT_AllowedTextOnly(
                                cGetDependentColumnVal, aSplittedVal)
                            if bMatchAllowedText == True:
                                bValidate = FunDCT_AllowedTextOnly(
                                    oPostParameters['oCellVal'], aSplittedVal)
                                break
                            else:
                                oPostParameters['cValidationString'] = aPrimaryValidators[1]
                                bValidate = FunDCT_RangeValidations(oPostParameters)
                        
                        elif aPrimaryValidators[1].startswith('DEFAULT_'):
                            aSplittedVal = aPrimaryValidators[2].replace(')', '').split(
                                loadConfig.PY_RANGE_VALUE_SEPARATOR)
                            bMatchAllowedText = FunDCT_AllowedTextOnly(
                                cGetDependentColumnVal, aSplittedVal)
                            if bMatchAllowedText == True:
                                bValidate = FunDCT_AllowedTextOnly(
                                    oPostParameters['oCellVal'], aSplittedVal)
                                break
                            else:
                                oPostParameters['cValidationString'] = aPrimaryValidators[1]
                                bValidate = FunDCT_DefaultValidations(oPostParameters)

                        elif aPrimaryValidators[1].startswith('MIXED_DATA('):
                            aSplittedVal = aPrimaryValidators[2].replace('))', '').split(
                                loadConfig.PY_RANGE_VALUE_SEPARATOR)
                            bMatchAllowedText = FunDCT_AllowedTextOnly(
                                cGetDependentColumnVal, aSplittedVal)
                            if bMatchAllowedText == True:
                                bValidate = FunDCT_AllowedTextOnly(
                                    oPostParameters['oCellVal'], aSplittedVal)
                                break
                            else:
                                aExplRange = aPrimaryValidators[1].split(
                                    'MIXED_DATA(')[1]

                                if loadConfig.PY_RANGE_VALUE_SEPARATOR in aExplRange:
                                    aSplittedVal = aExplRange.split(
                                        loadConfig.PY_RANGE_VALUE_SEPARATOR)

                                    iCount = len(aSplittedVal)
                                    bIsDigit = all(
                                        map(str.isdigit, aSplittedVal))
                                    bIsAlpha = all(
                                        map(str.isalpha, aSplittedVal))

                                    if bIsDigit and iCount == 2:
                                        [iRangeFrom, iRangeTo] = [int(iRanges) for iRanges in aExplRange.split(
                                            loadConfig.PY_RANGE_VALUE_SEPARATOR)]
                                        
                                        if (isinstance(oPostParameters['oCellVal'], int) or isinstance(oPostParameters['oCellVal'], float)) and (iRangeFrom == 0 and iRangeTo == 0):
                                            bValidate = True
                                        elif isinstance(oPostParameters['oCellVal'], int):
                                            bValidate = FunDCT_NumericIntRange(
                                                oPostParameters['oCellVal'], iRangeFrom, iRangeTo)
                                        elif isinstance(oPostParameters['oCellVal'], float):
                                            bValidate = FunDCT_NumericFloatRange(
                                                oPostParameters['oCellVal'], iRangeFrom, iRangeTo)
                                        break
        return bValidate
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_DependentData Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DependentData Error while parsing file due to ' + str(e))

###############################
# NO CHANGE FUNCTIONS
###############################

def FunDCT_NoChangeValidations(oPostParameters, aSavedSearchedRow, oSheetTemplateUniqueSeq):
    try:
        if oPostParameters['_cUploadType'] != 'DISTRIBUTE':
            bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'NO_CHANGE')
            if oPostParameters['cValidationString'] == 'NO_CHANGE':
                bMatch = FunDCT_NoChange(oPostParameters, aSavedSearchedRow, oSheetTemplateUniqueSeq)
        else:
            bMatch = True
        return bMatch
    except Exception as e:
        LogService.log('Error :FunDCT_NoChangeValidations Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_NoChangeValidations Error while parsing file due to ' + str(e))

def FunDCT_NoChange(oPostParameters, aSavedSearchedRow, oSheetTemplateUniqueSeq):
    try:
       ## code for checking if cKeyUniqueID is None or it is present in aSavedSearchedRow added by spirgonde
        bMatch = True
        cKeyUniqueID = str(oPostParameters['oSheetUploaded'].iat[int(oPostParameters['iRow'])-1, int(oPostParameters['cKeyUniqueIDSequence'])-1])
        # print(f'{MemoryService.getMemoryUsedStr()}- No change validaiton fetibg uploaded sheet')
        if(cKeyUniqueID is None):
            bMatch = False
            return bMatch 
        else:
            if type(cKeyUniqueID) == str : cKeyUniqueID = int(cKeyUniqueID)
            if cKeyUniqueID in aSavedSearchedRow:
                iFindRow = int(aSavedSearchedRow[cKeyUniqueID])
                if iFindRow != '':
                    # jColumnPlusone = oPostParameters['iForwardValidator']+1
                    jColumnPlusone = oPostParameters['iForwardValidator']
                    cCellValOrg = oSheetTemplateUniqueSeq.get(int(iFindRow))[int(jColumnPlusone)]
                    # cCellValOrg = oSheetTemplateUniqueSeq.iat[int(iFindRow), int(jColumnPlusone)]
                    # cCellValOrg = oSheetTemplateUniqueSeq.cell(row=int(iFindRow), column=int(jColumnPlusone))
                    
                    if str(oPostParameters['oCellVal']).strip() != str(cCellValOrg).strip():
                        if cCellValOrg is None and str(oPostParameters['oCellVal']).strip().upper() == 'NA':
                            return True
                        bMatch = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE', 'NO_CHANGE')
                        bMatch = str(bMatch).replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$',str(cCellValOrg).strip())
                        bMatch = bMatch +  f"{str(oPostParameters['oCellVal']).strip()}"
                        return bMatch


            else :
                bMatch = False
                return bMatch        
        return bMatch
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_NoChange Error while parsing file due to ' + str(e) + ' On Line no = '+ str(line)+'cKeyUniqueID :'+str(cKeyUniqueID))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_NoChange Error while parsing file due to ' + str(e) + ' On Line no = '+ str(line)+'cKeyUniqueID :'+str(cKeyUniqueID))

def FunDCT_FindAndGetRow(oSheetTemplate, cSearchUniqueID, iColToSearch=1):
    try:
        for row in range(1, oSheetTemplate.max_row + 1):
            if oSheetTemplate[row][iColToSearch].value is not None:
                if str(oSheetTemplate[row][iColToSearch].value) == str(cSearchUniqueID):
                    return iColToSearch, row
        return iColToSearch, None
    except Exception as e:
        LogService.log('Error :FunDCT_FindAndGetRowToUpdate Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_FindAndGetRowToUpdate Error while parsing file due to ' + str(e))

def FunDCT_GetUniqueIDSequence(oSheetTemplate, cTemplateType):
    try:
        cKeyUniqueIDSequence = 'NA'
        # if cTemplateType == 'RFQTool':
        iMxRo = 2
        iMxRo = oSheetTemplate.max_row+1
        for iRow in range(1, iMxRo):
            # 
            # for jColumn in range(1, oSheetTemplate.max_column+1): making below as considoring 2 is max column for id col 
            for jColumn in range(1, 2):
                oCell = oSheetTemplate.cell(row=iRow, column=jColumn)
                cCellVal = oCell.value
                if cCellVal is not None:
                    if str(cCellVal).upper() in ('WWA LANE ID','WWA ID', 'CENTRIC ID','SR.NO.','INDEX ID', 'LANE ID') :
                        cKeyUniqueIDSequence = jColumn
                        break

        return cKeyUniqueIDSequence
    except Exception as e:
        LogService.log('Error :FunDCT_GetUniqueIDSequence Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetUniqueIDSequence Error while parsing file due to ' + str(e))

###############################
# CHECKED RATE INCREASE FUNCTIONS
###############################

def FunDCT_CheckedRateIncreaseValidations(oPostParameters, aSavedSearchedRow, oSheetTemplateUniqueSeq):
    try:
        if oPostParameters['_cUploadType'] != 'DISTRIBUTE':
            bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'CHECKED_RATE_INCREASE')
            if oPostParameters['cValidationString'] == 'CHECKED_RATE_INCREASE':
                if isinstance(oPostParameters['oCellVal'], int) or isinstance(oPostParameters['oCellVal'], float):
                    bMatch = FunDCT_ValidateRateIncrease(oPostParameters, aSavedSearchedRow, oSheetTemplateUniqueSeq)
                else:
                    bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'INVALID_NUMERIC')
        else:
            bMatch = True
        return bMatch
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_CheckedRateIncreaseValidations Error while parsing file due to ' + str(e) + ' Lineno = ' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CheckedRateIncreaseValidations Error while parsing file due to ' + str(e) + ' Lineno = ' + str(line))

def FunDCT_ValidateRateIncrease(oPostParameters, aSavedSearchedRow, oSheetTemplateUniqueSeq):
    try:
        bMatch = False
        cKeyUniqueID = str(oPostParameters['oSheetUploaded'].iat[int(oPostParameters['iRow'])-1,int(oPostParameters['cKeyUniqueIDSequence'])-1])
        if type(cKeyUniqueID) == str : cKeyUniqueID = int(cKeyUniqueID)
        # iFindRow = int(aSavedSearchedRow[str(cKeyUniqueID)])
        iFindRow = int(aSavedSearchedRow[cKeyUniqueID])
        if iFindRow != '':
            # jColumnPlusone = oPostParameters['iForwardValidator']+1
            jColumnPlusone = oPostParameters['iForwardValidator']
            cCellValOrg = oSheetTemplateUniqueSeq.iat[int(iFindRow), int(jColumnPlusone)]
            # cCellValOrg = oSheetTemplateUniqueSeq.cell(row=int(iFindRow), column=jColumnPlusone).value
            if int(float(cCellValOrg)) >= int(float(oPostParameters['oCellVal'])):
                bMatch = True
            else:
                bMatch = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE','CHECKED_RATE_INCREASE').replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$', str(cCellValOrg).strip())
        return bMatch
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_ValidateRateIncrease Error while parsing file due to ' + str(e) + '  ===> ' + str(line) + ' Org Val = ' + cCellValOrg + ' Current Cell Val = ' + str(oPostParameters['oCellVal']))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_ValidateRateIncrease Error while parsing file due to ' + str(e) + '  ===> ' + str(line) + ' Org Val = ' + cCellValOrg + ' Current Cell Val = ' + str(oPostParameters['oCellVal']))



###############################
# NO_DUPLICATE FUNCTIONS
###############################

def FunDCT_NoDuplicate(oPostParameters):
    try:
        if oPostParameters['_cUploadType'] != 'DISTRIBUTE':
            bMatch = MessageHandling.FunDCT_GetValidationDescription(
                'UPLOAD_TEMPLATE', 'NO_DUPLICATE')
            if oPostParameters['cValidationString'].startswith('NO_DUPLICATE'):
                oPostParameters['cDependentCol'] = oPostParameters['cValidationString'].replace(
                ')', '').split('NO_DUPLICATE(')[1]
                oPostParameters['cDependentCol'] = str(oPostParameters['cDependentCol']).replace("'", "")
                bMatch = FunDCT_ValidateDuplicate(oPostParameters)
        else:
            bMatch = True
        return bMatch
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_NoDuplicate Error while parsing file due to ' + str(e) + ' Lineno = ' + str(line))
        return MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE','NO_DUPLICATE').replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$',str(oPostParameters['oCellVal']))
        # return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_NoDuplicate Error while parsing file due to ' + str(e) + ' Lineno = ' + str(line))

def FunDCT_ValidateDuplicate(oPostParameters):
    try:
        bMatch = False
        cColumnIndex = FunDCT_GetColumnIndexFromCoordinate(oPostParameters['cDependentCol'])
        if cColumnIndex != '':
            cCellValOrg = str(oPostParameters['oSheetUploaded'].iat[int(oPostParameters['iRow'])-1, int(cColumnIndex)-1])
            if str(cCellValOrg) != str(oPostParameters['oCellVal']):
                bMatch = True
            else:
                bMatch = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE','NO_DUPLICATE').replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$',str(oPostParameters['oCellVal']))
        else:
            bMatch = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE','NO_DUPLICATE').replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$',str(oPostParameters['oCellVal']))
        return bMatch
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_ValidateDuplicate Error while parsing file due to ' + str(e) + '  ===> ' + str(line) + ' Org Val = ' + cCellValOrg + ' Current Cell Val = ' + str(oPostParameters['oCellVal']))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_ValidateDuplicate Error while parsing file due to ' + str(e) + '  ===> ' + str(line) + ' Org Val = ' + cCellValOrg + ' Current Cell Val = ' + str(oPostParameters['oCellVal']))

def FunDCT_GetColumnIndexFromCoordinate(cCoordinates):
    try:
        iExprn = 0
        iColNum = 0
        for cChar in reversed(cCoordinates):
            iColNum += (ord(cChar) - ord('A') + 1) * (26 ** iExprn)
            iExprn += 1

        return iColNum
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error :FunDCT_GetColumnIndexFromCoordinate Error while parsing file due to ' + str(e) + '  ===> ' + str(line) )
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetColumnIndexFromCoordinate Error while parsing file due to ' + str(e) + '  ===> ' + str(line) )


def FunDCT_ApplyHeaderFormating(wrkbkPath:str,iMaxDepthHeaders):
    wrbk = load_workbook(wrkbkPath)
    wrksheet = wrbk.active
    cColumnLetter1 = f"{get_column_letter(2)}1:{get_column_letter(wrksheet.max_column + 1)}1" 
    iii = 1
    jjj = 1
    for eachCell in wrksheet[cColumnLetter1]:
        for nCel in eachCell:
            jjj+=1
            if nCel.value is not None:
                ranges = f"{get_column_letter(iii)}1:{get_column_letter(jjj-1)}1"
                wrksheet.merge_cells(ranges)
                nCel.alignment =Alignment(horizontal="center", vertical="center")
                iii=jjj
    cColumnLetter = f"{get_column_letter(1)}1:{get_column_letter(wrksheet.max_column)}{iMaxDepthHeaders}"
    for rows in wrksheet[cColumnLetter]:
        for nCell in rows:
            thin_border = Side(style="thin")
            nCell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
            nCell.font = Font(bold=True)
    wrbk.save(wrkbkPath)
    wrbk.close()

def getDataValidationAndCordinates(oValidationDetails):
    dataValidationDict = dict()
    for oVal in oValidationDetails:
        cVal = oVal['cValidations']
        conVal = oVal['cValidationsCondition']
        if 'DATA_VALIDATION' in cVal:
              for eachVal in conVal:
                if 'DATA_VALIDATION' in eachVal: 
                    conditionalString = eachVal.replace('(','').replace(')','').replace('DATA_VALIDATION','')
                    listofdataValidation =conditionalString.split(',')
                    dataValidationDict[oVal['cHeaderLabel']] = listofdataValidation
    return dataValidationDict
def FunDCT_GetTextColor(oCell):
    try:
        cTextColor = '000000'
        if oCell.font.color:
            if 'rgb' in oCell.font.color.__dict__:
                cTextColor = oCell.font.color.rgb[2:]          

        return cTextColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))

def FunDCT_GetBackColor(oCell, cTextColor):
    try:
        cBackColor = 'ffffff'
        
        if oCell.fill.start_color != '' and type(oCell.fill.start_color.index) != int:                    
            cBackColor = oCell.fill.start_color.index[2:]
            if cTextColor == '000000' and cBackColor == '000000':
                cBackColor = 'ffffff'
        elif oCell.fill.start_color != '' and type(oCell.fill.start_color.index) == int:
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index][2:])
        elif isinstance(oCell.fill, PatternFill) and oCell.fill.patternType == 'solid':
            if oCell.fill.fgColor.type =='rgb':
                cBackColor = oCell.fill.fgColor.rgb[2:]
            elif oCell.fill.fgColor.type =='indexed':
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index -1][2:])
        
        return cBackColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))
