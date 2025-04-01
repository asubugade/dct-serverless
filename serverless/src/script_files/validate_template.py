import json
import openpyxl
import re

def lambda_handler(event, context):
    try:
        # Parse input payload
        path = event.get('path')
        aValidateRules = event.get('aValidateRules', [])
        aValidateRulesAll = event.get('aValidateRulesAll', [])

        # Load the file
        workbook = openpyxl.load_workbook(path['cDirFile'])
        sheet = workbook.active

        # Validation logic
        if sheet.max_row in [2, 3]:
            for j in range(1, sheet.max_column + 1):
                cell_value = str(sheet.cell(row=sheet.max_row, column=j).value).strip()
                if re.findall('[^a-zA-Z_]', cell_value):
                    if cell_value not in aValidateRules:
                        return {"statusCode": 400, "body": json.dumps("Validation failed")}

            return {"statusCode": 200, "body": json.dumps("Validation successful")}

        return {"statusCode": 400, "body": json.dumps("Invalid file format")}

    except Exception as e:
        return {"statusCode": 500, "body": json.dumps({"error": str(e)})}
