import shutil
from pathlib import Path
import sys
import inspect
import os
import json
import datetime
import xml.etree.cElementTree as ET
from collections import defaultdict
from shutil import copyfile
from datetime import datetime

from pandas.core.indexes.base import Index

from pyconfig.LogService import LogService
oNow = datetime.now()
import re
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

path = str(Path(Path(__file__).parent.absolute()).parent.absolute())
sys.path.insert(2, path)
file_dir = os.path.dirname(__file__)
sys.path.append(file_dir)
cmd_folder = os.path.realpath(os.path.abspath(os.path.split(inspect.getfile( inspect.currentframe() ))[0]))
if cmd_folder not in sys.path:
    sys.path.insert(3, cmd_folder)


import pandas
from openpyxl.reader.excel import load_workbook
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import xlsxwriter
import pydash
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode
from anytree.importer import JsonImporter
from anytree.importer import DictImporter
from anytree.exporter import DotExporter
from anytree import AnyNode, RenderTree

from script_files.MessageHandling import MessageHandling
from script_files import CommonValidations
from pyconfig import loadConfig
from pyconfig import awsCloudFileManager
from pyconfig.model_common import model_common_Cls

model_common = model_common_Cls()
MessageHandling = MessageHandling()
def FunDCT_DistributionAndConsolidation(_cCompanyname, cDirDistribution, cStatusFileFullFilePath, _cUploadType, oSheetUploaded, _iTemplateUploadLogID, cDirValidateTemplateStatusFile, iTemplateID, cDirConsolidation, iMaxDepthHeaders, aValidateContent, cS3DirStatusFile, cFullFilePath, cTemplateType):
    try:
        aSelectedMembers = defaultdict(dict)
        aMappedScacColumnsStat = defaultdict(dict)
        if len(CommonValidations.aDistributionMembers) <= 0:
            aGetSelectedMembers = model_common.FunDCT_GetMembersForDistribution(_iTemplateUploadLogID, _cUploadType)
            if _cUploadType != 'DISTRIBUTE':
                # oWorkbookMemberDistribute = openpyxl.Workbook()
                cMemberDistributionFilePath = cDirValidateTemplateStatusFile + aGetSelectedMembers + \
                    '_' + iTemplateID + '_' + \
                    str(oNow.strftime("%Y-%m-%d-%H-%M-%S")) + '.xlsx'
                # oWorkbookMemberDistribute.save(cMemberDistributionFilePath)
                # FunDCT_CopyUploadedFile(cMemberDistributionFilePath, oSheetUploaded)
                copyfile(cFullFilePath, cMemberDistributionFilePath)
                aSelectedMembers[len(aSelectedMembers)+1] = {
                    'cDistScaccode': aGetSelectedMembers,
                    'cMemberDistributionFilePath': cMemberDistributionFilePath,
                    'cMemberEmail': model_common.FunDCT_GetEmailFromCompanyName(aGetSelectedMembers),
                }
                if len(aGetSelectedMembers) > 0:
                    aSelectedMembers = FunDCT_DistributionS3Upload(aSelectedMembers, cS3DirStatusFile)
                # Template Consolidation File Member Wise
                aMappedScacColumnsStat = FunDCT_CreateMemberWiseConsolidationFile(cStatusFileFullFilePath, aGetSelectedMembers, iTemplateID, oSheetUploaded, _cCompanyname, cDirConsolidation, iMaxDepthHeaders, aValidateContent, _iTemplateUploadLogID)
            else:
                if len(aValidateContent) <= 0:
                    for i, cScacVal in enumerate(aGetSelectedMembers):
                        Path(cDirDistribution+str(iTemplateID)).mkdir(exist_ok=True)
                        cDistributeMemberFile =  str(cDirDistribution) + str(iTemplateID) + '/' + str(cScacVal).strip() + '_' + str(iTemplateID) + '.xlsx'
                        os.remove(cDistributeMemberFile) if os.path.exists(cDistributeMemberFile) else None
                        copyfile(cStatusFileFullFilePath, cDistributeMemberFile)
                        aSelectedMembers[len(aSelectedMembers)+1] = {
                            'cDistScaccode': cScacVal,
                            'cMemberDistributionFilePath': cDistributeMemberFile,
                            'cMemberEmail': model_common.FunDCT_GetEmailFromCompanyName(cScacVal),
                        }
                    if len(aGetSelectedMembers) > 0:
                        aSelectedMembers = FunDCT_DistributionS3Upload(aSelectedMembers, cS3DirStatusFile)
        else:
            if len(aValidateContent) <= 0:
                aSelectedMembers = FunDCT_DistributionS3Upload(CommonValidations.aDistributionMembers, cS3DirStatusFile)
            else:
                shutil.rmtree(str(cDirDistribution)+str(iTemplateID)) if os.path.exists(str(cDirDistribution)+str(iTemplateID)) else None
                aSelectedMembers = defaultdict(dict)
                model_common.FunDCT_ResetMappingColumnDetails(iTemplateID)

        return aSelectedMembers
    except Exception as e:
        LogService.log('Error : FuunDCT_DistributionAndConsolidation Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DistributionAndConsolidation Error while parsing file due to ' + str(e))

def FunDCT_DistributionS3Upload(aSelectedMembers, cS3DirStatusFile):
    try:
        for cKey in aSelectedMembers:
            cFullPath = str(aSelectedMembers[cKey]['cMemberDistributionFilePath'])
            cS3UrlStatusFile = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirStatusFile), cFullPath)
            # cS3UrlStatusFile = cFullPath
            aSelectedMembers[cKey]['cMemberDistributionFilePath'] = cS3UrlStatusFile
        return aSelectedMembers
    except Exception as e:
        LogService.log('Error : FuunDCT_DistributionS3Upload Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DistributionS3Upload Error while parsing file due to ' + str(e))

def FunDCT_CopyUploadedFile(cMemberDistributionFilePath, oSheetUploaded):
    try:
        oWorkBookWriteMember = openpyxl.load_workbook(
            cMemberDistributionFilePath)
        oWorkSheetMember = oWorkBookWriteMember.active
        iMaxRowUploaded = oSheetUploaded.max_row
        iMaxColumnUploaded = oSheetUploaded.max_column

        for iIndexRow in range(1, iMaxRowUploaded + 1):
            for jIndexCol in range(1, iMaxColumnUploaded + 1):
                oMemberCellVal = oSheetUploaded.cell(
                    row=iIndexRow, column=jIndexCol)
                oWorkSheetMember.cell(
                    row=iIndexRow, column=jIndexCol).value = oMemberCellVal.value
        oWorkBookWriteMember.save(str(cMemberDistributionFilePath))
    except Exception as e:
        LogService.log('Error : FuunDCT_CopyUploadedFile Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CopyUploadedFile Error while parsing file due to ' + str(e))

def FunDCT_ExcelToDataFrameDistribution(oFilePath:openpyxl):
    try:
        oDataFrame = pandas.ExcelFile(oFilePath,engine='openpyxl').parse(keep_default_na=False)
        iIndexDataFrame = iter(range(1, len(oDataFrame.columns) + 1))
        oDataFrame.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in oDataFrame.columns]
        return oDataFrame.to_json()
    except Exception as e:
        LogService.log('Error : FuunDCT_ExcelToDataFrameDistribution Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_ExcelToDataFrameDistribution Error while parsing file due to ' + str(e))

def FunDCT_CreateMemberWiseConsolidationFile(cSrcFile, cMemberScac, iTemplateID, oSheetUploaded, _cCompanyname, cDirConsolidation, iMaxDepthHeaders, aValidateContent, _iTemplateUploadLogID):
    try:
        aMemberScacColIndex = []
        aMemberScacColLabel = defaultdict(dict)
        aCountMappedColumn = defaultdict(dict)
        aGetMemberColumns = model_common.FunDCT_GetMemberScacValidationCol(iTemplateID)
        for cMemberColumnKey, aMemberColumnDetails in enumerate(aGetMemberColumns):
            if aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel'] == 'SERVICE INFO':
                continue
            iColumnIndex = CommonValidations.FunDCT_GetColumnIndex(oSheetUploaded, aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel'])
            aMemberScacColIndex.append(iColumnIndex)
            aMemberScacColLabel[iColumnIndex]['cHeaderLabel'] = str(aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel']).strip()
            aCountMappedColumn[str(_cCompanyname).strip()][aMemberScacColLabel[iColumnIndex]['cHeaderLabel']] = {
                'iTotalLanes': 0,
                'iSubmittedLanes': 0,
                'iRemainingLanes': 0,
            }
        # aCountMappedColumn,iMaxRowTemplate = FunDCT_DeleteErrorRows(cSrcFile, aMemberScacColIndex, aMemberScacColLabel, aCountMappedColumn, iMaxDepthHeaders, _cCompanyname)
        oWorkBookWriteMember = openpyxl.load_workbook(cSrcFile,read_only=True)
        oWorkSheetMember = oWorkBookWriteMember.active
        iMaxRowTemplate = oWorkSheetMember.max_row
        if iMaxDepthHeaders>1:
            iMaxRowTemplate = iMaxRowTemplate+1
        if iMaxRowTemplate > iMaxDepthHeaders:
            iTotalUploadedRows = FunDCT_MergeExistingConsolidateFile(cSrcFile, cMemberScac, cDirConsolidation, iTemplateID, iMaxDepthHeaders, aValidateContent, _iTemplateUploadLogID)
            cConsolidationFile = cDirConsolidation + str(iTemplateID) + '/' + str(iTemplateID) + '_' + str(cMemberScac) + '.xlsx'
            cMemberTemplateFile = Path(cConsolidationFile)
            oDataFrame = pandas.read_excel(cConsolidationFile,engine='openpyxl')
            iIndexDataFrame = iter(range(1, len(oDataFrame.columns) + 1))
            for iColumnIndex in aMemberScacColIndex:
                cHeaderLabel = aMemberScacColLabel[iColumnIndex]['cHeaderLabel']
                oDataFrame.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in oDataFrame.columns]
                iSubmittedLanes = len(oDataFrame[oDataFrame[str(oDataFrame.columns[iColumnIndex-1])] == str(cMemberScac)])
                model_common.FunDCT_UpdateLaneSubmission(iTemplateID, str(cMemberScac), cHeaderLabel, iSubmittedLanes)
        return cSrcFile
    except Exception as e:
        LogService.log('Error : FuunDCT_CreateMemberWiseConsolidationFile Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CreateMemberWiseConsolidationFile Error while parsing file due to ' + str(e))

def FunDCT_DeleteErrorRows(cSrcFile, aMemberScacColIndex, aMemberScacColLabel, aCountMappedColumn, iMaxDepthHeaders, _cCompanyname):
    try:
        bBook = load_workbook(cSrcFile)
        oEcl = pandas.ExcelFile(bBook,engine='openpyxl')
        # oDataFrame = pandas.read_excel(cSrcFile,dtype=str,engine='openpyxl')
        oDataFrame = oEcl.parse()
        iIndexDataFrame = iter(range(1, len(oDataFrame.columns) + 1))
        
        oDataFrame.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in oDataFrame.columns]
        # Below Line is govong error ==== C:\Users\mpatil\app\WWA-IT\dct-backend\src\script_files\Distribution.py:181: UserWarning: Boolean Series key will be reindexed to match DataFrame index.
        oDataFrame = oDataFrame[oDataFrame.stack().str.contains('Error -').any(level=0) == False]
        
        oDataFrame.columns = oDataFrame.columns.str.replace('#DCT#.*', '')
        r,c = oDataFrame.shape
        oDataFrame.to_excel(cSrcFile, index=False)

        return aCountMappedColumn,r
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : FuunDCT_DeleteErrorRows Error while parsing file due to ' + str(e) + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_DeleteErrorRows Error while parsing file due to ' + str(e) + str(line))


def FunDCT_MergeExistingConsolidateFile(cSrcFile, cMemberScac, cDirConsolidation, iTemplateID, iMaxDepthHeaders, aValidateContent, _iTemplateUploadLogID):
    try:
        iTotalUploadedRows = 0
        # oDataframeSrcFile = pandas.read_excel(cSrcFile, na_values = "Missing", na_filter=True,engine='openpyxl')
        # oDataframeSrcFile = CommonValidations.FunDCT_TrimAllColumnsDataFrame(oDataframeSrcFile)
        Path(cDirConsolidation+str(iTemplateID)).mkdir(exist_ok=True)
        cConsolidationFile = cDirConsolidation + str(iTemplateID) + '/' + str(iTemplateID) + '_' + str(cMemberScac) + '.xlsx'
        cMemberTemplateFile = Path(cConsolidationFile)
        aConsolidationData = defaultdict(dict)
        iExistingRows = 0
        if cMemberTemplateFile.is_file():
            oConsolidationFile = load_workbook(cConsolidationFile)
            oSheetUploadedConsolidation = oConsolidationFile.active
            iExistingRows = oSheetUploadedConsolidation.max_row
            oConsolidationFile.save(str(cConsolidationFile))

            oDataFrame = pandas.DataFrame()
            oFiles = [cSrcFile, cConsolidationFile]
            for file in oFiles:
                if file.endswith('.xlsx'):
                    bTbook = openpyxl.load_workbook(file)
                    bT = pandas.ExcelFile(bTbook,'openpyxl').parse(keep_default_na=False)
                    oDataFrame = oDataFrame.append(bT, ignore_index=True)
                    # oDataFrame.columns[0] should be dynamic, for WWA LANE ID, Unique value for each row
                    oDataFrame = oDataFrame.drop_duplicates(subset=[oDataFrame.columns[0]], keep='first')

            oDataFrame.columns = oDataFrame.columns.str.replace('Unnamed.*', '')
            oDataFrame.head()
            oDataFrame.reset_index(drop=True, inplace=True)
            oDataFrame.to_excel(cConsolidationFile, index=False)
        else:
            copyfile(cSrcFile, cConsolidationFile)
        
        oConsolidationFile = load_workbook(cConsolidationFile)
        oSheetUploadedConsolidation = oConsolidationFile.active
        iCurrentRows = oSheetUploadedConsolidation.max_row
        oConsolidationFile.save(str(cConsolidationFile))

        bExceptionfound = 'Y'
        if iCurrentRows >= iExistingRows and len(aValidateContent) > 0:
            bExceptionfound = 'P'
        elif iCurrentRows <= iExistingRows and len(aValidateContent) > 0:
            bExceptionfound = 'Y'
        elif iCurrentRows >= iExistingRows and len(aValidateContent) <= 0:
            bExceptionfound = 'N'
        
        model_common.FunDCT_UpdateExceptionFlag(_iTemplateUploadLogID, bExceptionfound)

        cConsolidationData = FunDCT_ExcelToDataFrameDistribution(oConsolidationFile)
        cConsolidationFileJSONFile = cDirConsolidation + str(iTemplateID) + '/' + str(iTemplateID) + '_' + str(cMemberScac) + '.json'

        with open(cConsolidationFileJSONFile, "w") as oFilePointer:
            json.dump(cConsolidationData , oFilePointer)
        cS3DirJSONFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_MEMBER) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_STATUSFILE) + '/'
    
        cS3UrlJSONFile = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirJSONFile), cConsolidationFileJSONFile)
        aConsolidationData[str(cMemberScac)] = { 
            'aConsolidationData': cS3UrlJSONFile
        }
        # aConsolidationData[str(cMemberScac)] = { 
        #     'aConsolidationData': cConsolidationData
        # }
        model_common.FunDCT_UpdateMappingColumnDetails(iTemplateID, cMemberScac, aConsolidationData, "aConsolidationData")

        return iTotalUploadedRows
    except Exception as e:
        LogService.log('Error : FuunDCT_MergeExistingConsolidateFile Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_MergeExistingConsolidateFile Error while parsing file due to ' + str(e))

