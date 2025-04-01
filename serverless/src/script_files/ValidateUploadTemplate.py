##
# Written by spirgonde for validating no. of rows in uploaded sample template file
# and also to check if validations provided in file for a selected Template are according to the mapping provided in database .
##

import warnings
import sys
import json
from pathlib import Path
import inspect
import os
import openpyxl
from io import BytesIO
import re
from pyconfig.LogService import LogService
warnings.simplefilter(action='ignore', category=FutureWarning)

path = str(Path(Path(__file__).parent.absolute()).parent.absolute())
sys.path.insert(0, path)
file_dir = os.path.dirname(__file__)
sys.path.append(file_dir)
cmd_folder = os.path.realpath(os.path.abspath(
    os.path.split(inspect.getfile(inspect.currentframe()))[0]))
if cmd_folder not in sys.path:
    sys.path.insert(0, cmd_folder)

from script_files.MessageHandling import MessageHandling
from pyconfig import loadConfig
MessageHandling = MessageHandling()
PY_VALIDATION_SEPARATOR = loadConfig.PY_VALIDATION_SEPARATOR
PY_RANGE_VALUE_SEPARATOR = loadConfig.PY_RANGE_VALUE_SEPARATOR
PY_BRACKET_VALIDATION_SEPARATOR = loadConfig.PY_BRACKET_VALIDATION_SEPARATOR
LogService.log(f'pyvalidatate seperatetor ==> {PY_VALIDATION_SEPARATOR}')
LogService.log(f'PY_RANGE_VALUE_SEPARATOR seperatetor ==> {PY_RANGE_VALUE_SEPARATOR}')
LogService.log(f'PY_BRACKET_VALIDATION_SEPARATOR seperatetor ==> {PY_BRACKET_VALIDATION_SEPARATOR}')
path = json.loads(sys.argv[1])
with open(r'c:\vartemp\createtemp.txt','w') as f:
    f.write(str(sys.argv[1]))
# with open(r'c:\vartemp\Distribute_Release.txt','r') as f:
#     oPostParameters = json.loads(f.read())


# with open(r'c:\vartemp\tempvalidateuploadtemplate.txt','w') as f:
#     f.write(str(sys.argv[1]))
# with open(r'c:\vartemp\tempvalidateuploadtemplate.txt','r') as f:
#     path = json.loads(f.read())


LogService.log(f'Argument recieved ==> {path}')

cFileContent = path['cFileContent']
aValidateRules = path['aValidateRules']
aValidateRulesAll = path['aValidateRulesAll']

bValidation = True
iValidatoinRow = 2  # Row in xlsx file in which Validations are present


def FunDCT_ValidateRows(content, aValidateRules, aValidateRulesAll):
    try:
        LogService.log('Rrow Validator Started ')
        file_stream = BytesIO(content)
        oWorkBook = openpyxl.load_workbook(file_stream)
        oSheet = oWorkBook.active
        if oSheet.max_row == 3 or oSheet.max_row ==2:
            bValidation = True
            for j in range(1, oSheet.max_column+1):
                cell_obj = oSheet.cell(row=oSheet.max_row, column=j)
                cColumnValue = str(cell_obj.value).strip()

                if re.findall('[^a-zA-Z_]', cColumnValue):
                    bValidation = FunDCT_CheckMultipleValidations(
                        cColumnValue, aValidateRules, aValidateRulesAll)
                    if bValidation == False:
                        break

                elif cell_obj.value not in aValidateRules:
                    bValidation = False
                    break
            # LogService.log(f'Validation istatus == {bValidation}')
            if bValidation:
                return MessageHandling.FunDCT_MessageHandling('Success', aValidateRules)
            else:
                LogService.log(f'ErrorValidation ==> {cColumnValue}')
                return MessageHandling.FunDCT_MessageHandling('ErrorValidation', cColumnValue)

        else:
            return MessageHandling.FunDCT_MessageHandling('Error', 'Please Select a valid file for Creating Template !')

    except Exception as e:
        LogService.log(e)
        print(e)


def FunDCT_CheckMultipleValidations(cColumnValue, aValidateRules, aValidateRulesAll):
    aValidations = cColumnValue.replace(PY_VALIDATION_SEPARATOR, ',').replace(
        PY_RANGE_VALUE_SEPARATOR, ',').replace(PY_BRACKET_VALIDATION_SEPARATOR, ',')
    aValidationFile = re.sub('[^a-zA-Z_]', ' ', aValidations).split(" ")
    # to convert string validations to list and to Filter out ' ' .
    liFileValidationsRaw = list(filter(None, aValidationFile))
    liFileValidations = []

    ##
    # loop to remove NS,NA,OR from validation provided in file such as MIXED_DATA(DEFAULT_LOCATION%%NS&&NA&&OR) or
    # to remove USD from MIXED_DATA(DEFAULT_CURRENCY%%USD)
    # to remove SUN,MON,TUE,WED from MIXED_DATA(SUN&&MON&&TUE&&WED)
    # and to append the remaining strings in liFileValidations list as ['MIXED_DATA','DEFAULT_LOCATION'] ,so that it can be cheecked
    # against validation rules present in db (aValidateRules) agains a provided template type.
    ##
    for av in aValidateRulesAll:
        for v in liFileValidationsRaw:
            if v == av:
                liFileValidations.append(v)

    # loop to check validations provided in file against the one present in db for a provided template type
    ##
    for v in liFileValidations:
        if v not in aValidateRules:
            return False

    return True


FunDCT_ValidateRows(cFileContent, aValidateRules,
                    aValidateRulesAll)
