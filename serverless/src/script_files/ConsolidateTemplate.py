import shutil
from pathlib import Path
import sys
import json
import pandas
from openpyxl.reader.excel import load_workbook
import openpyxl
import os
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors
from openpyxl import styles
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import xlsxwriter
import pydash
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode
from anytree.importer import JsonImporter
from anytree.importer import DictImporter
from anytree.exporter import DotExporter
from anytree import AnyNode, RenderTree
import pandas.api.types as pdt
from MessageHandling import MessageHandling
import CommonValidations
# import datetime
import xml.etree.cElementTree as ET
from collections import defaultdict

from datetime import datetime
from dateutil import parser

from pyconfig.LogService import LogService
oNow = datetime.now()

import re
from pyconfig import loadConfig
from pyconfig.model_common import model_common_Cls 
from pyconfig import awsCloudFileManager
import prefillNomemberupload
from shutil import copyfile
import glob

from bson import ObjectId

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

sys.path.insert(0, os.path.abspath(__file__))

# Explore Parameteres
oPostParameters = json.loads(sys.argv[1])
# with open(r'c:\vartemp\consolidate.txt','w') as f:
#     f.write(str(sys.argv[1]))
# with open(r'c:\vartemp\consolidate.txt','r') as f:
#     oPostParameters = json.loads(f.read())
# with open(r'c:\vartemp\consolidate_CASE_Compenent_missMatch1.txt','r') as f:
#     oPostParameters = json.loads(f.read())
iTemplateID = oPostParameters['iTemplateID']
model_common = model_common_Cls()
MessageHandling = MessageHandling()
oDataFrameList = model_common.FunDCT_GetTemplateLog('gen_templates',iTemplateID)
cTemplateType = oDataFrameList[0]['cTemplateType']
cFileDir = oPostParameters['cDirFile']
cFileType = '.xlsx'
cRequestType = oPostParameters['cRequestType'] # Either ALL / MEMEBER
aUnMappedColumnLabels = defaultdict(dict)
aUnMappedColumnRange = defaultdict(dict)
aUnMappedColumnRangeDetails = defaultdict(dict)
aRangeStart = []
aRangeEnd = []

aRangeMappedStart = []
aRangeMappedEnd = []
aMappedColumnRangeDetails = defaultdict(dict)
aMappedColumnsDetail = defaultdict(dict)

aUniqueIDsRow = defaultdict(dict)
joDistributedMetaData = {}
cSamplfilepath = str(cFileDir) +  'Template_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,oDataFrameList[0]['cTemplateSampleFile'],cSamplfilepath)
oWorkbookTemplate = load_workbook(cSamplfilepath)
oSheetTemplate = oWorkbookTemplate.active
iMaxRowTemplate = oSheetTemplate.max_row
iMaxColumnTemplate = oSheetTemplate.max_column


aHeaderDetails = model_common.FunDCT_GET_HEADER_VALIDATION(iTemplateID)
validations = model_common.FunDCT_GET_HEADER_VALIDATIONonlyChild(iTemplateID)
# added this for single header level
if not validations: validations = aHeaderDetails
aGetMemberColumns = model_common.FunDCT_GetMemberScacValidationCol(iTemplateID)
aMemberScacColIndex = []
aMemberScacColLabel = defaultdict(dict)
mappingSCACDict =  CommonValidations.getSCAC_columnMapping(validations)
service_col_list = [x+1 for x in  mappingSCACDict.get('SERVICEINFO_COLUMN')] if mappingSCACDict.get('SERVICEINFO_COLUMN') else []
unlockedif_col_list = [x+1 for x in  mappingSCACDict.get('UNLOCKEDIF')] if mappingSCACDict.get('UNLOCKEDIF') else []
def FunDCT_GetUnMappednColumnDetailsConsolidation(_cCompanyname):
    try:
        aRangeStart = []
        aRangeEnd = []
        aGetUnMappedColumnLabels = model_common.FunDCT_GetUnMappednColumns(iTemplateID)
        for cMemberColumnKey, aMemberColumnDetails in enumerate(aGetUnMappedColumnLabels):
            if aMemberColumnDetails.get('aUnMappedColumns'):
                for iIndex in range(len(aMemberColumnDetails['aUnMappedColumns'])):
                    if _cCompanyname in aMemberColumnDetails['aUnMappedColumns'].keys():
                        for cKey, cVal in aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0].items():
                            iRangeStart = int(aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0][cKey]['iActualCellCol'])
                            iTotalLen = int(aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0][cKey]['iTotalLen'])
                            iRangeEnd = int(iRangeStart) + int(iTotalLen-1)
                            
                            aRangeStart.append(iRangeStart)
                            aRangeEnd.append(int(iRangeStart) + int(iTotalLen)-1)
                            aUnMappedColumnRange[iRangeStart] = {
                                'iRangeStart': iRangeStart,
                                'iRangeEnd': iRangeEnd,
                                'iTotalLen': iTotalLen,
                            }
                            aUnMappedColumnRangeDetails[str(iRangeStart)+"-"+str(iRangeEnd)] = {
                                'iRangeStart': iRangeStart,
                                'iRangeEnd': iRangeEnd,
                                'iTotalLen': iTotalLen,
                            }
                            aUnMappedColumnLabels[_cCompanyname] = {
                                '_cCompanyname': aMemberColumnDetails['aUnMappedColumns'][_cCompanyname],
                            }
            if aMemberColumnDetails.get('aMappedColumns'):
                for iIndex in range(len(aMemberColumnDetails['aMappedColumns'])):
                    if _cCompanyname in aMemberColumnDetails['aMappedColumns'].keys():
                        for cKey, cVal in aMemberColumnDetails['aMappedColumns'][_cCompanyname][0].items():
                            iRangeStartMapped = int(aMemberColumnDetails['aMappedColumns'][_cCompanyname][0][cKey]['iActualCellCol'])
                            iTotalLenMapped = int(aMemberColumnDetails['aMappedColumns'][_cCompanyname][0][cKey]['iTotalLen'])
                            iTotalLen = int(aMemberColumnDetails['aMappedColumns'][_cCompanyname][0][cKey]['iTotalLen'])
                            aMappedColumnRangeDetails[_cCompanyname][cKey] = {
                                'iRangeStart' : iRangeStartMapped,
                                'iTotalLen' : iTotalLenMapped,
                                'iRangeEnd' : int(iRangeStartMapped) + int(iTotalLenMapped-1)
                            }
                            aRangeMappedStart.append(iRangeStartMapped)
                            aRangeMappedEnd.append(int(iRangeStartMapped) + int(iTotalLen)-1)
        return aRangeStart,aRangeEnd
    except Exception as e:
        LogService.log('Error : FunDCT_GetUnMappednColumnDetailsConsolidation Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetUnMappednColumnDetailsConsolidation Error while parsing file due to ' + str(e))

def FunDCT_GetNextColumnConsolidate(jColumn, iMaxColumnUploaded, aRangeStart, aRangeEnd):
    try:
        iSkipRows = 0
        for iLowerRange, iUpperRange in zip(aRangeStart, aRangeEnd):
            if iLowerRange <= jColumn <= iUpperRange:
                iSkipRows += aUnMappedColumnRangeDetails[str(iLowerRange)+"-"+str(iUpperRange)]['iTotalLen']
                jColumn += iSkipRows
                if jColumn <= iMaxColumnUploaded+1:
                    FunDCT_GetNextColumnConsolidate(jColumn, iMaxColumnUploaded, aRangeStart, aRangeEnd)
        return iSkipRows
    except Exception as e:
        LogService.log('Error : FunDCT_GetNextColumnConsolidate Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetNextColumnConsolidate Error while parsing file due to ' + str(e))

def FunDCT_PrepareRowData(aConsolidateData):
    try:
        aRowData = defaultdict(dict)
        for cLabel in aConsolidateData:
            iIndex = 1
            for cNewLabel in aConsolidateData[cLabel]:
                aRowData[iIndex][len(aRowData[iIndex])+1] = aConsolidateData[cLabel][cNewLabel]
                iIndex += 1
        return aRowData
    except Exception as e:
        LogService.log('Error : FunDCT_PrepareRowData Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_PrepareRowData Error while parsing file due to ' + str(e))

def FunDCT_GetMappedCol(aTemplateHeader, aRowData, aMemberScac):
    indexValue = list(aRowData.keys())[list(aRowData.values()).index(aMemberScac)]
    indexValue = indexValue -1 if indexValue else indexValue
    headersWith = aTemplateHeader[indexValue]
    headerWithValidation = headersWith.get('cValidations')
    mappinheaderValidation = ''
    if 'DEFAULT_ORIGIN_SCAC' in headerWithValidation:
        mappinheaderValidation = 'ORIGIN_COLUMN'
    elif 'DEFAULT_DESTINATION_SCAC' in headerWithValidation:
        mappinheaderValidation = 'DESTINATION_COLUMN'
    mappedListColumn = list()
    iCol = 1
    for eachHeader in aTemplateHeader:
        if mappinheaderValidation in  eachHeader.get('cValidations'):
            mappedListColumn.append(iCol)
        iCol+=1

    return mappedListColumn
def FunDCT_GetMappedColumnsDetail(aRowConsolidateData, cColumnKey, cMemberScac,OFR='',FOB='',PLC=''):
    try:
        aScacAvailableInColumn = defaultdict(dict)
        for cRowKey in aRowConsolidateData[cColumnKey]:
            if aRowConsolidateData[cColumnKey][cRowKey]:
                if str(aRowConsolidateData[cColumnKey][cRowKey]).strip().upper() == str(cMemberScac).strip().upper():
                    cHeaderVal = aRowConsolidateData[1][cRowKey]
                    if cHeaderVal == OFR or cHeaderVal == FOB or cHeaderVal == PLC:
                        aScacAvailableInColumn[cColumnKey][cHeaderVal] = cHeaderVal
        return aScacAvailableInColumn
    except Exception as e:
        LogService.log('Error : FunDCT_GetMappedColumnsDetail Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetMappedColumnsDetail Error while parsing file due to ' + str(e))

def FunDCT_InsertDefaultHeader():
    try:
        cConsolidateTemplateFileName = str(cFileDir) + 'CONSOLIDATE-' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
        os.remove(cConsolidateTemplateFileName) if os.path.exists(cConsolidateTemplateFileName) else None
        cSampleFile = Path(cConsolidateTemplateFileName)
        
        if not cSampleFile.is_file():
            oDataFrameList = model_common.FunDCT_GetTemplateLog('gen_templates',iTemplateID)
            oDataFrameSampleFile = pandas.DataFrame(eval(oDataFrameList[0]['cUploadedFile']))
            oDataFrameSampleFile.columns = oDataFrameSampleFile.columns.str.replace('#DCT#.*', '')
            oDataFrameSampleFile.to_excel(cConsolidateTemplateFileName, index=False)
        return cConsolidateTemplateFileName
    except Exception as e:
        LogService.log('Error : FunDCT_InsertDefaultHeader Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_InsertDefaultHeader Error while parsing file due to ' + str(e))

def FunDCT_SetMappedColRange(aMappedColumnsDetail, cColumnKey, cMemberScac):
    try:
        aRangeMappedStart = []
        aRangeMappedEnd = []
        for cMappedColumn in aMappedColumnsDetail[cColumnKey]:
            if cMappedColumn == 'FOB':
                iStartFromRange = int(aMappedColumnRangeDetails[cMemberScac]['ORIGIN CHARGES']['iRangeStart'])
                iStartToRange = int(aMappedColumnRangeDetails[cMemberScac]['ORIGIN CHARGES']['iRangeEnd'])
            elif cMappedColumn == 'OFR':
                iStartFromRange = aMappedColumnRangeDetails[cMemberScac]['OCEAN CHARGES']['iRangeStart']
                iStartToRange = aMappedColumnRangeDetails[cMemberScac]['OCEAN CHARGES']['iRangeEnd']
            elif cMappedColumn == 'PLC':
                iStartFromRange = aMappedColumnRangeDetails[cMemberScac]['DESTINATION CHARGES']['iRangeStart']
                iStartToRange = aMappedColumnRangeDetails[cMemberScac]['DESTINATION CHARGES']['iRangeEnd']
            aRangeMappedStart.append(iStartFromRange)
            aRangeMappedEnd.append(iStartToRange)
        
        return aRangeMappedStart, aRangeMappedEnd
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : FunDCT_SetMappedColRange Error while parsing file due to ' + str(e) + ' Line no = '+ str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_SetMappedColRange Error while parsing file due to ' + str(e) + ' Line no = '+ str(line))

def handle_blank_cells(data):
    
    if isinstance(data, dict):
        return {key: handle_blank_cells(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [handle_blank_cells(item) for item in data]
    elif data is None:
        return "" 
    else:
        return data
    

def FunDCT_FillRowOfUniqueID(cUniqueID, iFindRow):
    try:
        aUniqueIDsRow[str(cUniqueID)] = {
            'iFindRow':iFindRow
        }
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : FunDCT_FillRowOfUniqueID Error while parsing file due to ' + str(e) + ' Line no ' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_FillRowOfUniqueID Error while parsing file due to ' + str(e) + ' Line no ' + str(line))

def FunDCT_InsertNewRow(iMaxDepthHeaders, iMaxColumnTemplate, oWorksheet, cMemberScac, aLaneIDs, aRowConsolidateData, cKeyUniqueIDSequenceFind, aUniqueIDsRow):
    try:
        aRangeStart,aRangeEnd = FunDCT_GetUnMappednColumnDetailsConsolidation(cMemberScac)
        aKeys = []
        aUnMappedColumnRangeList = [(aRangeStart[iRangeIndex], aRangeEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeStart))]
        bInUnmappedRange = True
        aKeys = aUnMappedColumnLabels[cMemberScac]['_cCompanyname'][0].keys() if len(aUnMappedColumnLabels[cMemberScac]) > 0 else []
        
        iMergedCellLen = 0
        jColumn = 1
        iRow = oWorksheet.max_row + 1
        cKeyUniqueID = 0
        cKeyUniqueID = cKeyUniqueIDSequenceFind
        MappingSCAC = CommonValidations.getSCAC_columnMapping(aHeaderDetails)
        FOB,PLC,OFR = ('','','')
        
        try:

            FOB = aHeaderDetails[MappingSCAC.get('DEFAULT_ORIGIN_SCAC')]['cHeaderLabel']
            PLC = aHeaderDetails[MappingSCAC.get('DEFAULT_DESTINATION_SCAC')]['cHeaderLabel']
            OFR = aHeaderDetails[MappingSCAC.get('DEFAULT_OCEAN_SCAC')]['cHeaderLabel']
        except:
            pass
        iRow = oWorksheet.max_row + 1
        for cColumnKey in aRowConsolidateData:
            if cColumnKey < iMaxDepthHeaders:
                continue
            bSkip = False
            iSubsctractCols = 0
            jColumn = 1

            aMappedColumnsDetail = defaultdict(dict)
            # if iMaxDepthHeaders > 1:
            #     aMappedColumnsDetail = FunDCT_GetMappedColumnsDetail(aRowConsolidateData, cColumnKey, cMemberScac,FOB,OFR,PLC)
            #     aRangeMappedStart,aRangeMappedEnd = FunDCT_SetMappedColRange(aMappedColumnsDetail, cColumnKey, cMemberScac)
            aMappedColumnRangeList = [(aRangeMappedStart[iRangeIndex], aRangeMappedEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeMappedStart))]
            iMergedCellLen = 0
            for cVal in aRowConsolidateData[cColumnKey]:
                if len(aKeys) > 0:
                    
                    bInUnmappedRange = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aUnMappedColumnRangeList)
                    binMappedRange  = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aMappedColumnRangeList)
                    if bInUnmappedRange == True:
                        if FunDCT_GetNextColumnConsolidate(jColumn-iSubsctractCols,iMaxColumnTemplate, aRangeStart, aRangeEnd) > iSubsctractCols:
                                iSubsctractCols = FunDCT_GetNextColumnConsolidate(jColumn-iSubsctractCols,iMaxColumnTemplate, aRangeStart, aRangeEnd)
                        # iMergedCellLen = FunDCT_GetNextColumnConsolidate(jColumn, iMaxColumnTemplate, aRangeStart, aRangeEnd)
                    else:
                        if binMappedRange:
                            ccd1 = FunDCT_GetNextColumnConsolidate(jColumn+iSubsctractCols,iMaxColumnTemplate, aRangeStart, aRangeEnd)
                            ccd = FunDCT_GetNextColumnConsolidate(jColumn-iSubsctractCols,iMaxColumnTemplate, aRangeStart, aRangeEnd)
                            if ccd1 > 0 and ccd > 0:
                                iSubsctractCols = ccd1 + ccd
                # if (bInUnmappedRange is True) and iMergedCellLen > 0:
                #     iSubsctractCols = iMergedCellLen
                # elif (bInUnmappedRange is True) and iMergedCellLen <= 0:
                #     iSubsctractCols = iMergedCellLen
                                
                iForwardValidator = int(int(jColumn) + int(iSubsctractCols))
                try:
                    dValidations = validations[iForwardValidator-1]
                except:
                    continue
                conditinalValidation = dValidations.get('cValidationsCondition')
                sSCACname = ''
                if 'ORIGIN_COLUMN' in conditinalValidation:
                    sSCACname = 'DEFAULT_ORIGIN_SCAC' 
                elif 'OCEAN_COLUMN' in conditinalValidation:
                    sSCACname = 'DEFAULT_OCEAN_SCAC' 
                elif 'DESTINATION_COLUMN' in conditinalValidation:
                    sSCACname = 'DEFAULT_DESTINATION_SCAC' 
                elif 'SERVICEINFO_COLUMN' in conditinalValidation:
                    sSCACname ='SERVICEINFO_HEADER'
                bLiable = None
                if mappingSCACDict.get(sSCACname) is not None:
                    try:
                        if mappingSCACDict.get(sSCACname):
                            scacint = int(mappingSCACDict.get(sSCACname)) + 1
                            liablitySCAC = aRowConsolidateData[cColumnKey][scacint]
                            bLiable = liablitySCAC == cMemberScac

                    except:
                        pass
                
                if str(aRowConsolidateData[cColumnKey][cKeyUniqueID]) not in aLaneIDs:
                    valec  = aRowConsolidateData[cColumnKey][cVal]
                    if str(valec).lower().startswith('='):
                        formula = valec
                        matches = re.findall(r"([A-Z]+[0-9]+)\b",formula)
                        for mtch in matches:
                            rownumber = re.search(r'([0-9]+)\b',str(mtch)).group(1)
                            newR = mtch.replace(rownumber,str(iRow))
                            valec = valec.replace(mtch,newR) 
                    if bLiable is None or bLiable is True :
                        oWorksheet.cell(iRow, iForwardValidator).value = valec
                    if jColumn == 1 :
                        FunDCT_FillRowOfUniqueID(aRowConsolidateData[cColumnKey][cVal], iRow)
                    jColumn += 1
                        
                elif str(aRowConsolidateData[cColumnKey][cKeyUniqueID]) in aLaneIDs:
                    bInMappedRange = any(iLowerRange <= iForwardValidator <= iUpperRange for (iLowerRange, iUpperRange) in aMappedColumnRangeList)
                    iFindRow = 0
                    if bInMappedRange == True:
                        bSkip = True                        
                        iFindRow = 0
                        iFindRow = int(aUniqueIDsRow[str(aRowConsolidateData[cColumnKey][cKeyUniqueID])]['iFindRow'])
                        valec  =aRowConsolidateData[cColumnKey][cVal]
                        if str(valec).lower().startswith('='):
                            formula = valec
                            matches = re.findall(r"([A-Z]+[0-9]+)\b",formula)
                            for mtch in matches:
                                rownumber = re.search(r'([0-9]+)\b',str(mtch)).group(1)
                                newR = mtch.replace(rownumber,str(iFindRow))
                                valec = valec.replace(mtch,newR) 
                        if bLiable is None or bLiable :
                            oWorksheet.cell(iFindRow, iForwardValidator).value = valec
                        
                        if jColumn == 1 :
                            FunDCT_FillRowOfUniqueID(aRowConsolidateData[cColumnKey][cVal], iRow)
                        jColumn += 1
                    else:
                        iFindRow = 0
                        iFindRow = int(aUniqueIDsRow[str(aRowConsolidateData[cColumnKey][cKeyUniqueID])]['iFindRow'])
                        valec  =aRowConsolidateData[cColumnKey][cVal]
                        if str(valec).lower().startswith('='):
                            formula = valec
                            matches = re.findall(r"([A-Z]+[0-9]+)\b",formula)
                            for mtch in matches:
                                rownumber = re.search(r'([0-9]+)\b',str(mtch)).group(1)
                                newR = mtch.replace(rownumber,str(iFindRow))
                                valec = valec.replace(mtch,newR) 
                        if iForwardValidator in unlockedif_col_list:
                            if oWorksheet.cell(iFindRow, iForwardValidator).value != valec:
                                oWorksheet.cell(iFindRow, iForwardValidator).value = valec
                        if iForwardValidator in service_col_list:
                            if bLiable :
                                oWorksheet.cell(iFindRow, iForwardValidator).value = valec
                                
                        # if bLiable is None or bLiable :
                        #     oWorksheet.cell(iFindRow, iForwardValidator).value = valec
                        jColumn += 1
                         
                       
            if not bSkip : iRow += 1
        return aUniqueIDsRow
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : FunDCT_InsertNewRow Error while parsing file due to ' + str(e) + ' Line no ' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_InsertNewRow Error while parsing file due to ' + str(e) + ' Line no ' + str(line))

def FunDCT_LoadMemberDistributedData(cScac, iTemplateID):
    LogService.log(f'{iTemplateID} , {cScac}')
    objeyKey = model_common.FunDCT_GetMemeberDistributedFile(iTemplateID,cScac)
    LogService.log(f'Fetching file => {objeyKey}')
    # objeyKey = f'{loadConfig.AWS_BUCKET_ENV}/{loadConfig.AWS_BUCKET_TEMPLATES}/{loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE}/{loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_INTERTEAM}/{loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_STATUSFILE}/{cScac}_{iTemplateID}.xlsx'
    if objeyKey  == '':
        return ''
    try:
        cTempfilePath = str(cFileDir) + cScac +'_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
        awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,objeyKey,cTempfilePath)
        
        loadwb = load_workbook(cTempfilePath,read_only=True)
        odataframe = pandas.ExcelFile(loadwb,engine="openpyxl").parse(keep_default_na=False)
        odataframe = odataframe.dropna(axis=1, how='all')
        odataframe = odataframe.dropna(how='all')
        # odataframe = odataframe.loc[:, (odataframe != '').all(axis=0)]

        
        if iMaxRowTemplate > 1 :
            col = ''
            ii = 0
            for x in odataframe.iloc[0]:

                if x:
                    if 'date' in str(x) or 'Date' in str(x):
                        col = list(odataframe.columns)[ii]                        
                        odataframe[col] = odataframe[col].apply(dateconc)
                        # odataframe[col] = odataframe[col].dt.strftime('%Y-%m-%d %M:%H:%S')

                ii+=1
        else:
            for x in odataframe:
                if pdt.is_datetime64_any_dtype(odataframe[x]):
                        odataframe[x] = odataframe[x].dt.strftime('%Y-%m-%d %M:%H:%S')

        
        # odataframe = pandas.read_excel(oData, engine='openpyxl',dtype = str,na_values=str,)
        iIndexDataFrame = iter(range(1, len(odataframe.columns) + 1))
        odataframe.columns = [cColumns if cColumns.startswith('Unnamed') is not None else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in odataframe.columns]
        jData = odataframe.to_json()
        # loadwb.close()
        if os.path.exists(cTempfilePath): os.remove(cTempfilePath)
    except Exception as E:
        LogService.log(f'Errot while fetching data from file ==> {objeyKey}')
        return ''
    return jData
def dateconc(val):
    try:
        if isinstance(val, str):
            val = val.strip().strip("'")
        
        if isinstance(val, (int, float)) or (isinstance(val, str) and val.isdigit()):
            timestamp = int(val)
            if len(str(timestamp)) == 13: 
                # return datetime.utcfromtimestamp(timestamp / 1000).strftime('%Y-%m-%d %H:%M:%S')
                return datetime.fromtimestamp(timestamp / 1000).strftime('%Y-%m-%d %H:%M:%S')
        
        if isinstance(val, str):
            try:
                return datetime.strptime(val, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
            except ValueError:
                return val 
        
        return val 
    except Exception as e:
        print(f"Error processing value {val}: {e}")
        return val
    
def FunDCT_LoadJsonByObjectKey(s3objKey):
    LogService.log(f'Fetching file => {s3objKey}')
    if s3objKey  == '':
        return ''
    try:
        filetype = s3objKey.split('.')[-1]
        if filetype.startswith('xls'):
            cTempfilePath = str(cFileDir) + 'Temple_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
            awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,s3objKey,cTempfilePath)
        
            loadwb = load_workbook(cTempfilePath,read_only=True)
            odataframe = pandas.ExcelFile(loadwb,engine="openpyxl").parse(na_filter=True,)
            odataframe = odataframe.dropna(axis=1, how='all') #to Nan, none empty column 
            # odataframe = odataframe.loc[:, (odataframe != '').all(axis=0)] #to delete empty column 
            if iMaxRowTemplate > 1 :
                col = ''
                ii = 0
                for x in odataframe.iloc[0]:

                    if x:
                        if 'date' in str(x) or 'Date' in str(x):
                            col = list(odataframe.columns)[ii]                        
                            odataframe[col] = odataframe[col].apply(dateconc)
                            # odataframe[col] = odataframe[col].dt.strftime('%Y-%m-%d %M:%H:%S')

                    ii+=1
            else:
                for x in odataframe:
                    if pdt.is_datetime64_any_dtype(odataframe[x]):
                            odataframe[x] = odataframe[x].dt.strftime('%Y-%m-%d %M:%H:%S')
                # oData = awsCloudFileManager.getfile(loadConfig.AWS_BUCKET,s3objKey)
                # odataframe = pandas.read_excel(oData, engine='openpyxl',dtype = str)
            iIndexDataFrame = iter(range(1, len(odataframe.columns) + 1))
            odataframe.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in odataframe.columns]
            jData = odataframe.to_json()
            if os.path.exists(cTempfilePath): os.remove(cTempfilePath)
        else:

            oData = awsCloudFileManager.download_data_from_bucket(loadConfig.AWS_BUCKET,s3objKey)
            oData =str(oData).replace('null','\\"\\"')
            jsoLod = json.loads(oData)
            odataframe = pandas.DataFrame(eval(jsoLod),dtype=str,)
            iIndexDataFrame = iter(range(1, len(odataframe.columns) + 1))
            odataframe.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in odataframe.columns]
            df = odataframe.applymap(lambda x: x.replace("\\", "") if isinstance(x, str) else x)
            jData = df.to_json()
        return jData
    except Exception as E:
        LogService.log(f'Errot while fetching data from file ==> {s3objKey}')
        return ''


def FunDCT_GetTextColor(oCell):
    try:
        cTextColor = '000000'
        if oCell.font.color:
            if 'rgb' in oCell.font.color.__dict__:
                cTextColor = oCell.font.color.rgb[2:]          

        return cTextColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))

def FunDCT_GetBackColor(oCell, cTextColor):
    try:
        cBackColor = 'ffffff'
        
        if oCell.fill.start_color != '' and type(oCell.fill.start_color.index) != int:                    
            cBackColor = oCell.fill.start_color.index[2:]
            if cTextColor == '000000' and cBackColor == '000000':
                cBackColor = 'ffffff'
        elif oCell.fill.start_color != '' and type(oCell.fill.start_color.index) == int:
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index][2:])
        elif isinstance(oCell.fill, PatternFill) and oCell.fill.patternType == 'solid':
            if oCell.fill.fgColor.type =='rgb':
                cBackColor = oCell.fill.fgColor.rgb[2:]
            elif oCell.fill.fgColor.type =='indexed':
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index -1][2:])
        
        return cBackColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))

def FunDCT_GetMergeCellRangeLen(oSheetTemplate, iRow, jColumn):
    try:
        iTotalLen = 0
        oCell = oSheetTemplate.cell(iRow, jColumn)
        for mergedCell in oSheetTemplate.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                cCellRangeStart, cCellRangeEnd = str(mergedCell).split(':')
                iCellCol = CommonValidations.FunDCT_GetColumnNumber(
                    ''.join([i for i in cCellRangeStart if not i.isdigit()]))
                iTotalLen = int(CommonValidations.FunDCT_GetColumnNumber(
                    ''.join([i for i in cCellRangeEnd if not i.isdigit()]))) - (int(iCellCol)-1)
        return iTotalLen
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetMergeCellRangeLen Error while parsing file due to ' + str(e))


    
def FunDCT_MergeAllTemplateFile():
    try:
        if(str(oPostParameters['preFillData']).lower() == 'yes'.lower()):
            cConsolidateTemplateFileName = FunDCT_InsertDefaultHeader()
            oConsolidationTemplateLog = model_common.FunDCT_GetConsolidationData(iTemplateID)
            oDistributedMeta = model_common.FunDCT_GetDistributionData(iTemplateID)
            if cRequestType == 'ALL':
                
                for cMemberColumnKey, aMemberColumnDetails in enumerate(oConsolidationTemplateLog):
                    for cMemberScac in oConsolidationTemplateLog[cMemberColumnKey]['aConsolidationData'].keys():
                        # if 'aConsolidationData' in oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]:
                            # cConsolidationData = oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData']
                            cConsolidationData = oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]
                            cConsolidationData2 = cConsolidationData.get('aConsolidationData')
                            if cConsolidationData2 == '' or cConsolidationData2 is None:
                                oDistributedOriginalData = FunDCT_LoadMemberDistributedData(cMemberScac,iTemplateID)
                                if oDistributedOriginalData == '':
                                    continue
                                oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData'] = oDistributedOriginalData
                            else:
                                oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData'] = FunDCT_LoadJsonByObjectKey(cConsolidationData2)
            else:
                for cMemberColumnKey, aMemberColumnDetails in enumerate(oConsolidationTemplateLog):
                    for cMemberScac in oConsolidationTemplateLog[cMemberColumnKey]['aConsolidationData'].keys():
                            cConsolidationData = oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]
                            cConsolidationData2 = cConsolidationData.get('aConsolidationData')
                            if cConsolidationData2:
                                oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData'] = FunDCT_LoadJsonByObjectKey(cConsolidationData2)

                        

            oWorkbook = load_workbook(cConsolidateTemplateFileName)
            oWorksheet = oWorkbook.active
            oWorksheet.title = str(cRequestType) + '_CONSOLIDATE' if cRequestType else 'CONSOLIDATE'
            iMaxColumnTemplate = oWorksheet.max_column + 1
            iMaxDepthHeaders = model_common.FunDCT_GetMaxDepthHeader(iTemplateID)
            # cColumnLetter1 = f"{get_column_letter(2)}1:{get_column_letter(oWorksheet.max_column + 1)}1" 
            cKeyUniqueIDSequenceFind = CommonValidations.FunDCT_GetUniqueIDSequence(oWorksheet, cTemplateType)
            cKeyUniqueIDSequence = cKeyUniqueIDSequenceFind
            if cKeyUniqueIDSequence != 'NA' and cKeyUniqueIDSequence >= 0:
                for iRow in range(1, oWorksheet.max_row + 1):
                    if iRow <= iMaxDepthHeaders:
                        continue

            cTempfilePathUploaded = str(cFileDir) + 'UploadedFile_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
            awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,oDistributedMeta[0]['aDistributedData'],cTempfilePathUploaded)
            tBook = load_workbook(cTempfilePathUploaded,read_only=True)
            oPadasExcel = pandas.ExcelFile(tBook,'openpyxl')
            oDataFrame = oPadasExcel.parse(keep_default_na=False)
            oDataFrame = oDataFrame.dropna(how='all')
            # for column in oDataFrame.columns:
            #         if pandas.api.types.is_datetime64_any_dtype(oDataFrame[column]):
            #             oDataFrame[column] = oDataFrame[column].apply(lambda x: x.strftime('%Y-%m-%d') if pandas.notnull(x) else x)            
            oJson = oDataFrame.to_json()

            oDistributedMetaData = json.loads(oJson)
            oDistributedMetaData = handle_blank_cells(oDistributedMetaData)
            aLaneIDs = []
            joDistributedMetaData=FunDCT_PrepareRowData(oDistributedMetaData)
            FunDCT_InsertNewRow(iMaxDepthHeaders, iMaxColumnTemplate, oWorksheet, 'NO_SCAC', aLaneIDs, joDistributedMetaData, cKeyUniqueIDSequenceFind, aUniqueIDsRow)
            # oWorkbook.save(cConsolidateTemplateFileName)
            tBook.close()
            if os.path.exists(cTempfilePathUploaded): os.remove(cTempfilePathUploaded)
            iIndexseq = 1
            for cMemberColumnKey, aMemberColumnDetails in enumerate(oConsolidationTemplateLog):
                for cMemberScac in oConsolidationTemplateLog[cMemberColumnKey]['aConsolidationData'].keys():
                    if 'aConsolidationData' in oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]:
                        cConsolidationData = oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData']
                        if cConsolidationData != '':
                            
                            aRowConsolidateData = []
                            if iIndexseq <= 0:
                                # oDistributedOriginalData = FunDCT_LoadMemberDistributedData(cMemberScac,iTemplateID)

                                aConsolidateData = json.loads(cConsolidationData)
                                aRowConsolidateData = FunDCT_PrepareRowData(aConsolidateData)
                                FunDCT_InsertNewRow(iMaxDepthHeaders, iMaxColumnTemplate, oWorksheet, cMemberScac, aLaneIDs, aRowConsolidateData, cKeyUniqueIDSequenceFind, aUniqueIDsRow,joDistributedMetaData)
                                iIndexseq += 1
                            elif iIndexseq > 0:
                                iRow = 1
                                for aRows in oWorksheet.rows:
                                    if iRow > iMaxDepthHeaders:
                                        aLaneIDs.append(str(aRows[0].value))
                                    iRow += 1
                                aLaneIDsToSkipSeq = []
                                aConsolidateData = json.loads(cConsolidationData)
                                # if cTemplateType == 'RFQTool':
                                #     for cLabel in aConsolidateData['QUOTING MEMBER']:
                                #         if str(aConsolidateData['QUOTING MEMBER'][cLabel]) in aLaneIDs :
                                #             aLaneIDsToSkipSeq.append(str(aConsolidateData['QUOTING MEMBER'][cLabel]))
                                # else:
                                for cLabel in aConsolidateData[str(next(iter(aConsolidateData)))]:
                                    if str(aConsolidateData[str(next(iter(aConsolidateData)))][cLabel]) in aLaneIDs :
                                        aLaneIDsToSkipSeq.append(str(aConsolidateData[next(iter(aConsolidateData))][cLabel]))

                                aRowConsolidateData = FunDCT_PrepareRowData(aConsolidateData)
                                
                                FunDCT_InsertNewRow(iMaxDepthHeaders, iMaxColumnTemplate, oWorksheet, cMemberScac, aLaneIDsToSkipSeq, aRowConsolidateData, cKeyUniqueIDSequenceFind, aUniqueIDsRow)
                                iIndexseq += 1
            oWorkbook.save(cConsolidateTemplateFileName)
            if cRequestType == 'ALL':
                FunDCT_GapFilling(joDistributedMetaData,oWorksheet)
            oWorkbook.save(cConsolidateTemplateFileName)


            oBookSheet = oWorkbook.active
            FunDCT_Percentage(oBookSheet)
            #######Onlly header Mapping #####################
            aKeys = []
            bColumnHeaderNotInRange = False
            bInUnmappedRange = False
            # if len(aUnMappedColumnLabels) > 0:
            #     aUnMappedColumnRangeList = [(aRangeStart[iRangeIndex], aRangeEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeStart))]
            #     bInUnmappedRange = True
            #     bColumnHeaderNotInRange = True
            #     aKeys = aUnMappedColumnLabels[_cCompanyname]['_cCompanyname'][0].keys()
            
            iIndex = 0
            for iRow in range(1, iMaxRowTemplate+1):
                iSubsctractCols = 0
                for jColumn in range(1, iMaxColumnTemplate+1):
                    oSampleCell = oSheetTemplate.cell(row=iRow, column=jColumn)
                    oSampleCellVal = oSampleCell.value
                    # if len(aKeys) > 0:
                    #     bColumnHeaderNotInRange = True if str(oSampleCellVal) in aKeys else False
                    #     bInUnmappedRange = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aUnMappedColumnRangeList)
                    
                    iMergedCellLen = FunDCT_GetMergeCellRangeLen(oSheetTemplate, iRow, jColumn)
                    if (oSampleCellVal is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen > 0:
                        iSubsctractCols += iMergedCellLen
                        continue
                    elif (oSampleCellVal is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen <= 0:
                        iSubsctractCols += 1
                        continue
                    # elif (oSampleCellVal is None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True):
                    #     iSubsctractCols += 1
                    #     continue

                    # iBackwardCol = int(jColumn) - int(iSubsctractCols)
                    # oSampleCell = oSheetTemplate.cell(row=iRow, column=iBackwardCol)
                    # oSampleCellVal = oSampleCell.value
                    

                    if oSampleCellVal is not None:
                        iBackwardCol = int(jColumn) - int(iSubsctractCols)
                        oUploadedCell = oBookSheet.cell(row=iRow, column=iBackwardCol)
                        oUploadedCellVal = str(oUploadedCell.value)
                        bMatchHeaderValue = True if str(oSampleCellVal).strip() == str(oUploadedCellVal) else False

                        # if not bMatchHeaderValue:
                        #     cErrorMessage = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE','TEMPLATE_VALIDATE_HEADER_MISMATCH').replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$', str(oSampleCellVal)).replace('#$DCTVARIABLE_INCORRECT_CELL_VAL#$', str(oUploadedCellVal) )
                        #     FunDCT_PrepareTemplateStatusFile(iIndex, oUploadedCell.row, oUploadedCell.column, str(oUploadedCell.coordinate), cErrorMessage)
                        #     iIndex += 1
                        # else:
                        oBookSheet.cell(iRow, iBackwardCol).value = oSampleCell.value                       
                        oBookSheet.cell(iRow, iBackwardCol).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cTextColor = FunDCT_GetTextColor(oSampleCell)
                        cBackColor = FunDCT_GetBackColor(oSampleCell,cTextColor)
                        thin_border = Side(style="thin")
                        oBookSheet.cell(row=iRow, column=iBackwardCol).border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                        oBookSheet.cell(row=iRow, column=iBackwardCol).font = Font(color=cTextColor)
                        oBookSheet.cell(row=iRow, column=iBackwardCol).fill = PatternFill(fgColor=cBackColor,fill_type = "solid")
                        if FunDCT_CellMerged(oSheetTemplate, iRow, iBackwardCol):
                            if iMergedCellLen > 0:
                                ranges = f"{get_column_letter(iBackwardCol)}{iRow}:{get_column_letter((iBackwardCol - 1)+iMergedCellLen)}{iRow}"
                                oBookSheet.merge_cells(ranges)
                ################## Heading End


            
            oWorkbook.save(cConsolidateTemplateFileName)
            oWorkbook.close()
            oWorkbookTemplate.close()
                

            if os.path.exists(cSamplfilepath): os.remove(cSamplfilepath)

            cS3DirUploadFile = ''
            cS3DirUploadFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_CONSOLIDATION) + '/'
            cS3UrlUploadedOrg = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirUploadFile), cConsolidateTemplateFileName)
        
            oStatusDetails = model_common.statusActive[0]
            iStatusID = oStatusDetails.get('_id')
            pending_consl_req_id = ObjectId(oPostParameters.get('oPendingConslReq', ''))
            existing_data = model_common.FunDCT_GetPendingConsolidationData(iStatusID,pending_consl_req_id)
            is_consolidation = existing_data.get("IsConsolidationRequested", {})

            IsConsolidationRequested = {
                "cRequestType": is_consolidation.get("cRequestType", ""),
                "iEnteredby": is_consolidation.get("iEnteredby", 0),
                "iTemplateID": is_consolidation.get("iTemplateID", 0),
                "IsAlreadyRequested": False
            }
            aRequestDetails = {
                "cTemplateStatusFile": cS3UrlUploadedOrg,
                "IsConsolidationRequested": IsConsolidationRequested,
                "bProcesslock": "N",
                "bProcessed": "Y",
                "tProcessEnd": datetime.utcnow(),
                "tUpdated": datetime.utcnow()
            }

            model_common.FunDCT_UpdateConsolidationData(iStatusID,pending_consl_req_id,aRequestDetails)

            oTemplateDetails = model_common.FunDCT_GetTemplateDetails(iTemplateID)
            model_common.FunDCT_SendConsolidateFile(cS3UrlUploadedOrg, oTemplateDetails, oPostParameters)

            return MessageHandling.FunDCT_MessageHandling('Success', cS3UrlUploadedOrg)
        else:
            iMaxDepthHeaders = model_common.FunDCT_GetMaxDepthHeader(iTemplateID)
            oPostParameters['iMaxDepthHeaders']=iMaxDepthHeaders
            
            prefillNomemberupload.PreFilledDataConsolidation(oPostParameters)

    except Exception as e:
        
        oStatusDetails = model_common.statusActive[0]
        iStatusID = oStatusDetails.get('_id')
        pending_consl_req_id = ObjectId(oPostParameters.get('oPendingConslReq', ''))
        existing_data = model_common.FunDCT_GetPendingConsolidationData(iStatusID,pending_consl_req_id)
        is_consolidation = existing_data.get("IsConsolidationRequested", {})
        IsConsolidationRequested = {
            "cRequestType": is_consolidation.get("cRequestType", ""),
            "iEnteredby": is_consolidation.get("iEnteredby", 0),
            "iTemplateID": is_consolidation.get("iTemplateID", 0),
            "IsAlreadyRequested": False
        }
        aRequestDetails = {
            "IsConsolidationRequested": IsConsolidationRequested,
			"bExceptionfound": 'Y',
            "bProcesslock": "N",
            "bProcessed": "F",
            "tProcessEnd": datetime.utcnow(),
            "tUpdated": datetime.utcnow()
        }
        model_common.FunDCT_UpdateConsolidationData(iStatusID,pending_consl_req_id,aRequestDetails)
        
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : FunDCT_MergeAllTemplateFile Error while parsing file due to ' + str(e) + 'Line no - '+ str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_MergeAllTemplateFile Error while parsing file due to ' + str(e) + 'Line no - '+ str(line))

def FunDCT_Percentage(oWorksheet):
    percentage_columns = [] 

    try:
        parent_row = next(oWorksheet.iter_rows(values_only=True, max_row=1))
    except StopIteration:
        print("No rows found in the worksheet")
        return

    try:
        subheader_row = next(oWorksheet.iter_rows(values_only=True, min_row=2, max_row=2))
    except StopIteration:
        subheader_row = [] 

    for idx, parent_value in enumerate(parent_row):
        if 'SERVICE INFORMATION' in str(parent_value).upper():
            for sub_idx, subheader_value in enumerate(subheader_row):
                if 'PERCENTAGE' in str(subheader_value).upper():
                    percentage_columns.append(sub_idx)
                    break
            if not subheader_row: 
                percentage_columns.append(idx)

    for ii, row in enumerate(oWorksheet.iter_rows(values_only=False), start=1):
        formatted_row = []
        for idx, cell in enumerate(row):
            formatted_value = cell.value

            if isinstance(cell.value, str):
                try:
                    formatted_value = float(cell.value)
                except ValueError:
                    formatted_value = cell.value

            if isinstance(formatted_value, (int, float)) and idx in percentage_columns:
                formatted_value = f"{round(formatted_value * 100, 2)}%"

            formatted_value = dateconc(formatted_value)

            formatted_row.append(formatted_value)

        for idx, cell in enumerate(row):
            cell.value = formatted_row[idx]
                
def FunDCT_CellMerged(oSheet, iRow, jColumn):
    try:
        oCell = oSheet.cell(iRow, jColumn)
        for mergedCell in oSheet.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                return True
        return False
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CellMerged Error while parsing file due to ' + str(e))
def FunDCT_GapFilling(joDistributedMetaData, oWorksheet):
    ft = Font(color="FF0000")
    pendingRows=list()
    for index in joDistributedMetaData:
        if iMaxRowTemplate > 1 and index <= 1:
            continue

        iFindRow = 0
        try:
            iFindRow = int(aUniqueIDsRow[str(joDistributedMetaData[index][1])]['iFindRow'])
        except:
            pass
        if iFindRow == 0 :
            pendingRows.append(joDistributedMetaData[index])
        else:
            for cellindex in range(1,len(joDistributedMetaData[index])+1):
                if oWorksheet.cell(iFindRow,cellindex).value is None or oWorksheet.cell(iFindRow,cellindex).value  == '':
                    oWorksheet.cell(iFindRow,cellindex).value = str(joDistributedMetaData[index][cellindex])
                elif str(joDistributedMetaData[index][cellindex]) != str(oWorksheet.cell(iFindRow,cellindex).value):

                    original_value = joDistributedMetaData[index][cellindex]
                    if isinstance(original_value, float):
                        cmp1 = f"{round(original_value * 100, 2)}%"
                    else:
                        cmp1 = str(original_value)
                    cmp2 = str(oWorksheet.cell(iFindRow,cellindex).value)
                    normalized_cmp1 = normalize_date(cmp1)
                    normalized_cmp2 = normalize_date(cmp2)
                    cmp = normalize_values(cmp1, cmp2)
                    if cmp1 == cmp2:
                        # oWorksheet.cell(iFindRow,cellindex).value = str(joDistributedMetaData[index][cellindex])
                        oWorksheet.cell(iFindRow,cellindex).value = cmp1
                    elif cmp[0] == cmp[1]:
                        oWorksheet.cell(iFindRow,cellindex).value = cmp[0]
                    elif normalized_cmp1 == normalized_cmp2 and (normalized_cmp1!=None or normalized_cmp2!=None):
                        oWorksheet.cell(iFindRow,cellindex).value = str(joDistributedMetaData[index][cellindex])
                    else:
                        # if isinstance(oWorksheet.cell(iFindRow,cellindex).value,datetime):
                        #     try:
                        #         date1 = datetime.strftime(oWorksheet.cell(iFindRow,cellindex).value,oWorksheet.cell(iFindRow,cellindex).number_format)
                        #         date2 =  datetime.strftime(cmp1,'%Y-%m-%d')
                        #         if date1 != date2:
                        #             oWorksheet.cell(iFindRow,cellindex).font = ft
                        #     except:
                        #         pass
                        # else:
                        oWorksheet.cell(iFindRow,cellindex).font = ft
                # elif not oWorksheet.cell(iFindRow,cellindex).value:
                #     oWorksheet.cell(iFindRow,cellindex).value = str(joDistributedMetaData[index][cellindex])

def normalize_values(cmp1, cmp2):
    # Check if one value is "100%" and the other is 1
    if (cmp1 == '1' and cmp2 == "100%") or (cmp1 == "100%" and cmp2 == '1'):
        return "100%", "100%"  # Normalize both to "100%"
    return cmp1, cmp2

def normalize_date(value):
    try:
        if value.isdigit() and len(value) >= 10:  
            timestamp = int(value)
            if timestamp > 10**10:  
                timestamp /= 1000  
            return datetime.fromtimestamp(timestamp).date()
        else:
            return parser.parse(value).date()
    except (ValueError, TypeError):
        return
   

FunDCT_MergeAllTemplateFile()
