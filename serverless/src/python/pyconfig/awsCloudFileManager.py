from copy import copy
import io
import sys
import json
import pandas
import numpy as np
from openpyxl.reader.excel import load_workbook
import openpyxl
import os
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import xlsxwriter
import pydash
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode
from anytree.importer import JsonImporter
from anytree.importer import DictImporter
from anytree.exporter import DotExporter
from anytree import AnyNode, RenderTree
import MessageHandling
import datetime
import xml.etree.cElementTree as ET
from collections import defaultdict

from shutil import copyfile
from pathlib import Path
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

import re
from pyconfig import loadConfig
from datetime import datetime
oNow = datetime.now()

import boto3

# cmd_folder = os.path.realpath(os.path.abspath(os.path.split(inspect.getfile( inspect.currentframe() ))[0]))
# if cmd_folder not in sys.path:
#     sys.path.insert(0, cmd_folder)

def aws_session():
    return boto3.session.Session(aws_access_key_id=loadConfig.AWS_ACCESS_KEY,
                                aws_secret_access_key=loadConfig.AWS_SECRET_KEY,
                                region_name=loadConfig.AWS_REGION)


def make_bucket(name, acl):
    session = aws_session()
    s3_resource = session.resource('s3')
    return s3_resource.create_bucket(Bucket=name, ACL=acl)


def upload_file_to_bucket(bucket_name, uploadTo, file_path):
    session = aws_session()
    s3_resource = session.resource('s3')
    file_dir, file_name = os.path.split(file_path)

    bucket = s3_resource.Bucket(bucket_name)
    bucket.upload_file(
      Filename=file_path,
      Key=uploadTo + file_name,
    )
    s3_url = uploadTo + file_name
    return s3_url

def getPresignedUrlPresingedURL(bucket_name, uploadTo, file_path):
    session = aws_session()
    
    s3_resource = session.resource('s3')
    file_dir, file_name = os.path.split(file_path)

    bucket = s3_resource.Bucket(bucket_name)
    bucket.upload_file(
      Filename=file_path,
      Key=uploadTo,
    )
    s3_url = uploadTo
    presigned_url = session.client('s3').generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket_name, 'Key': s3_url},
            ExpiresIn=3600
        )
    return presigned_url


def download_file_from_bucket(bucket_name, s3_key, dst_path):
    session = aws_session()
    s3_resource = session.resource('s3')
    bucket = s3_resource.Bucket(bucket_name)
    bucket.download_file(Key=s3_key, Filename=dst_path)
 

def upload_data_to_bucket(bytes_data, bucket_name, s3_key):
    session = aws_session()
    s3_resource = session.resource('s3')
    obj = s3_resource.Object(bucket_name, s3_key)
    obj.put(ACL='private', Body=bytes_data)

    s3_url = f"https://{bucket_name}.s3.amazonaws.com/{s3_key}"
    return s3_url


def download_data_from_bucket(bucket_name, s3_key):
    session = aws_session()
    s3_resource = session.resource('s3')
    obj = s3_resource.Object(bucket_name, s3_key)
    io_stream = io.BytesIO()
    obj.download_fileobj(io_stream)

    io_stream.seek(0)
    data = io_stream.read().decode('utf-8')

    return data
def getfile(bucket_name, s3_key):
    session = aws_session()
    s3_resource = session.resource('s3')
    obj = s3_resource.Object(bucket_name, s3_key)
    io_stream = io.BytesIO()
    obj.download_fileobj(io_stream)

    io_stream.seek(0)
    data = io_stream.read()
    return data