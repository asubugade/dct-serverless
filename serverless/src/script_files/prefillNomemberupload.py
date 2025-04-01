import openpyxl.worksheet
import pandas as pd
import sys,inspect,os
from pathlib import Path
path = str(Path(Path(__file__).parent.absolute()).parent.absolute())
sys.path.insert(0, path)
file_dir = os.path.dirname(__file__)
sys.path.append(file_dir)
cmd_folder = os.path.realpath(os.path.abspath(os.path.split(inspect.getfile( inspect.currentframe() ))[0]))
if cmd_folder not in sys.path:
    sys.path.insert(0, cmd_folder)

from script_files.pyconfig.LogService import LogService
from pyconfig.model_common import model_common_Cls
from pyconfig import awsCloudFileManager
from pyconfig import loadConfig
import json
from openpyxl.reader.excel import load_workbook
import openpyxl
from openpyxl.workbook import Workbook
from collections import defaultdict
from script_files.MessageHandling import MessageHandling
import warnings
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors,protection,Protection
from openpyxl.cell import Cell
from datetime import datetime
oNow = datetime.now()
model_common = model_common_Cls()
MessageHandling = MessageHandling()
def FunDCT_PreFillNoDataUpload(oPostParameters):
    try:
        exl = pd.ExcelFile(str(oPostParameters['cFullFilePath']),engine='openpyxl')
        sheet_to_df_map = {}
        totrowsumiteed = 0
        exlbook = exl.book
        for sht in exl.sheet_names:
            if exlbook[sht].sheet_state == 'visible':
                warnings.simplefilter(action='ignore',category=UserWarning)
                tdf = exl.parse(sht,int(oPostParameters['iMaxDepthHeaders']) + int(oPostParameters['startHeaderRowIndex'])-1)
                
                columss = list(tdf.columns)
                tdf=tdf.dropna(subset=columss[:5],how='all')
                r,c = tdf.shape
                totrowsumiteed +=r
                for column in tdf.columns:
                    if pd.api.types.is_datetime64_any_dtype(tdf[column]):
                        tdf[column] = tdf[column].apply(lambda x: x.strftime('%d-%b-%Y') if pd.notnull(x) else x)
                sheet_to_df_map[sht] = tdf.to_json()

        exl.close()
        # cConsolidationData = json.dumps(sheet_to_df_map,)
        cS3DirJSONFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_MEMBER) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_STATUSFILE) + '/'
        cConsolidationFileJSONFile = os.path.abspath(str(oPostParameters['cFileDir'])) + '/' + oPostParameters['_iTemplateUploadLogID'] + str(oNow.strftime("%Y-%m-%d-%H-%M-%S")) + '.json'
        # cConsolidationFileJSONFile = os.path.abspath(str(oPostParameters['cFullFilePath'])) + '/ ' + oPostParameters['_iTemplateUploadLogID'] + '.json'
        with open(cConsolidationFileJSONFile, "w") as oFilePointer:
            json.dump(sheet_to_df_map , oFilePointer)    
        cS3UrlJSONFile = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirJSONFile), cConsolidationFileJSONFile)
        os.remove(cConsolidationFileJSONFile)
        aConsolidationData = {}
        aConsolidationData[str(oPostParameters['_cCompanyname'])] = { 
                'aConsolidationData': cS3UrlJSONFile
            }
        model_common.FunDCT_UpdateMappingColumnDetails(oPostParameters['iTemplateID'], oPostParameters['_cCompanyname'], aConsolidationData, "aConsolidationData")
        model_common.FunDCT_UpdateExceptionFlag(oPostParameters['_iTemplateUploadLogID'], 'N')
        model_common.FunDCT_AddLaneSubmission(oPostParameters['iTemplateID'], str(oPostParameters['_cCompanyname']), 'PreFilledNo', totrowsumiteed)
        return str(oPostParameters['_cCompanyname'])
    except Exception as err:
        LogService.log(f'Error while PreFillNo file member upload =={err}')
        model_common.FunDCT_UpdateExceptionFlag(oPostParameters['_iTemplateUploadLogID'], 'Y')
        


def PreFilledDataConsolidation(oPostParameters):
    tmpl = model_common.FunDCT_GetTemplateUplodedFile('tmpl_uploadlogs',oPostParameters['iTemplateID'])
    tmplS3File = tmpl[0]['cTemplateFile']
    fFile = oPostParameters['cDirFile'] + 'CONSOLIDATE-' + str(oPostParameters['iTemplateID']) + '_' + str(oNow.strftime("%Y-%m-%d-%H-%M-%S")) +'.xlsx'

    # data = awsCloudFileManager.getfile(loadConfig.AWS_BUCKET,tmplS3File)
    warnings.simplefilter(action='ignore',category=UserWarning)
    awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,tmplS3File,fFile)
    # exl = pd.ExcelFile(fFile,engine='openpyxl',)
    # exl = pd.ExcelFile("C:/Users/atpathan/Downloads/WWA_Member_Data_Import_Template_20241.xlsx",engine='openpyxl')
    # openxlBok = exl.book
    # openxlBok.save(fFile)
    
    # opnxl = openpyxl.Workbook()
    # opnxl.save(fFile)
    opnxl =openpyxl.load_workbook(fFile,read_only=False,keep_links=False)
    # copy_sheets(exl.book,opnxl)
    # opnxl.save(fFile)
    for sh in opnxl.worksheets:
        if sh.sheet_state == 'visible':
            mr = sh.max_row
            sh.delete_rows(int(oPostParameters['iMaxDepthHeaders']) + int(oPostParameters['startHeaderRowIndex'])+1,mr)

        
    oConsolidationTemplateLog = model_common.FunDCT_GetConsolidationData(oPostParameters['iTemplateID'])
    for oWorksheet in opnxl.worksheets:
        
        if oWorksheet.sheet_state == 'visible':
            for cMemberColumnKey, aMemberColumnDetails in enumerate(oConsolidationTemplateLog):
                for cMemberScac in oConsolidationTemplateLog[cMemberColumnKey]['aConsolidationData'].keys():
                    for cMemberColumnKey, aMemberColumnDetails in enumerate(oConsolidationTemplateLog):
                        if 'aConsolidationData' in oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]:
                            cConsolidationData = oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData']
                            if cConsolidationData != '':
                                oData = awsCloudFileManager.download_data_from_bucket(loadConfig.AWS_BUCKET,cConsolidationData)
                                # oDataL = json.loads(oData)
                                # eva = eval(oDataL)
                                # odataframe = pd.DataFrame(eva)
                                # iIndexDataFrame = iter(range(1, len(odataframe.columns) + 1))
                                # odataframe.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in odataframe.columns]
                                # jcConsolidationData = odataframe.to_json()
                                # aConsolidateData = json.loads(jcConsolidationData)
                                # aRowConsolidateData = FunDCT_PrepareRowData(aConsolidateData)
                                aConsolidateData = json.loads(oData)
                                aRowConsolidateData = FunDCT_PrepareRowData(json.loads(aConsolidateData[oWorksheet.title]))
                                
                                for cColumnKey in aRowConsolidateData:
                                    
                                    iRow = oWorksheet.max_row + 1
                                    iSubsctractCols = 0
                                    jColumn = 1
                                    for cVal in aRowConsolidateData[cColumnKey]:
                                        oWorksheet.cell(iRow, jColumn).value = aRowConsolidateData[cColumnKey][cVal]
                                        jColumn+=1
                                    oWorksheet.cell(iRow, jColumn+1).value = cMemberScac
        # openxlBok.save('C:/Users/atpathan/Downloads/WWA_Member_Data_Import_Template_2024.xlsx')
    opnxl.save(fFile)
    
    cS3DirUploadFile = str(loadConfig.AWS_BUCKET) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_CONSOLIDATION) + '/'
    cS3UrlUploadedOrg = awsCloudFileManager.upload_file_to_bucket(loadConfig.AWS_BUCKET,str(cS3DirUploadFile),fFile)
    return MessageHandling.FunDCT_MessageHandling('Success', cS3UrlUploadedOrg)

    

def FunDCT_PrepareRowData(aConsolidateData):
    try:
        aRowData = defaultdict(dict)
        for cLabel in aConsolidateData:
            iIndex = 1
            for cNewLabel in aConsolidateData[cLabel]:
                aRowData[iIndex][len(aRowData[iIndex])+1] = aConsolidateData[cLabel][cNewLabel]
                iIndex += 1
        return aRowData
    except Exception as e:
        LogService.log('Error : FunDCT_PrepareRowData Error while parsing file due to ' + str(e))

def copy_sheets(source_wb:openpyxl, dest_wb:openpyxl)->None:

  for sheet in source_wb.worksheets:
    new_sheet = dest_wb.create_sheet(sheet.title)  # Create sheet with same title
    
    for row in sheet.iter_rows():
      for cell in row:
        if cell:
            if cell.value:
                new_sheet.cell(cell.row, cell.column).value = cell.value
            # To copy formatting (optional)
            if cell.fill:
                new_sheet.cell(cell.row, cell.column).fill = cell.fill
            if cell.font:
                new_sheet.cell(cell.row, cell.column).font = cell.font
            if cell.border:
                new_sheet.cell(cell.row, cell.column).border = cell.border
            # if cell.is_date:
            #     new_sheet.cell(cell.row, cell.column).is_date = cell.is_date
            if cell.number_format:
                new_sheet.cell(cell.row, cell.column).number_format = cell.number_format
            if cell.alignment:
                new_sheet.cell(cell.row, cell.column).alignment = cell.alignment
            # if cell.data_type:
            #     new_sheet.cell(cell.row, cell.column).data_type = cell.data_type
