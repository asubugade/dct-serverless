import shutil
from pathlib import Path
import sys
import json
import pandas
from openpyxl.reader.excel import load_workbook
import openpyxl
import os
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors
from openpyxl import styles
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import xlsxwriter
import pydash
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode
from anytree.importer import JsonImporter
from anytree.importer import DictImporter
from anytree.exporter import DotExporter
from anytree import AnyNode, RenderTree
import pandas.api.types as pdt
from MessageHandling import MessageHandling
import CommonValidations
import datetime
import xml.etree.cElementTree as ET
from collections import defaultdict

from datetime import datetime

from pyconfig.LogService import LogService
oNow = datetime.now()

import re
from pyconfig import loadConfig
from pyconfig.model_common import model_common_Cls 
from pyconfig import awsCloudFileManager
import prefillNomemberupload
from shutil import copyfile
import glob

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

sys.path.insert(0, os.path.abspath(__file__))

# Explore Parameteres
oPostParameters = json.loads(sys.argv[1])
# with open(r'c:\vartemp\consolidate_CASE_Compenent_missMatch1_MEBNER.txt','w') as f:
#     f.write(str(sys.argv[1]))
# with open(r'c:\vartemp\consolidate_CASE_Compenent_missMatch1_all.txt','r') as f:
#     oPostParameters = json.loads(f.read())
iTemplateID = oPostParameters['iTemplateID']
model_common = model_common_Cls()
MessageHandling = MessageHandling()
oDataFrameList = model_common.FunDCT_GetTemplateLog('gen_templates',iTemplateID)
cTemplateType = oDataFrameList[0]['cTemplateType']
cFileDir = oPostParameters['cDirFile']
cFileType = '.xlsx'
cRequestType = oPostParameters['cRequestType'] # Either ALL / MEMEBER
aUnMappedColumnLabels = defaultdict(dict)
aUnMappedColumnRange = defaultdict(dict)
aUnMappedColumnRangeDetails = defaultdict(dict)
aRangeStart = []
aRangeEnd = []

aRangeMappedStart = []
aRangeMappedEnd = []
aMappedColumnRangeDetails = defaultdict(dict)
aMappedColumnsDetail = defaultdict(dict)

aUniqueIDsRow = defaultdict(dict)

cSamplfilepath = str(cFileDir) +  'Template_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,oDataFrameList[0]['cTemplateSampleFile'],cSamplfilepath)
oWorkbookTemplate = load_workbook(cSamplfilepath)
oSheetTemplate = oWorkbookTemplate.active
iMaxRowTemplate = oSheetTemplate.max_row
iMaxColumnTemplate = oSheetTemplate.max_column


aHeaderDetails = model_common.FunDCT_GET_HEADER_VALIDATION(iTemplateID)
validations = model_common.FunDCT_GET_HEADER_VALIDATIONonlyChild(iTemplateID)
aGetMemberColumns = model_common.FunDCT_GetMemberScacValidationCol(iTemplateID)
aMemberScacColIndex = []
aMemberScacColLabel = defaultdict(dict)
mappingSCACDict =  CommonValidations.getSCAC_columnMapping(validations)

def FunDCT_GetUnMappednColumnDetailsConsolidation(_cCompanyname):
    try:
        aRangeStart = []
        aRangeEnd = []
        aGetUnMappedColumnLabels = model_common.FunDCT_GetUnMappednColumns(iTemplateID)
        for cMemberColumnKey, aMemberColumnDetails in enumerate(aGetUnMappedColumnLabels):
            if aMemberColumnDetails.get('aUnMappedColumns'):
                for iIndex in range(len(aMemberColumnDetails['aUnMappedColumns'])):
                    if _cCompanyname in aMemberColumnDetails['aUnMappedColumns'].keys():
                        for cKey, cVal in aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0].items():
                            iRangeStart = int(aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0][cKey]['iActualCellCol'])
                            iTotalLen = int(aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0][cKey]['iTotalLen'])
                            iRangeEnd = int(iRangeStart) + int(iTotalLen-1)
                            
                            aRangeStart.append(iRangeStart)
                            aRangeEnd.append(int(iRangeStart) + int(iTotalLen)-1)
                            aUnMappedColumnRange[iRangeStart] = {
                                'iRangeStart': iRangeStart,
                                'iRangeEnd': iRangeEnd,
                                'iTotalLen': iTotalLen,
                            }
                            aUnMappedColumnRangeDetails[str(iRangeStart)+"-"+str(iRangeEnd)] = {
                                'iRangeStart': iRangeStart,
                                'iRangeEnd': iRangeEnd,
                                'iTotalLen': iTotalLen,
                            }
                            aUnMappedColumnLabels[_cCompanyname] = {
                                '_cCompanyname': aMemberColumnDetails['aUnMappedColumns'][_cCompanyname],
                            }
            if aMemberColumnDetails.get('aMappedColumns'):
                for iIndex in range(len(aMemberColumnDetails['aMappedColumns'])):
                    if _cCompanyname in aMemberColumnDetails['aMappedColumns'].keys():
                        for cKey, cVal in aMemberColumnDetails['aMappedColumns'][_cCompanyname][0].items():
                            iRangeStartMapped = int(aMemberColumnDetails['aMappedColumns'][_cCompanyname][0][cKey]['iActualCellCol'])
                            iTotalLenMapped = int(aMemberColumnDetails['aMappedColumns'][_cCompanyname][0][cKey]['iTotalLen'])
                            aMappedColumnRangeDetails[_cCompanyname][cKey] = {
                                'iRangeStart' : iRangeStartMapped,
                                'iTotalLen' : iTotalLenMapped,
                                'iRangeEnd' : int(iRangeStartMapped) + int(iTotalLenMapped-1)
                            }
        return aRangeStart,aRangeEnd
    except Exception as e:
        LogService.log('Error : FunDCT_GetUnMappednColumnDetailsConsolidation Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetUnMappednColumnDetailsConsolidation Error while parsing file due to ' + str(e))

def FunDCT_GetNextColumnConsolidate(jColumn, iMaxColumnUploaded, aRangeStart, aRangeEnd):
    try:
        iSkipRows = 0
        for iLowerRange, iUpperRange in zip(aRangeStart, aRangeEnd):
            if iLowerRange <= jColumn <= iUpperRange:
                iSkipRows += aUnMappedColumnRangeDetails[str(iLowerRange)+"-"+str(iUpperRange)]['iTotalLen']
                jColumn += iSkipRows
                if jColumn <= iMaxColumnUploaded+1:
                    FunDCT_GetNextColumnConsolidate(jColumn, iMaxColumnUploaded, aRangeStart, aRangeEnd)
        return iSkipRows
    except Exception as e:
        LogService.log('Error : FunDCT_GetNextColumnConsolidate Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetNextColumnConsolidate Error while parsing file due to ' + str(e))

def FunDCT_PrepareRowData(aConsolidateData):
    try:
        aRowData = defaultdict(dict)
        for cLabel in aConsolidateData:
            iIndex = 1
            for cNewLabel in aConsolidateData[cLabel]:
                aRowData[iIndex][len(aRowData[iIndex])+1] = aConsolidateData[cLabel][cNewLabel]
                iIndex += 1
        return aRowData
    except Exception as e:
        LogService.log('Error : FunDCT_PrepareRowData Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_PrepareRowData Error while parsing file due to ' + str(e))

def FunDCT_GetMappedCol(aTemplateHeader, aRowData, aMemberScac):
    indexValue = list(aRowData.keys())[list(aRowData.values()).index(aMemberScac)]
    indexValue = indexValue -1 if indexValue else indexValue
    headersWith = aTemplateHeader[indexValue]
    headerWithValidation = headersWith.get('cValidations')
    mappinheaderValidation = ''
    if 'DEFAULT_ORIGIN_SCAC' in headerWithValidation:
        mappinheaderValidation = 'ORIGIN_COLUMN'
    elif 'DEFAULT_DESTINATION_SCAC' in headerWithValidation:
        mappinheaderValidation = 'DESTINATION_COLUMN'
    mappedListColumn = list()
    iCol = 1
    for eachHeader in aTemplateHeader:
        if mappinheaderValidation in  eachHeader.get('cValidations'):
            mappedListColumn.append(iCol)
        iCol+=1

    return mappedListColumn
def FunDCT_GetMappedColumnsDetail(aRowConsolidateData, cColumnKey, cMemberScac,OFR='',FOB='',PLC=''):
    try:
        aScacAvailableInColumn = defaultdict(dict)
        for cRowKey in aRowConsolidateData[cColumnKey]:
            if aRowConsolidateData[cColumnKey][cRowKey]:
                if str(aRowConsolidateData[cColumnKey][cRowKey]).strip().upper() == str(cMemberScac).strip().upper():
                    cHeaderVal = aRowConsolidateData[1][cRowKey]
                    if cHeaderVal == OFR or cHeaderVal == FOB or cHeaderVal == PLC:
                        aScacAvailableInColumn[cColumnKey][cHeaderVal] = cHeaderVal
        return aScacAvailableInColumn
    except Exception as e:
        LogService.log('Error : FunDCT_GetMappedColumnsDetail Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetMappedColumnsDetail Error while parsing file due to ' + str(e))

def FunDCT_InsertDefaultHeader():
    try:
        cConsolidateTemplateFileName = str(cFileDir) + 'CONSOLIDATE-' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
        os.remove(cConsolidateTemplateFileName) if os.path.exists(cConsolidateTemplateFileName) else None
        cSampleFile = Path(cConsolidateTemplateFileName)
        
        if not cSampleFile.is_file():
            oDataFrameList = model_common.FunDCT_GetTemplateLog('gen_templates',iTemplateID)
            oDataFrameSampleFile = pandas.DataFrame(eval(oDataFrameList[0]['cUploadedFile']))
            oDataFrameSampleFile.columns = oDataFrameSampleFile.columns.str.replace('#DCT#.*', '')
            oDataFrameSampleFile.to_excel(cConsolidateTemplateFileName, index=False)
        return cConsolidateTemplateFileName
    except Exception as e:
        LogService.log('Error : FunDCT_InsertDefaultHeader Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_InsertDefaultHeader Error while parsing file due to ' + str(e))

def FunDCT_SetMappedColRange(aMappedColumnsDetail, cColumnKey, cMemberScac):
    try:
        aRangeMappedStart = []
        aRangeMappedEnd = []
        for cMappedColumn in aMappedColumnsDetail[cColumnKey]:
            if cMappedColumn == 'FOB':
                iStartFromRange = int(aMappedColumnRangeDetails[cMemberScac]['ORIGIN CHARGES']['iRangeStart'])
                iStartToRange = int(aMappedColumnRangeDetails[cMemberScac]['ORIGIN CHARGES']['iRangeEnd'])
            elif cMappedColumn == 'OFR':
                iStartFromRange = aMappedColumnRangeDetails[cMemberScac]['OCEAN CHARGES']['iRangeStart']
                iStartToRange = aMappedColumnRangeDetails[cMemberScac]['OCEAN CHARGES']['iRangeEnd']
            elif cMappedColumn == 'PLC':
                iStartFromRange = aMappedColumnRangeDetails[cMemberScac]['DESTINATION CHARGES']['iRangeStart']
                iStartToRange = aMappedColumnRangeDetails[cMemberScac]['DESTINATION CHARGES']['iRangeEnd']
            aRangeMappedStart.append(iStartFromRange)
            aRangeMappedEnd.append(iStartToRange)
        
        return aRangeMappedStart, aRangeMappedEnd
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : FunDCT_SetMappedColRange Error while parsing file due to ' + str(e) + ' Line no = '+ str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_SetMappedColRange Error while parsing file due to ' + str(e) + ' Line no = '+ str(line))

def FunDCT_FillRowOfUniqueID(cUniqueID, iFindRow):
    try:
        aUniqueIDsRow[str(cUniqueID)] = {
            'iFindRow':iFindRow
        }
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : FunDCT_FillRowOfUniqueID Error while parsing file due to ' + str(e) + ' Line no ' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_FillRowOfUniqueID Error while parsing file due to ' + str(e) + ' Line no ' + str(line))

def FunDCT_InsertNewRow(iMaxDepthHeaders, iMaxColumnTemplate, oWorksheet, cMemberScac, aLaneIDs, aRowConsolidateData, cKeyUniqueIDSequenceFind, aUniqueIDsRow):
    try:
        aRangeStart,aRangeEnd = FunDCT_GetUnMappednColumnDetailsConsolidation(cMemberScac)
        aKeys = []
        aUnMappedColumnRangeList = [(aRangeStart[iRangeIndex], aRangeEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeStart))]
        bInUnmappedRange = True
        aKeys = aUnMappedColumnLabels[cMemberScac]['_cCompanyname'][0].keys() if len(aUnMappedColumnLabels[cMemberScac]) > 0 else []
        iSubsctractCols = 0
        iMergedCellLen = 0
        jColumn = 1
        iRow = oWorksheet.max_row + 1
        cKeyUniqueID = 0
        cKeyUniqueID = cKeyUniqueIDSequenceFind
        MappingSCAC = CommonValidations.getSCAC_columnMapping(aHeaderDetails)
        FOB,PLC,OFR = ('','','')
        
        try:

            FOB = aHeaderDetails[MappingSCAC.get('DEFAULT_ORIGIN_SCAC')]['cHeaderLabel']
            PLC = aHeaderDetails[MappingSCAC.get('DEFAULT_DESTINATION_SCAC')]['cHeaderLabel']
            OFR = aHeaderDetails[MappingSCAC.get('DEFAULT_OCEAN_SCAC')]['cHeaderLabel']
        except:
            pass
        iRow = oWorksheet.max_row + 1
        for cColumnKey in aRowConsolidateData:
            if cColumnKey < iMaxDepthHeaders:
                continue
            bSkip = False
            iSubsctractCols = 0
            jColumn = 1

            aMappedColumnsDetail = defaultdict(dict)
            aMappedColumnsDetail = FunDCT_GetMappedColumnsDetail(aRowConsolidateData, cColumnKey, cMemberScac,FOB,OFR,PLC)
            
            aRangeMappedStart,aRangeMappedEnd = FunDCT_SetMappedColRange(aMappedColumnsDetail, cColumnKey, cMemberScac)
            aMappedColumnRangeList = [(aRangeMappedStart[iRangeIndex], aRangeMappedEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeMappedStart))]
            iMergedCellLen = 0
            for cVal in aRowConsolidateData[cColumnKey]:
                if len(aKeys) > 0:
                    
                    bInUnmappedRange = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aUnMappedColumnRangeList)
                    binMappedRange  = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aMappedColumnRangeList)
                    if bInUnmappedRange == True:
                        if FunDCT_GetNextColumnConsolidate(jColumn-iMergedCellLen,iMaxColumnTemplate, aRangeStart, aRangeEnd) > iMergedCellLen:
                                iMergedCellLen = FunDCT_GetNextColumnConsolidate(jColumn-iMergedCellLen,iMaxColumnTemplate, aRangeStart, aRangeEnd)
                        # iMergedCellLen = FunDCT_GetNextColumnConsolidate(jColumn, iMaxColumnTemplate, aRangeStart, aRangeEnd)
                    else:
                        if binMappedRange:
                            ccd1 = FunDCT_GetNextColumnConsolidate(jColumn+iSubsctractCols,iMaxColumnTemplate, aRangeStart, aRangeEnd)
                            ccd = FunDCT_GetNextColumnConsolidate(jColumn-iSubsctractCols,iMaxColumnTemplate, aRangeStart, aRangeEnd)
                            if ccd1 > 0 and ccd > 0:
                                iSubsctractCols = ccd1 + ccd
                if (bInUnmappedRange is True) and iMergedCellLen > 0:
                    iSubsctractCols = iMergedCellLen
                elif (bInUnmappedRange is True) and iMergedCellLen <= 0:
                    iSubsctractCols = iMergedCellLen
                                
                iForwardValidator = int(int(jColumn) + int(iSubsctractCols))

                dValidations = validations[iForwardValidator-1]
                conditinalValidation = dValidations.get('cValidationsCondition')
                sSCACname = ''
                if 'ORIGIN_COLUMN' in conditinalValidation:
                    sSCACname = 'DEFAULT_ORIGIN_SCAC' 
                elif 'OCEAN_COLUMN' in conditinalValidation:
                    sSCACname = 'DEFAULT_OCEAN_SCAC' 
                elif 'DESTINATION_COLUMN' in conditinalValidation:
                    sSCACname = 'DEFAULT_DESTINATION_SCAC' 
                elif 'SERVICEINFO_COLUMN' in conditinalValidation:
                    sSCACname ='SERVICEINFO_HEADER'
                bLiable = None
                if mappingSCACDict.get(sSCACname) is not None:
                    try:
                        if mappingSCACDict.get(sSCACname):
                            scacint = int(mappingSCACDict.get(sSCACname)) + 1
                            liablitySCAC = aRowConsolidateData[cColumnKey][scacint]
                            bLiable = liablitySCAC == cMemberScac

                    except:
                        pass
                
                if str(aRowConsolidateData[cColumnKey][cKeyUniqueID]) not in aLaneIDs:
                    valec  = aRowConsolidateData[cColumnKey][cVal]
                    if str(valec).lower().startswith('='):
                        formula = valec
                        matches = re.findall(r"([A-Z]+[0-9]+)\b",formula)
                        for mtch in matches:
                            rownumber = re.search(r'([0-9]+)\b',str(mtch)).group(1)
                            newR = mtch.replace(rownumber,str(iRow))
                            valec = valec.replace(mtch,newR) 
                    if bLiable is None or bLiable is True :
                        oWorksheet.cell(iRow, iForwardValidator).value = valec
                    if jColumn == 1 :
                        FunDCT_FillRowOfUniqueID(aRowConsolidateData[cColumnKey][cVal], iRow)
                    jColumn += 1
                        
                elif str(aRowConsolidateData[cColumnKey][cKeyUniqueID]) in aLaneIDs:
                    # aMappedColumnsDetail = defaultdict(dict)
                    # aMappedColumnsDetail = FunDCT_GetMappedColumnsDetail(aRowConsolidateData, cColumnKey, cMemberScac,FOB,OFR,PLC)
                    # aRangeMappedStart,aRangeMappedEnd = FunDCT_SetMappedColRange(aMappedColumnsDetail, cColumnKey, cMemberScac)
                    # aMappedColumnRangeList = [(aRangeMappedStart[iRangeIndex], aRangeMappedEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeMappedStart))]
                    # bInMappedRange = any(iLowerRange <= iForwardValidator <= iUpperRange for (iLowerRange, iUpperRange) in aMappedColumnRangeList)
                    
                    # if cTemplateType != 'RFQTool':
                    #     aMappedColumnRangeList = FunDCT_GetMappedCol(aHeaderDetails,aRowConsolidateData[cColumnKey],cMemberScac)
                    #     bInMappedRange = True if jColumn in aMappedColumnRangeList else False
                    bInMappedRange = any(iLowerRange <= iForwardValidator <= iUpperRange for (iLowerRange, iUpperRange) in aMappedColumnRangeList)
                    iFindRow = 0
                    if bInMappedRange == True:
                        bSkip = True                        
                        iFindRow = 0
                        iFindRow = int(aUniqueIDsRow[str(aRowConsolidateData[cColumnKey][cKeyUniqueID])]['iFindRow'])
                        valec  =aRowConsolidateData[cColumnKey][cVal]
                        if str(valec).lower().startswith('='):
                            formula = valec
                            matches = re.findall(r"([A-Z]+[0-9]+)\b",formula)
                            for mtch in matches:
                                rownumber = re.search(r'([0-9]+)\b',str(mtch)).group(1)
                                newR = mtch.replace(rownumber,str(iRow))
                                valec = valec.replace(mtch,newR) 
                        if bLiable is None or bLiable :
                            oWorksheet.cell(iFindRow, iForwardValidator).value = valec
                        
                        if jColumn == 1 :
                            FunDCT_FillRowOfUniqueID(aRowConsolidateData[cColumnKey][cVal], iRow)
                        jColumn += 1
                    else:
                        iFindRow = 0
                        iFindRow = int(aUniqueIDsRow[str(aRowConsolidateData[cColumnKey][cKeyUniqueID])]['iFindRow'])
                        valec  =aRowConsolidateData[cColumnKey][cVal]
                        if str(valec).lower().startswith('='):
                            formula = valec
                            matches = re.findall(r"([A-Z]+[0-9]+)\b",formula)
                            for mtch in matches:
                                rownumber = re.search(r'([0-9]+)\b',str(mtch)).group(1)
                                newR = mtch.replace(rownumber,str(iRow))
                                valec = valec.replace(mtch,newR) 
                        if bLiable is None or bLiable :
                            oWorksheet.cell(iFindRow, iForwardValidator).value = valec
                        jColumn += 1
                         
                       
            if not bSkip : iRow += 1
        return aUniqueIDsRow
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : FunDCT_InsertNewRow Error while parsing file due to ' + str(e) + ' Line no ' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_InsertNewRow Error while parsing file due to ' + str(e) + ' Line no ' + str(line))

def FunDCT_LoadMemberDistributedData(cScac, iTemplateID):
    LogService.log(f'{iTemplateID} , {cScac}')
    objeyKey = model_common.FunDCT_GetMemeberDistributedFile(iTemplateID,cScac)
    LogService.log(f'Fetching file => {objeyKey}')
    # objeyKey = f'{loadConfig.AWS_BUCKET_ENV}/{loadConfig.AWS_BUCKET_TEMPLATES}/{loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE}/{loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_INTERTEAM}/{loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_STATUSFILE}/{cScac}_{iTemplateID}.xlsx'
    if objeyKey  == '':
        return ''
    try:
        cTempfilePath = str(cFileDir) + cScac +'_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
        awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,objeyKey,cTempfilePath)
        
        loadwb = load_workbook(cTempfilePath,read_only=True)
        odataframe = pandas.ExcelFile(loadwb,engine="openpyxl").parse()
        odataframe = odataframe.dropna(axis=1, how='all')
        
        if iMaxRowTemplate > 1 :
            col = ''
            ii = 0
            for x in odataframe.iloc[0]:

                if x:
                    if 'date' in str(x) or 'Date' in str(x):
                        col = list(odataframe.columns)[ii]                        
                        odataframe[col] = odataframe[col].apply(dateconc)
                        # odataframe[col] = odataframe[col].dt.strftime('%Y-%m-%d %M:%H:%S')

                ii+=1
        else:
            for x in odataframe:
                if pdt.is_datetime64_any_dtype(odataframe[x]):
                        odataframe[x] = odataframe[x].dt.strftime('%Y-%m-%d %M:%H:%S')

        
        # odataframe = pandas.read_excel(oData, engine='openpyxl',dtype = str,na_values=str,)
        iIndexDataFrame = iter(range(1, len(odataframe.columns) + 1))
        odataframe.columns = [cColumns if cColumns.startswith('Unnamed') is not None else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in odataframe.columns]
        jData = odataframe.to_json()
        # loadwb.close()
        if os.path.exists(cTempfilePath): os.remove(cTempfilePath)
    except Exception as E:
        LogService.log(f'Errot while fetching data from file ==> {objeyKey}')
        return ''
    return jData
def dateconc(val):
    try:
        return val.strftime('%Y-%m-%d %H:%M:%S')
    except:
        return val
def FunDCT_LoadJsonByObjectKey(s3objKey):
    LogService.log(f'Fetching file => {s3objKey}')
    if s3objKey  == '':
        return ''
    try:
        filetype = s3objKey.split('.')[-1]
        if filetype.startswith('xls'):
            cTempfilePath = str(cFileDir) + 'Temple_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
            awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,s3objKey,cTempfilePath)
        
            loadwb = load_workbook(cTempfilePath,read_only=True)
            odataframe = pandas.ExcelFile(loadwb,engine="openpyxl").parse(na_filter=True,)
            odataframe = odataframe.dropna(axis=1, how='all')
            if iMaxRowTemplate > 1 :
                col = ''
                ii = 0
                for x in odataframe.iloc[0]:

                    if x:
                        if 'date' in str(x) or 'Date' in str(x):
                            col = list(odataframe.columns)[ii]                        
                            odataframe[col] = odataframe[col].apply(dateconc)
                            # odataframe[col] = odataframe[col].dt.strftime('%Y-%m-%d %M:%H:%S')

                    ii+=1
            else:
                for x in odataframe:
                    if pdt.is_datetime64_any_dtype(odataframe[x]):
                            odataframe[x] = odataframe[x].dt.strftime('%Y-%m-%d %M:%H:%S')
                # oData = awsCloudFileManager.getfile(loadConfig.AWS_BUCKET,s3objKey)
                # odataframe = pandas.read_excel(oData, engine='openpyxl',dtype = str)
            iIndexDataFrame = iter(range(1, len(odataframe.columns) + 1))
            odataframe.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in odataframe.columns]
            jData = odataframe.to_json()
            if os.path.exists(cTempfilePath): os.remove(cTempfilePath)
        else:

            oData = awsCloudFileManager.download_data_from_bucket(loadConfig.AWS_BUCKET,s3objKey)
            oData =str(oData).replace('null','\\"\\"')
            jsoLod = json.loads(oData)
            odataframe = pandas.DataFrame(eval(jsoLod),dtype=str,)
            iIndexDataFrame = iter(range(1, len(odataframe.columns) + 1))
            odataframe.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in odataframe.columns]
            jData = odataframe.to_json()
        return jData
    except Exception as E:
        LogService.log(f'Errot while fetching data from file ==> {s3objKey}')
        return ''


def FunDCT_GapFilling1(oDistributedMeta, oWorksheet):
    iMaxDepthHeaders = model_common.FunDCT_GetMaxDepthHeader(iTemplateID)
    cTempfilePathUploaded = str(cFileDir) + 'UploadedFile_SampleFile' + str(iTemplateID) + "_" + datetime.strftime(oNow,'%Y%m%d%H%M%S%f') + '.xlsx'
    awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,oDistributedMeta[0]['aDistributedData'],cTempfilePathUploaded)
    wb =load_workbook(cTempfilePathUploaded,read_only=True)
    wbs = wb.active
    # iMaxr = wbs.max_row
    # iMaxCr = wbs.max_column
    iRowc = 1
    ft = Font(color="FF0000")
    for ii,row in enumerate(wbs.iter_rows(values_only=True),1):
        if ii <= iMaxDepthHeaders:
            iRowc += 1
            continue
        for k,cln in enumerate(row):
            if str(cln).lower().startswith('='):
                oWorksheet.cell(iRowc,k+1).value = cln
            elif oWorksheet.cell(iRowc,k+1).value is None or oWorksheet.cell(iRowc,k+1).value  == '':
                oWorksheet.cell(iRowc,k+1).value = cln
                oWorksheet.cell(iRowc,k+1).font = ft
            elif str(cln).strip().upper() != str(oWorksheet.cell(iRowc,k+1).value).strip().upper():
                oWorksheet.cell(iRowc,k+1).font = ft
        iRowc += 1
    wb.close()
    if os.path.exists(cTempfilePathUploaded): os.remove(cTempfilePathUploaded)


def FunDCT_GapFilling(oDistributedMeta, oWorksheet):
    iMaxDepthHeaders = model_common.FunDCT_GetMaxDepthHeader(iTemplateID)
    bytesData = awsCloudFileManager.getfile(loadConfig.AWS_BUCKET,oDistributedMeta[0]['aDistributedData'])
    oPadasExcel = pandas.ExcelFile(bytesData,'openpyxl')
    oDataFrame = oPadasExcel.parse(na_values = "Missing", na_filter=True, skiprows = iMaxDepthHeaders - 1)
    # oDataFrame = pandas.read_excel(wb,skiprows = iMaxDepthHeaders - 1)
    oJson = oDataFrame.to_json()
    oDistributedMetaData = json.loads(oJson)
    joDistributedMetaData=FunDCT_PrepareRowData(oDistributedMetaData)
    ft = Font(color="FF0000")
    pendingRows=list()
    for index in joDistributedMetaData:
        if index <= 1:
            continue

        iFindRow = 0
        try:
            iFindRow = int(aUniqueIDsRow[str(joDistributedMetaData[index][1])]['iFindRow'])
        except:
            pass
        if iFindRow == 0 :
            pendingRows.append(joDistributedMetaData[index])
        else:
            for cellindex in range(1,len(joDistributedMetaData[index])):
                if oWorksheet.cell(iFindRow,cellindex).value is None or oWorksheet.cell(iFindRow,cellindex).value  == '':
                    oWorksheet.cell(iFindRow,cellindex).value = str(joDistributedMetaData[index][cellindex])
                elif str(joDistributedMetaData[index][cellindex]) != str(oWorksheet.cell(iFindRow,cellindex).value):
                    oWorksheet.cell(iFindRow,cellindex).font = ft
    if len(pendingRows)  > 0:
        for inx, eachRowData in enumerate(pendingRows):
            valueList = list(eachRowData.values())
            oWorksheet.append(valueList)

def FunDCT_GetTextColor(oCell):
    try:
        cTextColor = '000000'
        if oCell.font.color:
            if 'rgb' in oCell.font.color.__dict__:
                cTextColor = oCell.font.color.rgb[2:]          

        return cTextColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))

def FunDCT_GetBackColor(oCell, cTextColor):
    try:
        cBackColor = 'ffffff'
        
        if oCell.fill.start_color != '' and type(oCell.fill.start_color.index) != int:                    
            cBackColor = oCell.fill.start_color.index[2:]
            if cTextColor == '000000' and cBackColor == '000000':
                cBackColor = 'ffffff'
        elif oCell.fill.start_color != '' and type(oCell.fill.start_color.index) == int:
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index][2:])
        elif isinstance(oCell.fill, PatternFill) and oCell.fill.patternType == 'solid':
            if oCell.fill.fgColor.type =='rgb':
                cBackColor = oCell.fill.fgColor.rgb[2:]
            elif oCell.fill.fgColor.type =='indexed':
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index -1][2:])
        
        return cBackColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))

def FunDCT_GetMergeCellRangeLen(oSheetTemplate, iRow, jColumn):
    try:
        iTotalLen = 0
        oCell = oSheetTemplate.cell(iRow, jColumn)
        for mergedCell in oSheetTemplate.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                cCellRangeStart, cCellRangeEnd = str(mergedCell).split(':')
                iCellCol = CommonValidations.FunDCT_GetColumnNumber(
                    ''.join([i for i in cCellRangeStart if not i.isdigit()]))
                iTotalLen = int(CommonValidations.FunDCT_GetColumnNumber(
                    ''.join([i for i in cCellRangeEnd if not i.isdigit()]))) - (int(iCellCol)-1)
        return iTotalLen
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetMergeCellRangeLen Error while parsing file due to ' + str(e))


    
def FunDCT_MergeAllTemplateFile():
    try:
        if(str(oPostParameters['preFillData']).lower() == 'yes'.lower()):
            cConsolidateTemplateFileName = FunDCT_InsertDefaultHeader()
            oConsolidationTemplateLog = model_common.FunDCT_GetConsolidationData(iTemplateID)
            oDistributedMeta = model_common.FunDCT_GetDistributionData(iTemplateID)
            if cRequestType == 'ALL':
                
                for cMemberColumnKey, aMemberColumnDetails in enumerate(oConsolidationTemplateLog):
                    for cMemberScac in oConsolidationTemplateLog[cMemberColumnKey]['aConsolidationData'].keys():
                        # if 'aConsolidationData' in oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]:
                            # cConsolidationData = oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData']
                            cConsolidationData = oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]
                            cConsolidationData2 = cConsolidationData.get('aConsolidationData')
                            if cConsolidationData2 == '' or cConsolidationData2 is None:
                                oDistributedOriginalData = FunDCT_LoadMemberDistributedData(cMemberScac,iTemplateID)
                                if oDistributedOriginalData == '':
                                    continue
                                oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData'] = oDistributedOriginalData
                            else:
                                oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData'] = FunDCT_LoadJsonByObjectKey(cConsolidationData2)
            else:
                for cMemberColumnKey, aMemberColumnDetails in enumerate(oConsolidationTemplateLog):
                    for cMemberScac in oConsolidationTemplateLog[cMemberColumnKey]['aConsolidationData'].keys():
                            cConsolidationData = oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]
                            cConsolidationData2 = cConsolidationData.get('aConsolidationData')
                            if cConsolidationData2:
                                oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData'] = FunDCT_LoadJsonByObjectKey(cConsolidationData2)

                        

            oWorkbook = load_workbook(cConsolidateTemplateFileName)
            oWorksheet = oWorkbook.active
            iMaxColumnTemplate = oWorksheet.max_column + 1
            iMaxDepthHeaders = model_common.FunDCT_GetMaxDepthHeader(iTemplateID)
            cColumnLetter1 = f"{get_column_letter(2)}1:{get_column_letter(oWorksheet.max_column + 1)}1" 
            iii = 1
            jjj = 1
            # for eachCell in oWorksheet[cColumnLetter1]:
            #     for nCel in eachCell:
            #         jjj+=1
            #         if nCel.value is not None:
            #             ranges = f"{get_column_letter(iii)}1:{get_column_letter(jjj-1)}1"
            #             oWorksheet.merge_cells(ranges)
            #             nCel.alignment =Alignment(horizontal="center", vertical="center")
            #             iii=jjj
            cColumnLetter = f"{get_column_letter(1)}1:{get_column_letter(oWorksheet.max_column)}{iMaxDepthHeaders}"
            for rows in oWorksheet[cColumnLetter]:
                for nCell in rows:
                    thin_border = Side(style="thin")
                    nCell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                    nCell.font = Font(bold=True)

            oWorkbook.save(cConsolidateTemplateFileName)
            cKeyUniqueIDSequenceFind = CommonValidations.FunDCT_GetUniqueIDSequence(oWorksheet, cTemplateType)
            oWorkbook = load_workbook(cConsolidateTemplateFileName)
            oWorksheet = oWorkbook.active
            cKeyUniqueIDSequence = cKeyUniqueIDSequenceFind
            if cKeyUniqueIDSequence != 'NA' and cKeyUniqueIDSequence >= 0:
                for iRow in range(1, oWorksheet.max_row + 1):
                    if iRow <= iMaxDepthHeaders:
                        continue
            
            iIndexseq = 0
            for cMemberColumnKey, aMemberColumnDetails in enumerate(oConsolidationTemplateLog):
                for cMemberScac in oConsolidationTemplateLog[cMemberColumnKey]['aConsolidationData'].keys():
                    if 'aConsolidationData' in oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]:
                        cConsolidationData = oConsolidationTemplateLog[0]['aConsolidationData'][cMemberScac][0]['aConsolidationData']
                        if cConsolidationData != '':
                            aLaneIDs = []
                            aRowConsolidateData = []
                            if iIndexseq <= 0:
                                # oDistributedOriginalData = FunDCT_LoadMemberDistributedData(cMemberScac,iTemplateID)

                                aConsolidateData = json.loads(cConsolidationData)
                                aRowConsolidateData = FunDCT_PrepareRowData(aConsolidateData)
                                FunDCT_InsertNewRow(iMaxDepthHeaders, iMaxColumnTemplate, oWorksheet, cMemberScac, aLaneIDs, aRowConsolidateData, cKeyUniqueIDSequenceFind, aUniqueIDsRow)
                                iIndexseq += 1
                            elif iIndexseq > 0:
                                iRow = 1
                                for aRows in oWorksheet.rows:
                                    if iRow > iMaxDepthHeaders:
                                        aLaneIDs.append(str(aRows[0].value))
                                    iRow += 1
                                aLaneIDsToSkipSeq = []
                                aConsolidateData = json.loads(cConsolidationData)
                                if cTemplateType == 'RFQTool':
                                    for cLabel in aConsolidateData['QUOTING MEMBER']:
                                        if str(aConsolidateData['QUOTING MEMBER'][cLabel]) in aLaneIDs :
                                            aLaneIDsToSkipSeq.append(str(aConsolidateData['QUOTING MEMBER'][cLabel]))
                                else:
                                    for cLabel in aConsolidateData[str(next(iter(aConsolidateData)))]:
                                        if str(aConsolidateData[str(next(iter(aConsolidateData)))][cLabel]) in aLaneIDs :
                                            aLaneIDsToSkipSeq.append(str(aConsolidateData[next(iter(aConsolidateData))][cLabel]))

                                aRowConsolidateData = FunDCT_PrepareRowData(aConsolidateData)
                                
                                FunDCT_InsertNewRow(iMaxDepthHeaders, iMaxColumnTemplate, oWorksheet, cMemberScac, aLaneIDsToSkipSeq, aRowConsolidateData, cKeyUniqueIDSequenceFind, aUniqueIDsRow)
                                iIndexseq += 1
            
            # if cRequestType == 'ALL': 
            #     FunDCT_GapFilling(oDistributedMeta,oWorksheet)
            
            # oWorkbook.save(cConsolidateTemplateFileName)
            df = pandas.ExcelFile(oWorkbook,engine='openpyxl')
            tdf = df.parse(skiprows=iMaxDepthHeaders-1,na_values = "Mssing",na_filter=True,date_parser=True)
            for x in tdf:
                if pdt.is_datetime64_any_dtype(tdf[x]):
                    tdf[x] = tdf[x].dt.strftime('%Y-%m-%d %M:%H:%S')
            # tdf = pandas.DataFrame(oWorksheet.values)
            colname = list(tdf.columns)[0]
            for cl in list(tdf.columns):
                if 'Unnamed' in str(cl):
                    tdf.drop(columns=[str(cl)],inplace=True)
            tdf.sort_values(by=[str(colname)],inplace=True,ignore_index=True)
            # oWorkbook.save(cConsolidateTemplateFileName)
            tdf.to_excel(cConsolidateTemplateFileName,startrow=iMaxDepthHeaders-1,index=False)
            oWorkbook = load_workbook(cConsolidateTemplateFileName)
            oBookSheet = oWorkbook.active
            #######Onlly header Mapping #####################
            aKeys = []
            bColumnHeaderNotInRange = False
            bInUnmappedRange = False
            # if len(aUnMappedColumnLabels) > 0:
            #     aUnMappedColumnRangeList = [(aRangeStart[iRangeIndex], aRangeEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeStart))]
            #     bInUnmappedRange = True
            #     bColumnHeaderNotInRange = True
            #     aKeys = aUnMappedColumnLabels[_cCompanyname]['_cCompanyname'][0].keys()
            
            iIndex = 0
            for iRow in range(1, iMaxRowTemplate+1):
                iSubsctractCols = 0
                for jColumn in range(1, iMaxColumnTemplate+1):
                    oSampleCell = oSheetTemplate.cell(row=iRow, column=jColumn)
                    oSampleCellVal = oSampleCell.value
                    # if len(aKeys) > 0:
                    #     bColumnHeaderNotInRange = True if str(oSampleCellVal) in aKeys else False
                    #     bInUnmappedRange = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aUnMappedColumnRangeList)
                    
                    iMergedCellLen = FunDCT_GetMergeCellRangeLen(oSheetTemplate, iRow, jColumn)
                    if (oSampleCellVal is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen > 0:
                        iSubsctractCols += iMergedCellLen
                        continue
                    elif (oSampleCellVal is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen <= 0:
                        iSubsctractCols += 1
                        continue
                    # elif (oSampleCellVal is None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True):
                    #     iSubsctractCols += 1
                    #     continue

                    # iBackwardCol = int(jColumn) - int(iSubsctractCols)
                    # oSampleCell = oSheetTemplate.cell(row=iRow, column=iBackwardCol)
                    # oSampleCellVal = oSampleCell.value
                    

                    if oSampleCellVal is not None:
                        iBackwardCol = int(jColumn) - int(iSubsctractCols)
                        oUploadedCell = oBookSheet.cell(row=iRow, column=iBackwardCol)
                        oUploadedCellVal = str(oUploadedCell.value)
                        bMatchHeaderValue = True if str(oSampleCellVal).strip() == str(oUploadedCellVal) else False

                        # if not bMatchHeaderValue:
                        #     cErrorMessage = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE','TEMPLATE_VALIDATE_HEADER_MISMATCH').replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$', str(oSampleCellVal)).replace('#$DCTVARIABLE_INCORRECT_CELL_VAL#$', str(oUploadedCellVal) )
                        #     FunDCT_PrepareTemplateStatusFile(iIndex, oUploadedCell.row, oUploadedCell.column, str(oUploadedCell.coordinate), cErrorMessage)
                        #     iIndex += 1
                        # else:
                        oBookSheet.cell(iRow, iBackwardCol).value = oSampleCell.value                       
                        oBookSheet.cell(iRow, iBackwardCol).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cTextColor = FunDCT_GetTextColor(oSampleCell)
                        cBackColor = FunDCT_GetBackColor(oSampleCell,cTextColor)
                        thin_border = Side(style="thin")
                        oBookSheet.cell(row=iRow, column=iBackwardCol).border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                        oBookSheet.cell(row=iRow, column=iBackwardCol).font = Font(color=cTextColor)
                        oBookSheet.cell(row=iRow, column=iBackwardCol).fill = PatternFill(fgColor=cBackColor,fill_type = "solid")
                        if FunDCT_CellMerged(oSheetTemplate, iRow, iBackwardCol):
                            if iMergedCellLen > 0:
                                ranges = f"{get_column_letter(iBackwardCol)}{iRow}:{get_column_letter((iBackwardCol - 1)+iMergedCellLen)}{iRow}"
                                oBookSheet.merge_cells(ranges)
                ################## Heading End


            if cRequestType == 'ALL': 
                FunDCT_GapFilling1(oDistributedMeta,oBookSheet)
            
            
            # if iMaxDepthHeaders>1:
            #     oWorkbook = load_workbook(cConsolidateTemplateFileName)
            #     oWorksheet = oWorkbook.active
            #     oDataFrameList = model_common.FunDCT_GetTemplateLog('gen_templates',iTemplateID)
            #     oDataFrameSampleFile = pandas.DataFrame(eval(oDataFrameList[0]['cUploadedFile']))
            #     oDataFrameSampleFile.columns = oDataFrameSampleFile.columns.str.replace('#DCT#.*', '')
            #     c =1
            #     for e in list(oDataFrameSampleFile.columns):
            #         oWorksheet.cell(row=1,column=c).value = str(e)
            #         thin_border = Side(style="thin")
            #         oWorksheet.cell(row=1,column=c).border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                    
            #         c +=1
            #     cColumnLetter1 = f"{get_column_letter(2)}1:{get_column_letter(oWorksheet.max_column + 1)}1" 
            #     iii = 1
            #     jjj = 1
            #     for eachCell in oWorksheet[cColumnLetter1]:
            #         for nCel in eachCell:
            #             jjj+=1
            #             if nCel.value != '':
            #                 ranges = f"{get_column_letter(iii)}1:{get_column_letter(jjj-1)}1"
            #                 oWorksheet.merge_cells(ranges)
            #                 nCel.alignment =Alignment(horizontal="left", vertical="center")
            #                 iii=jjj
            #     cColumnLetter1 = f"{get_column_letter(1)}2:{get_column_letter(oWorksheet.max_column + 1)}2"
            #     for eachCell in oWorksheet[cColumnLetter1]:
            #         for nCel in eachCell:
            #             nCel.alignment =Alignment(horizontal="left", vertical="center",wrap_text=True)


                

            oWorkbook.save(cConsolidateTemplateFileName)
            oWorkbook.close()
            oWorkbookTemplate.close()
                

            if os.path.exists(cSamplfilepath): os.remove(cSamplfilepath)

            cS3DirUploadFile = ''
            cS3DirUploadFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_CONSOLIDATION) + '/'
            cS3UrlUploadedOrg = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirUploadFile), cConsolidateTemplateFileName)

            return MessageHandling.FunDCT_MessageHandling('Success', cS3UrlUploadedOrg)
        else:
            iMaxDepthHeaders = model_common.FunDCT_GetMaxDepthHeader(iTemplateID)
            oPostParameters['iMaxDepthHeaders']=iMaxDepthHeaders
            
            prefillNomemberupload.PreFilledDataConsolidation(oPostParameters)

    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : FunDCT_MergeAllTemplateFile Error while parsing file due to ' + str(e) + 'Line no - '+ str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_MergeAllTemplateFile Error while parsing file due to ' + str(e) + 'Line no - '+ str(line))

def FunDCT_CellMerged(oSheet, iRow, jColumn):
    try:
        oCell = oSheet.cell(iRow, jColumn)
        for mergedCell in oSheet.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                return True
        return False
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CellMerged Error while parsing file due to ' + str(e))

FunDCT_MergeAllTemplateFile()
