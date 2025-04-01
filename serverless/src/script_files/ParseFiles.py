import shutil
from pathlib import Path
import sys
import inspect
import json
import datetime
import xml.etree.cElementTree as ET
from collections import defaultdict
from shutil import copyfile
from datetime import datetime
import pandas
from openpyxl.reader.excel import load_workbook
import openpyxl
import os
import xlrd
import xlwt
from xlutils.copy import copy

from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import xlsxwriter
import pydash
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode
from anytree.importer import JsonImporter
from anytree.importer import DictImporter
from anytree.exporter import DotExporter
from anytree import AnyNode, RenderTree

from datetime import datetime
import copy
oNow = datetime.now()
import re
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

path = str(Path(Path(__file__).parent.absolute()).parent.absolute())

sys.path.insert(0, path)
file_dir = os.path.dirname(__file__)
sys.path.append(file_dir)

from pyconfig.LogService import LogService
cmd_folder = os.path.realpath(os.path.abspath(os.path.split(inspect.getfile( inspect.currentframe() ))[0]))
if cmd_folder not in sys.path:
    sys.path.insert(0, cmd_folder)



from script_files.MessageHandling import MessageHandling
from script_files import CommonValidations
from pyconfig import loadConfig
# Explore Parameteres
oPostParameters = json.loads(sys.argv[1])
# with open(r'c:\vartemp\temp.txt','w') as f:
#     f.write(str(sys.argv[1]))

# with open(r'c:\vartemp\temp.txt','r') as f:
#     oPostParameters = json.loads(f.read())
    # sys.argv[1] = f.read()
MessageHandling = MessageHandling()
LogService.log(f'Revice arguarlmnet {sys.argv[1]}')
cUserID = oPostParameters['_cUserID']
_cImportedExcelToJSONSuccessFile = oPostParameters['_cImportedExcelToJSONSuccessFile']

cDirTemplateSampleFile = oPostParameters['cDirFile']
cFileName = oPostParameters['cFileName']
cFileType = oPostParameters['cFileType']
_cImportHeaderVaidationRuleFile = oPostParameters['_cImportHeaderVaidationRuleFile']
cRegexFileFullFilePath = cDirTemplateSampleFile + _cImportHeaderVaidationRuleFile

aResponse = defaultdict(dict)
aMergedHeaders = defaultdict(dict)
aParentHeaders = defaultdict(dict)
aChildHeaders = defaultdict(dict)
aValidations = defaultdict(dict)
aSecondaryValidations = defaultdict(dict)
aRangeValidations = defaultdict(dict)
aPrimaryValidations = defaultdict(dict)
aValidationsColors = defaultdict(dict)
oNow = datetime.now()

def FunDCT_ExplodePrimaryValidations(jColumn, cCellVal, cStartRange, oWorkbook, oSheet, iRow):
    try:
        # with open(r'C:\var\DCT_R\dct-backend\assets\templates\importtoheader\text.txt','a') as f:
        #         f.write(f'{loadConfig.PY_VALIDATION_SEPARATOR} in {cCellVal} \n')
        if loadConfig.PY_VALIDATION_SEPARATOR in cCellVal:
            
            aPrimaryValidations[str(cStartRange)] = cCellVal.split(loadConfig.PY_VALIDATION_SEPARATOR)
        else:
            ##below added if MIXED_DATA(' in cCellVal: condition by spirgonde for validation issue 
            
            # if 'MIXED_DATA(' in cCellVal:
            #     aPrimaryValidations[str(cStartRange)] = cCellVal
                
            # elif '(' in cCellVal:
            #     aPrimaryValidations[str(cStartRange)] = cCellVal.split('(')[0]
            # else:
            aPrimaryValidations[str(cStartRange)] = cCellVal
        
                
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_ExplodePrimaryValidations - due to ' + str(e) + ' Line no = ' + str(line) + ' CellValue = ' + str(cCellVal))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_ExplodePrimaryValidations - due to ' + str(e) + ' Line no = ' + str(line) + ' CellValue = ' + str(cCellVal))

def FunDCT_ExplodeSecondaryValidations(jColumn, iIndexSecondaryVal, cCellVal):
    try:
        cSecondaryExploder = '('
        if cSecondaryExploder in cCellVal:
            aExplSecondaryValidations = cCellVal.split(cSecondaryExploder)
            iIndexSecondaryVal = 0
            for cValidations in aExplSecondaryValidations:
                aSecondaryValidations[jColumn] = cValidations
                aRangeValidations[jColumn] = cValidations
                iIndexSecondaryVal += 1
    except Exception as e:
        LogService.log('Error : Error while parsing file - FunDCT_ExplodeSecondaryValidations - due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_ExplodeSecondaryValidations - due to ' + str(e))

def FunDCT_GetTextColor(oCell):
    try:
        cTextColor = '000000'
        if oCell.font.color:
            if 'rgb' in oCell.font.color.__dict__:
                cTextColor = oCell.font.color.rgb[2:]          

        return cTextColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))

def FunDCT_GetBackColor(oCell, cTextColor):
    try:
        cBackColor = 'ffffff'
        
        if oCell.fill.start_color != '' and type(oCell.fill.start_color.index) != int:                    
            cBackColor = oCell.fill.start_color.index[2:]
            if cTextColor == '000000' and cBackColor == '000000':
                cBackColor = 'ffffff'
        elif oCell.fill.start_color != '' and type(oCell.fill.start_color.index) == int:
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index][2:])
        elif isinstance(oCell.fill, PatternFill) and oCell.fill.patternType == 'solid':
            if oCell.fill.fgColor.type =='rgb':
                cBackColor = oCell.fill.fgColor.rgb[2:]
            elif oCell.fill.fgColor.type =='indexed':
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index -1][2:])
        
        return cBackColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))

def FunDCT_GetRegexValidations(iMaxRow, iMaxColumn, oSheet, oWorkbook):
    try:
        iRegexIndex = 0
        for iRow in range(iMaxRow, iMaxRow+1):
            for jColumn in range(1, iMaxColumn+1):
                if iRow != iMaxRow:
                    continue
                oCell = oSheet.cell(row=iRow, column=jColumn)
                if type(oCell.value) == str:
                    cCellVal = oCell.value.strip('\n')
                else:
                    cCellVal = oCell.value
                cStartRange = "".join(filter(lambda x: not x.isdigit(), oCell.coordinate))
                cTextColor = FunDCT_GetTextColor(oCell)
                cBackColor = FunDCT_GetBackColor(oCell,cTextColor)

                aValidationsColors[str(cStartRange)] = {
                    'cTextColor' : '#'+ str(cTextColor),
                    'cBackColor' : '#'+ str(cBackColor),
                    'cStartRange' : cStartRange,
                }
                
                if type(cCellVal) == str:
                    

                    FunDCT_ExplodePrimaryValidations(jColumn, cCellVal, cStartRange, oWorkbook, oSheet, iRow)
                    iIndexSecondaryVal = 0
                    for cValidations in aPrimaryValidations[str(cStartRange)]:
                        FunDCT_ExplodeSecondaryValidations(jColumn, iIndexSecondaryVal, cValidations)
                        iIndexSecondaryVal += 1
                    iRegexIndex += 1
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetRegexValidations - due to ' + str(e) + ' Line = ' + str(line))

        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetRegexValidations - due to ' + str(e) + ' Line = ' + str(line))

def FunDCT_CellMerged(oWorkbook, oSheet, iRow, jColumn):
    try:
        oCell = oSheet.cell(iRow, jColumn)
        for mergedCell in oSheet.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                return True
        return False
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file due to ' + str(e))

def FunDCT_ParseTemplateFileToJSON():
    try:
        LogService.log('Parseing Started')
        LogService.log('Loading workbood' + cDirTemplateSampleFile + cFileName + cFileType)
        oWorkbook = load_workbook(cDirTemplateSampleFile + cFileName + cFileType)
        oSheet = oWorkbook.active
        iMaxRow = oSheet.max_row
        iMaxColumn = oSheet.max_column
        aValues = {}
        iIndex = 0
        LogService.log('Workbook Loaded')
        LogService.log('REX Validation STarted')
        FunDCT_GetRegexValidations(iMaxRow, iMaxColumn, oSheet, oWorkbook)
        LogService.log('RED Validation Ended')
        cSuccessErrorFileName = cDirTemplateSampleFile + "CREATE_TEMPLATE-IMPORT-JSON-" + str(cUserID) + '-' + str(oNow.strftime("%Y-%m-%d-%H-%M-%S")) + '.json'

        for iRow in range(1, iMaxRow):
            for jColumn in range(1, iMaxColumn+1):
                oCell = oSheet.cell(row=iRow, column=jColumn)
                aValues[oCell.value] = oCell.value
                cCellVal = oCell.value

                if cCellVal is not None:
                    aCellCordinate = []
                    aMergedCellString = []
                    cStartRange = oCell.coordinate
                    cStartRangePrimaryVal = "".join(filter(lambda x: not x.isdigit(), cStartRange))
                    cEndRange = '-'
                    if FunDCT_CellMerged(oWorkbook, oSheet, iRow, jColumn):
                        cEndRange = '-'
                    
                    cTextColor = FunDCT_GetTextColor(oCell)
                    cBackColor = FunDCT_GetBackColor(oCell,cTextColor)

                    cComment = ''
                    if oCell.comment:
                        cComment = oCell.comment.text
                    if iRow <= 1:
                        aParentHeaders[iIndex] = {
                            'cHeaderLabel' : cCellVal,
                            'cTextColor' : '#'+ cTextColor,
                            'cBackColor' : '#'+ cBackColor,
                            'children' : [],
                            'cParentHeader' : '',
                            'aCellCordinate' : aCellCordinate,
                            'cComment' : cComment,
                            'cCellCordinate' : cStartRange,
                            'aMergedCellString' : aMergedCellString,
                            'cStartRange' : "".join(filter(lambda x: not x.isdigit(), cStartRange)),
                            'cEndRange' : cEndRange,
                            'iRow' : iRow,
                            'iMaxRow' : iMaxRow,
                            'jColumn' : jColumn,
                            'iIndex' : iIndex,
                            'aPrimaryValidations' :  aPrimaryValidations,
                            # 'aPrimaryValidations' :  dict(enumerate(aPrimaryValidations.get(cStartRangePrimaryVal))),
                            # 'aValidationsColors' : aValidationsColors[cStartRangePrimaryVal],
                            'aValidationsColors' : aValidationsColors,
                            # 'aValidationsColors' : dict(enumerate(aValidationsColors.get(cStartRangePrimaryVal))),
                        }
                    else:
                        aParentHeaders[iIndex] = {
                            'cHeaderLabel' : cCellVal,
                            'cTextColor' : '#'+ cTextColor,
                            'cBackColor' : '#'+ cBackColor,
                            'children' : [],
                            'cParentHeader' : '',
                            'aCellCordinate' : aCellCordinate,
                            'cComment' : cComment,
                            'cCellCordinate' : cStartRange,
                            'aMergedCellString' : aMergedCellString,
                            'cStartRange' : "".join(filter(lambda x: not x.isdigit(), cStartRange)),
                            'cEndRange' : cEndRange,
                            'iRow' : iRow,
                            'iMaxRow' : iMaxRow,
                            'jColumn' : jColumn,
                            'iIndex' : iIndex,
                            'aPrimaryValidations' :  aPrimaryValidations,
                            # 'aPrimaryValidations' :  dict(enumerate(aPrimaryValidations.get(cStartRangePrimaryVal))),
                            'aValidationsColors' : aValidationsColors,
                            # 'aValidationsColors' : aValidationsColors[cStartRangePrimaryVal],
                            # 'aValidationsColors' : dict(enumerate(aValidationsColors.get(cStartRangePrimaryVal))),
                        }
                   
                else: continue
                
                iIndex += 1
        LogService.log('Opening work book ==> ' + cDirTemplateSampleFile + _cImportedExcelToJSONSuccessFile)
        with open(cDirTemplateSampleFile + _cImportedExcelToJSONSuccessFile, "w") as oFilePointer:
            json.dump(aParentHeaders , oFilePointer)

        LogService.log(f'_cImportedExcelToJSONSuccessFile {_cImportedExcelToJSONSuccessFile}' )
        return MessageHandling.FunDCT_MessageHandling('Success', _cImportedExcelToJSONSuccessFile)
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file due to ' + str(e))
    
FunDCT_ParseTemplateFileToJSON()