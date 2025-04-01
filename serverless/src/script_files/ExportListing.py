import sys
import json
from openpyxl.reader.excel import load_workbook
import openpyxl
import os
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import xlsxwriter
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode

from pyconfig import loadConfig

from MessageHandling import MessageHandling
import CommonValidations
import datetime
import xml.etree.cElementTree as ET
import pandas as pd
from collections import defaultdict

from datetime import datetime

import re
from pyconfig.model_common import model_common_Cls
# import xlsxwriter module
from pprint import pprint
from pyconfig import awsCloudFileManager
from pyconfig import loadConfig
sys.path.insert(0, os.path.abspath(__file__))
oPostParameters = json.loads(sys.argv[1])
# with open(r'c:\vartemp\temp4.txt','w') as f:
#     f.write(str(sys.argv[1]))
# with open(r'c:\vartemp\temp4.txt','r') as f:
    # oPostParameters = json.loads(f.read())
model_common = model_common_Cls()
MessageHandling = MessageHandling()
oNow = datetime.now()
cDirExportListing = oPostParameters['cDirFile']
cFileName = oPostParameters['cFileName']
cType = oPostParameters['cType']
cSearchFilterQuery = oPostParameters['cSearchFilterQuery']
oSort = oPostParameters['oSort']
cTemplateTypes = oPostParameters['cTemplateTypes']
cUserID = oPostParameters['cUserID']
# Added by spirgonde for creating header from moongoose.
Schema = oPostParameters['Schema']
lStaticHeader =  oPostParameters['aStaticHeader']

cSuccessFileName = cDirExportListing + cFileName
oWorkbook = xlsxwriter.Workbook(cSuccessFileName)
oWorksheet = oWorkbook.add_worksheet(cType)
compName = model_common.FunDCT_GetCompny(cUserID)
userRole = model_common.FunDCT_getUserRole(cUserID)
valid_template_ids = model_common.get_valid_template_ids()
if not userRole:userRole = 'Member'
if compName:
    compName = compName[0].get('cCompanyname')
else:
    compName = ''

# String containing Value that needs to be stored in Cells that have no data 
cNoData:str='NO DATA'

def FunDCT_ExcelFile():
    try:
        arrydata = []
        if len(cTemplateTypes) > 0:
            arrydata = cTemplateTypes
        else:
            arrydata = model_common.get_user_template_type_details(cUserID)

        oExportListingData = model_common.FunDCT_ExportListing(cType,cSearchFilterQuery,compName,userRole,cUserID,oSort,arrydata,valid_template_ids)
        # aHeadder = oExportListingData[0].keys()
        aHeadder = Schema

        oCellFormat = oWorkbook.add_format()
        oCellFormat.set_font_color('#EE4443')
        oCellFormat.set_bg_color('#ff4040')
        oCellFormat.set_align('center')
        oCellFormat.set_valign('vcenter')

        cCol = 0
        ##  for cKey, cVal in enumerate(aHeadder):
        data_format3 = oWorkbook.add_format({'bold': True})
        data_format3.set_bg_color('#FFD970')
        for cKey, cVal in enumerate(lStaticHeader):
            oWorksheet.write(0, cCol, str(cVal),data_format3)
            cCol += 1

        for iRow, aRowDetails in enumerate(oExportListingData):
            iRow += 1
            for cKeyHeader, cVal in enumerate(aHeadder):
                sCellValue = ''
                for iColNum, cColumnKey in enumerate(aRowDetails):
                    # with open(r'temp.txt','w') as f:
                    #     f.write(f"\n {str(cVal)} -- {str(cColumnKey)}")
                    try:
                        if str(cVal) == str(cColumnKey):
                            sCellValue = aRowDetails[cColumnKey]
                            if str(cColumnKey) == "tEntered":
                                sCellValue = sCellValue.strftime("%Y-%m-%d %H:%M:%S")
                            if str(cColumnKey) == "tUpdated":
                                sCellValue = sCellValue.strftime("%Y-%m-%d %H:%M:%S")
                            if str(cColumnKey) == "tProcessStart":
                                sCellValue = sCellValue.strftime("%Y-%m-%d %H:%M:%S")
                            if str(cColumnKey) == "tProcessEnd":
                                sCellValue = sCellValue.strftime("%Y-%m-%d %H:%M:%S")
                            if str(cColumnKey) == "cAllowedScacCode":
                                lTemplateTypes: list = []
                                if aRowDetails["cAllowedScacCode"]:
                                # for cTemplateType in aRowDetails[cColumnKey]:
                                #     lTemplateTypes.append(
                                #         cTemplateType['cTemplateType'])
                                    sCellValue = ",".join(aRowDetails["cAllowedScacCode"])
                            if str(cColumnKey) == "cTemplateType":
                                lTemplateTypes: list = []
                                if isinstance(aRowDetails[cColumnKey],str):
                                    sCellValue = aRowDetails[cColumnKey]
                                elif aRowDetails[cColumnKey]:
                                    try:
                                        listd = model_common.getTemplateName(aRowDetails[cColumnKey])
                                        sCellValue = ",".join(listd)
                                    except:
                                        sCellValue = ""

                            if str(cVal) == "tCuttoffdate" and "tCuttoffdate" in aRowDetails:
                                sCellValue = aRowDetails["tCuttoffdate"]
                                if isinstance(sCellValue, datetime):
                                    sCellValue = sCellValue.strftime("%Y-%m-%d %H:%M:%S")
                                elif not sCellValue:
                                    sCellValue = cNoData 
                                break
                            if (str(cColumnKey) == "cAddtionalComment"):
                                sCellValue = aRowDetails[cColumnKey]
                                # sCellValue = aRowDetails[cColumnKey]['cAddtionalComment']
                                complilertext = re.compile('<.*?>')
                                clearedText = re.sub(complilertext,'',str(sCellValue))
                                sCellValue = clearedText

                            if (str(cColumnKey) =='bExceptionfound') :
                                if aRowDetails['bExceptionfound'] == 'Y' and aRowDetails['bProcesslock'] == 'N' and aRowDetails['bProcessed'] == 'Y':
                                    sCellValue = 'Exception'
                                elif aRowDetails['bExceptionfound'] == 'N' and aRowDetails['bProcesslock'] == 'N' and aRowDetails['bProcessed'] == 'Y':
                                    sCellValue = 'Success'
                                elif aRowDetails['bExceptionfound'] == 'N' and aRowDetails['bProcesslock'] == 'N' and aRowDetails['bProcessed'] == 'N':
                                    sCellValue = 'Pending'
                                elif aRowDetails['bExceptionfound'] == 'N' and aRowDetails['bProcesslock'] == 'Y' and aRowDetails['bProcessed'] == 'N':
                                    sCellValue = 'In Progress'
                                elif aRowDetails['bExceptionfound'] == 'P' and aRowDetails['bProcessed'] == 'Y':
                                    sCellValue = 'Partially Insert'
                                elif aRowDetails['bProcessed'] == 'F':
                                    sCellValue = 'Technical Failure'
                                else:
                                    sCellValue = '-'
                            break

                        elif (str(cVal) == "cAccessCode") and str(cColumnKey) == "oAccessTypeListing":
                            sCellValue = aRowDetails[cColumnKey]['cAccessCode']
                            break

                        elif (str(cVal) == "cStatus") and str(cColumnKey) == "oStatus":
                            sCellValue = aRowDetails[cColumnKey]['cStatusCode']
                            break

                        elif (str(cVal) == "cEnteredby") and str(cColumnKey) == "oEnteredby":
                            sCellValue = aRowDetails[cColumnKey]['cUsername']
                            break
                        elif (str(cVal) == "cUploadedBy") and str(cColumnKey) == "oEnteredby":
                            sCellValue = aRowDetails[cColumnKey]['cName']
                            break
                        elif (str(cVal) == "cEmailUploadedBy") and str(cColumnKey) == "oEnteredby":
                            sCellValue = aRowDetails[cColumnKey]['cEmail']
                            break
                        elif (str(cVal) == "cUpdatedby") and str(cColumnKey) == "oUpdatedby":
                            sCellValue = aRowDetails[cColumnKey]['cUsername']
                            break

                        elif (str(cVal) == "cProcessCode") and str(cColumnKey) == "oProcessListing":
                            sCellValue = aRowDetails[cColumnKey]['cProcessCode']
                            break

                        elif (str(cVal) == "cCodeRegion") and str(cColumnKey) == "oRegionListing":
                            sCellValue = aRowDetails[cColumnKey]['cCode']
                            break

                        elif (str(cVal) == "cTemplateTypeListing") and str(cColumnKey) == "cTemplateType":
                            # sCellValue = aRowDetails[cColumnKey].keys()
                            lTemplateTypes: list = []
                            for cTemplateType in aRowDetails['oTemplateTypeListing']:
                                # print(x['cTemplateType'])
                                lTemplateTypes.append(
                                    cTemplateType['cTemplateType'])
                            sCellValue = ",".join(lTemplateTypes)
                            break
                        
                        

                            

                        
                        else:
                            # sCellValue = f"\n {str(cVal)} -- {str(cColumnKey)}"
                            sCellValue = cNoData
                    except:
                        sCellValue = ''
            
       
                oWorksheet.write(iRow, cKeyHeader, str(sCellValue))

        oWorkbook.close()
        s3path = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/TemparyFile/' + str(cType) + '_' + datetime.strftime(oNow,'%Y-%m-%d_%H%M%S_%f') + '.xlsx'  
        presignedURL = uploadAndgetPresingUrl(cSuccessFileName, s3path)
        if os.path.exists(cSuccessFileName): os.remove(cSuccessFileName)

        return MessageHandling.FunDCT_MessageHandling('Success', presignedURL)
    except Exception as e:
        print(e)
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_ExcelFile - due to ' + str(e))

def uploadAndgetPresingUrl(cSuccessFileName,s3Path):

    url = awsCloudFileManager.getPresignedUrlPresingedURL(loadConfig.AWS_BUCKET,s3Path,cSuccessFileName)
    return url
FunDCT_ExcelFile()
