from pathlib import Path
import sys,os
import inspect
import re
import json
import datetime
from collections import defaultdict
from shutil import copyfile
from datetime import datetime
import openpyxl.workbook
from pyconfig.MemoryService import MemoryService
oNow = datetime.now()
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

path = str(Path(Path(__file__).parent.absolute()).parent.absolute())
sys.path.insert(0, path)
file_dir = os.path.dirname(__file__)
sys.path.append(file_dir)
cmd_folder = os.path.realpath(os.path.abspath(os.path.split(inspect.getfile( inspect.currentframe() ))[0]))
if cmd_folder not in sys.path:
    sys.path.insert(0, cmd_folder)

import pandas
from openpyxl.reader.excel import load_workbook
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, Border, Side, PatternFill, GradientFill, Alignment, colors,protection,Protection
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter


from script_files.MessageHandling import MessageHandling
from script_files import CommonValidations
from pyconfig import loadConfig
from pyconfig import awsCloudFileManager
from pyconfig.model_common import model_common_Cls
from pyconfig.LogService import LogService
import prefillNomemberupload
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# Explore Parameteres
aGlobalParameters = defaultdict(dict)
oPostParameters = json.loads(sys.argv[1])
# with open(r'c:\vartemp\Distribute_Release.txt','w') as f:
#     f.write(str(sys.argv[1]))
# with open(r'c:\vartemp\Distribute_Release.txt','r') as f:
#     oPostParameters = json.loads(f.read())    
# with open(r'c:\vartemp\Member_Upload.txt','r') as f:
#     oPostParameters = json.loads(f.read())
# with open(r'c:\vartemp\DHL.txt','r') as f:
#     oPostParameters = json.loads(f.read())
# with open(r'c:\vartemp\tempvalidationpymember2.txt','r') as f:
#     oPostParameters = json.loads(f.read())
# with open(r'c:\vartemp\tempvalidationpymember.txt','r') as f:
#     oPostParameters = json.loads(f.read())
# with open(r'c:\vartemp\tempvalidationpymember_alinkk_new.txt','r') as f:
#     oPostParameters = json.loads(f.read())
iTemplateID = oPostParameters['iTemplateID']
_cUploadType = oPostParameters['_cUploadType']
_cCompanyname = oPostParameters['_cCompanyname']
_iTemplateUploadLogID = oPostParameters['_iTemplateUploadLogID']
cTemplateType = oPostParameters['cTemplateType']
iMaxDepthHeaders = oPostParameters['iMaxDepthHeaders']
cAdditionalFieldValue = oPostParameters['cAdditionalFieldValue']
cDirDistribution = oPostParameters['cDirDistribution']
cDirConsolidation = oPostParameters['cDirConsolidation']
cFileDir = oPostParameters['cFileDir']
cFileName = oPostParameters['cFileName']
cFileType = oPostParameters['cFileType']
tCuttoffdate = oPostParameters['tCuttoffdate']
oPostParameters['cKeyUniqueIDSequence'] = 0
cFullFilePath = cFileDir + cFileName + cFileType
oPostParameters['cFullFilePath'] = cFullFilePath

cSampleTemplateFileDir = oPostParameters['cSampleTemplateFileDir']
cSampleTemplateFile = oPostParameters['cSampleTemplateFile']
cSampleFileType = oPostParameters['cSampleFileType']
cSampleFullFilePath = cSampleTemplateFileDir + cSampleTemplateFile + cSampleFileType

cDirValidateTemplateHeaderLevel = oPostParameters['cDirValidateTemplateHeaderLevel']
_cValidateTemplateFileJSON = oPostParameters['_cValidateTemplateFileJSON']

_cTemplateCommentsFileJSON = oPostParameters['_cTemplateCommentsFileJSON']

cDirValidateTemplateStatusFile = oPostParameters['cDirValidateTemplateStatusFile']
_cValidateTmplateStatusFile = oPostParameters['_cValidateTmplateStatusFile']
cStatusFileFullFilePath = cDirValidateTemplateStatusFile + _cValidateTmplateStatusFile
oPostParameters['cStatusFileFullFilePath'] = cStatusFileFullFilePath
succussFilePath = cDirValidateTemplateStatusFile + 'SuccessFile' + _cValidateTmplateStatusFile

_cValidatedExcelToJSONSuccessFile = oPostParameters['_cValidatedExcelToJSONSuccessFile']
with open(cDirValidateTemplateHeaderLevel + _cValidateTemplateFileJSON, "r") as oValidationsDetails:
    oValidationsDetails = json.load(oValidationsDetails)
# if iTemplateID:
#     oValidationsDetails = model_common.FunDCT_GET_HEADER_VALIDATION(iTemplateID)
    


with open(cDirValidateTemplateHeaderLevel + _cTemplateCommentsFileJSON, "r") as oCommentsDetails:
    oCommentsDetails = json.load(oCommentsDetails)

aValidateHeader = defaultdict(dict)
aValidateContent = defaultdict(dict)
aValidateResponse = defaultdict(dict)
aValidateContent_testing = defaultdict(dict)
cValidatePattern = ''
cKey = ''
cNewkeytype = ''
cNewval = ''
model_common = model_common_Cls()
MessageHandling = MessageHandling()
cSampleFile = Path(cSampleFullFilePath)
LogService.log(f'cSampleFile =={cSampleFile}')
LogService.log(f'cSampleFile.is_file() =={cSampleFile.is_file()}')

if not cSampleFile.is_file():
    LogService.log(f'cSampleFile.is_file() =={cSampleFile.is_file()}')

    try:
        # oDataFrameList = model_common.FunDCT_GetTemplateLog('gen_templates',iTemplateID)
        # cReplacedNull = oDataFrameList[0]['cUploadedFile'].replace("null", "'#DCT#'")
        # oDataFrame = pandas.DataFrame(eval(cReplacedNull))
        # oDataFrame.columns = oDataFrame.columns.str.replace('#DCT#.*', '')
        # oDataFrame.to_excel(cSampleFullFilePath, index=False)

        # cS3UrlUploadedOrg = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirUploadFile), cUploadedFileOrg)
        # cS3DirUploadFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_INTERTEAM) + '/'
        
        testpathname = str(cSampleTemplateFile) + str(cSampleFileType)
        
        cS3DirSampleFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_SAMPLEFILE) + '/' + str(testpathname)
        LogService.log(f'donwloading file,  {str(cSampleFullFilePath)}')
        
        awsCloudFileManager.download_file_from_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirSampleFile), str(cSampleFullFilePath))
    except Exception as e:
        LogService.log(f'Trackbak , hense hard exit {e}')

        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        # print("cSampleFile ERRRRRRRRRRRR - ",str(e) + 'Line no - '+ str(line) + ' cReplacedNull = '+cReplacedNull)
        # print("cSampleFile ERRRRRRRRRRRR - ",str(e) + 'Line no - '+ str(line) + ' AWS_BUCKET = ' +str(loadConfig.AWS_BUCKET) + 'AWS_BUCKET_TEMPLATES_SAMPLEFILE' + str(loadConfig.AWS_BUCKET_TEMPLATES_SAMPLEFILE) + ' testpathname =  ' + str(testpathname) + ' cS3DirSampleFile = ' + str(cS3DirSampleFile))
        exit
LogService.log(f'_cUploadType {_cUploadType}')
if _cUploadType != 'DISTRIBUTE' and str(oPostParameters['preFillData']).strip().lower() == 'yes':
    try:
        Path(cDirDistribution+str(iTemplateID)).mkdir(exist_ok=True)
        cDistributeFileToCompare =  str(cDirDistribution) + str(iTemplateID) + '/' + str(iTemplateID) + '.xlsx'
        bDistributeFileToCompare = Path(cDistributeFileToCompare)
        # if not bDistributeFileToCompare.is_file():
        oDataFrameDistributed = model_common.FunDCT_GetDistributionData(iTemplateID)
            # cReplacedNull = oDataFrameDistributed[0]['aDistributedData'].replace('null', "'#DCT#'")
        awsCloudFileManager.download_file_from_bucket(loadConfig.AWS_BUCKET,oDataFrameDistributed[0]['aDistributedData'],cDistributeFileToCompare)
        # bytesData = awsCloudFileManager.getfile(loadConfig.AWS_BUCKET,oDataFrameDistributed[0]['aDistributedData'])
        # oDataFrameDistributedData = pandas.read_excel(bytesData,na_values = "Missing", na_filter=True, skiprows = iMaxDepthHeaders - 1,engine='openpyxl')
        # oDataFrameDistributedData.columns = oDataFrameDistributedData.columns.str.replace('#DCT#.*', '')
        # oDataFrameDistributedData.to_excel(cDistributeFileToCompare, index=False)
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log(f"ERRRRRRRRRRRR - {str(e)} Line no - + {str(line)}")
        exit



if str(oPostParameters['preFillData']).strip().lower() == 'yes'.lower():
    # Initiating Workbooks
    oWorkbookTemplate = load_workbook(cSampleFullFilePath)
    # print(f'{MemoryService.getMemoryUsedStr()}- TEmplate load mery')
    oSheetTemplate = oWorkbookTemplate.active
    iMaxRowTemplate = oSheetTemplate.max_row
    iMaxColumnTemplate = oSheetTemplate.max_column
    oPostParameters['oWorkbookTemplate'] = oWorkbookTemplate

    # print(f'{MemoryService.getMemoryUsedStr()}- uploaded file membero loadss')
    # oWorkbookUploaded = pandas.ExcelFile(oPostParameters['cFullFilePath'],'openpyxl').book    
    oWorkbookUploaded = load_workbook(oPostParameters['cFullFilePath'])
    # print(f'{MemoryService.getMemoryUsedStr()}- uploaded file membero loadss')
    oSheetUploaded = oWorkbookUploaded.active
    iMaxRowUploaded = oSheetUploaded.max_row
    iMaxColumnUploaded = oSheetUploaded.max_column

    # Code Adeed by Asif to triming the values for header 
    # for iColx in range(iMaxColumnUploaded):
    # cColumnLetter = f"{get_column_letter(1)}1:{get_column_letter(iMaxColumnUploaded)}{iMaxRowUploaded}"
    
    # for eachCell in oSheetUploaded.iter_rows(min_row=iMaxDepthHeaders+1,max_col=iMaxColumnUploaded,max_row=iMaxRowUploaded):
    #     for cCell in eachCell:
    #         if cCell.value is not None:
    #             cCell.value = str(cCell.value).strip()
    oPostParameters['oSheetUploaded'] = oSheetUploaded
    oWorkbookStatusFile = openpyxl.Workbook()
    oWorkbookStatusFile.save(cStatusFileFullFilePath)
    oWorkbookStatusFile = load_workbook(cStatusFileFullFilePath)
    oWorkbookSuccessFile = openpyxl.Workbook()
    oSheetSuccessFile = oWorkbookSuccessFile.active
    oSheetStatusFile = oWorkbookStatusFile.active
    oSheetStatusFileDATA = []
    oSheetStatusFileDATASuccess = []
    # oPostParameters['oWorkbookStatusFile'] = oWorkbookStatusFile
    oPostParameters['oSheetTemplate'] = oSheetTemplate
    oPostParameters['oValidationsDetails'] = oValidationsDetails
    oPostParameters['cSampleFullFilePath'] = cSampleFullFilePath

cStringSeparator = '#$'

aUnMappedColumnLabels = defaultdict(dict)
aUnMappedColumnRange = defaultdict(dict)
aUnMappedColumnRangeDetails = defaultdict(dict)
aMappedColumnLabels = defaultdict(dict)
aMappedColumnRange = defaultdict(dict)
aMappedColumnRangeDetails = defaultdict(dict)

aRangeStart = []
aRangeEnd = []

aRangeStartmapped = []
aRangeEndmapped = []

cS3DirStatusFile = ''
cS3DirUploadFile = ''
if _cUploadType == 'DISTRIBUTE':
    cS3DirStatusFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_INTERTEAM) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_STATUSFILE) + '/'
    cS3DirUploadFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_INTERTEAM) + '/'
elif _cUploadType != 'DISTRIBUTE':
    cS3DirStatusFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_MEMBER) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_STATUSFILE) + '/'
    cS3DirUploadFile = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_UPLOADTEMPLATE_MEMBER) + '/'

from script_files import Distribution

aSavedSearchedRow = defaultdict(dict)
oSheetTemplateUniqueSeq = defaultdict(dict)
if oPostParameters['_cUploadType'] != 'DISTRIBUTE' and str(oPostParameters['preFillData']).strip().lower() == 'yes'.lower():
    cDistributeFileToCompare =  str(oPostParameters['cDirDistribution']) + str(oPostParameters['iTemplateID']) + '/' + str(oPostParameters['iTemplateID']) + '.xlsx'
    oWorkbookTemplateUniqueSeq = load_workbook(cDistributeFileToCompare,read_only=True,)
    data = []
    ws = oWorkbookTemplateUniqueSeq.active
    indexID = 0
    # for ii,row in enumerate(ws.iter_rows(values_only=True),1):
    #     # data.append(row)
    #     oSheetTemplateUniqueSeq[ii]=row
    # ------------------------------------------
    for ii, row in enumerate(ws.iter_rows(values_only=False), start=1):
        formatted_row = []
        for cell in row:
            if cell.number_format in ['0%', '0.00%']:  # Check if the cell is formatted as a percentage
                formatted_value = f"{round(cell.value * 100, 2)}%"  # Convert 0.3 to '30%'
            else:
                formatted_value = cell.value  # Keep the value as-is
            formatted_row.append(formatted_value)
        row = formatted_row
    # ------------------------------------------
        oSheetTemplateUniqueSeq[ii]=row
        if int(oPostParameters['iMaxDepthHeaders']) >= ii:
            for k,cln in enumerate(row):
                if str(cln).upper() in ('WWA LANE ID','WWA ID', 'CENTRIC ID','SR.NO.','INDEX ID', 'LANE ID') :
                    oPostParameters['cKeyUniqueIDSequence'] = k+1
                    indexID = k
                    break
                k+=1
            continue
        try:
            ds = int(row[indexID])
            if ds not in aSavedSearchedRow:
                aSavedSearchedRow[int(row[indexID])] = ii
            else:
                aSavedSearchedRow['Unknown_Row' + str(ii)] = ii    
        except:
            aSavedSearchedRow['Unknown_Row' + str(ii)] = ii
        ii +=1
    # pexb = pandas.ExcelFile(cDistributeFileToCompare,'openpyxl')
    # dpf =  pex.parse(skiprows=int(int(oPostParameters['iMaxDepthHeaders']) - 1 ), nrows=1000)
    # oSheetTemplateUniqueSeq = pandas.read_excel(pexb,header=int(oPostParameters['iMaxDepthHeaders'])-1,na_values='',engine='openpyxl')
    # data = data[int(int(oPostParameters['iMaxDepthHeaders']) - 1 ):]
    # oSheetTemplateUniqueSeq = pandas.DataFrame(data,dtype=str,)
    # oSheetTemplateUniqueSeq.columns = oSheetTemplateUniqueSeq.iloc[0]
    # print(f'{MemoryService.getMemoryUsedStr()}- dataframe loaded')
    # colName = 'WWA LANE ID'
    # i = 1
    # for cln in list(oSheetTemplateUniqueSeq.columns):
    #     if str(cln).upper() in ('WWA LANE ID','WWA ID', 'CENTRIC ID','SR.NO.','INDEX ID', 'LANE ID') :
    #                     colName = cln
    #                     oPostParameters['cKeyUniqueIDSequence'] = i
    #                     break
    #     i+=1
    # for chunk in pandas.read_excel(skiprows=int(int(oPostParameters['iMaxDepthHeaders']) - 1 ), nrows=5000,):
    # for i in range(0,10):
    #     chunk = pandas.read_excel(pex,header=int(oPostParameters['iMaxDepthHeaders'])-1,skiprows=(5000 * i), nrows=5000,engine='openpyxl')
    #     chunks.append(chunk)
    # oWorkbookTemplateUniqueSeq = pandas.concat(chunks,axis=0)
    # oWorkbookTemplateUniqueSeq = pandas.ExcelFile(cDistributeFileToCompare,'openpyxl',).parse(skiprows = int(int(oPostParameters['iMaxDepthHeaders'],) - 1),)
    # aSavedSearchedRow = oSheetTemplateUniqueSeq[colName].to_dict()
    
    # oSheetTemplateUniqueSeq =  oWorkbookTemplateUniqueSeq.worksheets[0]
    # c = oSheetTemplateUniqueSeq['C5000']
    # cCellValOrg = oSheetTemplateUniqueSeq.iter_rows(min_row=5000,min_col=3,max_col=3,max_row=5000)
    # for c in cCellValOrg:
    #     for x in c:
    #         cCellValOrg = x.value
    # oPostParameters['cKeyUniqueIDSequence'] = CommonValidations.FunDCT_GetUniqueIDSequence(oSheetTemplateUniqueSeq, cTemplateType)
    # cKeyUniqueIDSequence = int(oPostParameters['cKeyUniqueIDSequence'])-1
    # iMax = oSheetTemplateUniqueSeq.max_row + 1
    # if cKeyUniqueIDSequence != 'NA' and cKeyUniqueIDSequence >= 0:
        # colLetter = get_column_letter(oPostParameters['cKeyUniqueIDSequence'])
        # iRow = iMaxDepthHeaders if iMaxDepthHeaders > 1 else iMaxDepthHeaders+1
        # iRow = iMaxDepthHeaders + 1
        # for cRowIter in oSheetTemplateUniqueSeq.iter_rows(min_row=iRow,max_col=1,max_row=iMax):
        #     for cCelliter in cRowIter:
        #         if cCelliter.value is not None:
        #             aSavedSearchedRow[int(cCelliter.value)] = iRow
        #         iRow +=1
                
    # if cKeyUniqueIDSequence != 'NA' and cKeyUniqueIDSequence >= 0:
    #     for iRow in range(1, oSheetTemplateUniqueSeq.max_row + 1):
    #         if iRow <= iMaxDepthHeaders:
    #             continue
    #         if oSheetTemplateUniqueSeq[iRow][cKeyUniqueIDSequence].value is not None:
    #             aSavedSearchedRow[int(oSheetTemplateUniqueSeq[iRow][cKeyUniqueIDSequence].value)] = iRow

aRestrictedScacCodeAll = []
aRestrictedScacCode = []
if cTemplateType == 'RFQTool' and cAdditionalFieldValue != '-':
    aRestrictedScacCodeAll = model_common.FunDCT_GetRestrictedMembers(cTemplateType, cAdditionalFieldValue)

if len(aRestrictedScacCodeAll) > 0:
    for aRestrictedScacDetails in aRestrictedScacCodeAll:
        aRestrictedScacCode.append(aRestrictedScacDetails['cRestrictedScacCode'])

oPostParameters['aRestrictedScacCode'] = aRestrictedScacCode

def FunDCT_GetUnMappednColumnDetails(_cCompanyname):
    try:
        aGetUnMappedColumnLabels = model_common.FunDCT_GetUnMappednColumns(iTemplateID)
        for cMemberColumnKey, aMemberColumnDetails in enumerate(aGetUnMappedColumnLabels):
            for iIndex in range(len(aMemberColumnDetails['aUnMappedColumns'])):
                if _cCompanyname in aMemberColumnDetails['aUnMappedColumns'].keys():
                    for cKey, cVal in aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0].items():
                        iRangeStart = int(aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0][cKey]['iActualCellCol'])
                        iTotalLen = int(aMemberColumnDetails['aUnMappedColumns'][_cCompanyname][0][cKey]['iTotalLen'])
                        iRangeEnd = int(iRangeStart) + int(iTotalLen-1)
                        
                        aRangeStart.append(iRangeStart)
                        aRangeEnd.append(int(iRangeStart) + int(iTotalLen)-1)
                        aUnMappedColumnRange[iRangeStart] = {
                            'iRangeStart': iRangeStart,
                            'iRangeEnd': iRangeEnd,
                            'iTotalLen': iTotalLen,
                        }
                        aUnMappedColumnRangeDetails[str(iRangeStart)+"-"+str(iRangeEnd)] = {
                            'iRangeStart': iRangeStart,
                            'iRangeEnd': iRangeEnd,
                            'iTotalLen': iTotalLen,
                        }
                        aUnMappedColumnLabels[str(_cCompanyname)] = {
                            '_cCompanyname': aMemberColumnDetails['aUnMappedColumns'][_cCompanyname],
                        }
        for cMemberColumnKey, aMemberColumnDetails in enumerate(aGetUnMappedColumnLabels):
            for iIndex in range(len(aMemberColumnDetails['aMappedColumns'])):
                if _cCompanyname in aMemberColumnDetails['aMappedColumns'].keys():
                    for cKey, cVal in aMemberColumnDetails['aMappedColumns'][_cCompanyname][0].items():
                        iRangeStart = int(aMemberColumnDetails['aMappedColumns'][_cCompanyname][0][cKey]['iActualCellCol'])
                        iTotalLen = int(aMemberColumnDetails['aMappedColumns'][_cCompanyname][0][cKey]['iTotalLen'])
                        iRangeEnd = int(iRangeStart) + int(iTotalLen-1)
                        
                        aRangeStartmapped.append(iRangeStart)
                        aRangeEndmapped.append(int(iRangeStart) + int(iTotalLen)-1)
                        aMappedColumnRange[iRangeStart] = {
                            'iRangeStart': iRangeStart,
                            'iRangeEnd': iRangeEnd,
                            'iTotalLen': iTotalLen,
                        }
                        aMappedColumnRangeDetails[str(iRangeStart)+"-"+str(iRangeEnd)] = {
                            'iRangeStart': iRangeStart,
                            'iRangeEnd': iRangeEnd,
                            'iTotalLen': iTotalLen,
                        }
                        aMappedColumnLabels[str(_cCompanyname)] = {
                            '_cCompanyname': aMemberColumnDetails['aMappedColumns'][_cCompanyname],
                        }

    except Exception as e:
        LogService.log('FunDCT_GetUnMappednColumnDetails Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetUnMappednColumnDetails Error while parsing file due to ' + str(e))

if _cUploadType != 'DISTRIBUTE' and str(oPostParameters['preFillData']).strip().lower() == 'yes'.lower():
    FunDCT_GetUnMappednColumnDetails(_cCompanyname)

def FunDCT_PrepareTemplateStatusFile(iIndex, iRow, iColumn, cCoordinate, cMessage):
    aValidateHeader[iIndex] = {
        'Row': iRow,
        'Column':  get_column_letter(iColumn),
        'cStartRange': cCoordinate,
        'Message': cMessage
    }

    if oSheetStatusFile.cell(iRow, iColumn).value is None:
        oSheetStatusFile.cell(iRow, iColumn).value = cMessage
        oSheetStatusFile.cell(iRow, iColumn).font = Font(color="FF0000", italic=True)
        oSheetStatusFile.cell(iRow, iColumn).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    else:
        cCellValueNew = oSheetStatusFile.cell(iRow, iColumn).value + cStringSeparator + cMessage
        oSheetStatusFile.cell(iRow, iColumn).value = cCellValueNew
        oSheetStatusFile.cell(iRow, iColumn).font = Font(color="FF0000", italic=True)
        oSheetStatusFile.cell(iRow, iColumn).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

def FunDCT_GetFontColor(oCell):
    cTextColor = '000000'
    if oCell.font.color:
        if 'rgb' in oCell.font.color.__dict__:
            cTextColor = oCell.font.color.rgb[2:]
    return cTextColor

def FunDCT_GetBackGroundColor(oCell):
    cTextColor = '000000'
    cBackColor = 'ffffff'
    if oCell.fill.start_color != '':                    
        cBackColor = str(oCell.fill.start_color.index)[2:]
        if cTextColor == '000000' and cBackColor == '000000':
            cBackColor = 'ffffff'
    else :
        cBackColor = str(oCell.fill.bgColor.index).zfill(6)
    return cBackColor

def FunDCT_CellMerged(oSheet, iRow, jColumn):
    try:
        oCell = oSheet.cell(iRow, jColumn)
        for mergedCell in oSheet.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                return True
        return False
    except Exception as e:
        LogService.log('FunDCT_CellMerged Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_CellMerged Error while parsing file due to ' + str(e))

def FunDCT_GetMergeCellRangeLen(oSheetTemplate, iRow, jColumn):
    try:
        iTotalLen = 0
        oCell = oSheetTemplate.cell(iRow, jColumn)
        for mergedCell in oSheetTemplate.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                cCellRangeStart, cCellRangeEnd = str(mergedCell).split(':')
                iCellCol = CommonValidations.FunDCT_GetColumnNumber(
                    ''.join([i for i in cCellRangeStart if not i.isdigit()]))
                iTotalLen = int(CommonValidations.FunDCT_GetColumnNumber(
                    ''.join([i for i in cCellRangeEnd if not i.isdigit()]))) - (int(iCellCol)-1)
        return iTotalLen
    except Exception as e:
        LogService.log('Error , FunDCT_GetMergeCellRangeLen Error while parsing file due to ' + str(e))

        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetMergeCellRangeLen Error while parsing file due to ' + str(e))

# Validate Headers
test = defaultdict(dict)
def FunDCT_ValidateHeader(oPostParameters):
    try:
        aKeys = []
        bColumnHeaderNotInRange = False
        bInUnmappedRange = False
        if _cUploadType != 'DISTRIBUTE' and len(aUnMappedColumnLabels) > 0:
            aUnMappedColumnRangeList = [(aRangeStart[iRangeIndex], aRangeEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeStart))]
            bInUnmappedRange = True
            bColumnHeaderNotInRange = True
            aKeys = aUnMappedColumnLabels[_cCompanyname]['_cCompanyname'][0].keys()
        
        iIndex = 0
        for iRow in range(1, iMaxRowTemplate+1):
            iSubsctractCols = 0
            for jColumn in range(1, iMaxColumnTemplate+1):
                oSampleCell = oSheetTemplate.cell(row=iRow, column=jColumn)
                oSampleCellVal = oSampleCell.value
                if len(aKeys) > 0:
                    bColumnHeaderNotInRange = True if str(oSampleCellVal) in aKeys else False
                    bInUnmappedRange = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aUnMappedColumnRangeList)
                
                iMergedCellLen = FunDCT_GetMergeCellRangeLen(oSheetTemplate, iRow, jColumn)
                if (oSampleCellVal is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen > 0:
                    iSubsctractCols += iMergedCellLen
                    continue
                elif (oSampleCellVal is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen <= 0:
                    iSubsctractCols += 1
                    continue
                # elif (oSampleCellVal is None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True):
                #     iSubsctractCols += 1
                #     continue

                # iBackwardCol = int(jColumn) - int(iSubsctractCols)
                # oSampleCell = oSheetTemplate.cell(row=iRow, column=iBackwardCol)
                # oSampleCellVal = oSampleCell.value
                

                if oSampleCellVal is not None:
                    iBackwardCol = int(jColumn) - int(iSubsctractCols)
                    oUploadedCell = oSheetUploaded.cell(row=iRow, column=iBackwardCol)
                    oUploadedCellVal = str(oUploadedCell.value)
                    bMatchHeaderValue = True if str(oSampleCellVal).strip() == str(oUploadedCellVal) else False

                    if not bMatchHeaderValue:
                        cErrorMessage = MessageHandling.FunDCT_GetValidationDescription('UPLOAD_TEMPLATE','TEMPLATE_VALIDATE_HEADER_MISMATCH').replace('#$DCTVARIABLE_CORRECT_CELL_VAL#$', str(oSampleCellVal)).replace('#$DCTVARIABLE_INCORRECT_CELL_VAL#$', str(oUploadedCellVal) )
                        FunDCT_PrepareTemplateStatusFile(iIndex, oUploadedCell.row, oUploadedCell.column, str(oUploadedCell.coordinate), cErrorMessage)
                        iIndex += 1
                    else:                       
                        oSheetStatusFile.cell(iRow, iBackwardCol).value = oSampleCell.value
                        oSheetSuccessFile.cell(iRow, iBackwardCol).value = oSampleCell.value
                        oSheetStatusFile.cell(iRow, iBackwardCol).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        oSheetSuccessFile.cell(iRow, iBackwardCol).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cTextColor = FunDCT_GetTextColor(oSampleCell)
                        cBackColor = FunDCT_GetBackColor(oSampleCell,cTextColor)
                        oSheetStatusFile.cell(row=iRow, column=iBackwardCol).font = Font(color=cTextColor)
                        oSheetStatusFile.cell(row=iRow, column=iBackwardCol).fill = PatternFill(fgColor=cBackColor,fill_type = "solid")
                        thin_border = Side(style="thin")
                        oSheetStatusFile.cell(row=iRow, column=iBackwardCol).border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                        if FunDCT_CellMerged(oSheetTemplate, iRow, iBackwardCol):
                            if iMergedCellLen > 0:
                                ranges = f"{get_column_letter(iBackwardCol)}{iRow}:{get_column_letter((iBackwardCol - 1)+iMergedCellLen)}{iRow}"
                                oSheetStatusFile.merge_cells(ranges)

    
       
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error FunDCT_ValidateHeader Error while parsing file due to ' + str(e) + ' Lineno == '+str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_ValidateHeader Error while parsing file due to ' + str(e) + ' Lineno == '+str(line))

def FunDCT_GetTextColor(oCell):
    try:
        cTextColor = '000000'
        if oCell.font.color:
            if 'rgb' in oCell.font.color.__dict__:
                cTextColor = oCell.font.color.rgb[2:]          

        return cTextColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))

def FunDCT_GetBackColor(oCell, cTextColor):
    try:
        cBackColor = 'ffffff'
        
        if oCell.fill.start_color != '' and type(oCell.fill.start_color.index) != int:                    
            cBackColor = oCell.fill.start_color.index[2:]
            if cTextColor == '000000' and cBackColor == '000000':
                cBackColor = 'ffffff'
        elif oCell.fill.start_color != '' and type(oCell.fill.start_color.index) == int:
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index][2:])
        elif isinstance(oCell.fill, PatternFill) and oCell.fill.patternType == 'solid':
            if oCell.fill.fgColor.type =='rgb':
                cBackColor = oCell.fill.fgColor.rgb[2:]
            elif oCell.fill.fgColor.type =='indexed':
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index -1][2:])
        
        return cBackColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))


def check_value_exist(test_dict, value, cKey):
    try:
        do_exist = False
        for key, val in test_dict.items():
            if val == value and cKey == key:
                do_exist = True
        return do_exist
    except Exception as e:
        LogService.log('check_value_exist Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'check_value_exist Error while parsing file due to ' + str(e))

def FunDCT_MatchHeaderLabel(oValidationsDetails, iIndexValidatornext, cCellVal):
    try:
        bMatch = False
        for cValidationKeyObj, cValidationValObj in oValidationsDetails[iIndexValidatornext].items():
            if cValidationKeyObj == 'cHeaderLabel':
                if cValidationValObj == cCellVal:
                    bMatch = True
        return bMatch
    except Exception as e:
        LogService.log('Error FunDCT_MatchHeaderLabel Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_MatchHeaderLabel Error while parsing file due to ' + str(e))

def FunDCT_RemoveUnnecessaryChar(cCellVal):
    try:
        cCellVal = cCellVal.strip()
        cCellVal = cCellVal.strip('\n')
        return cCellVal
    except Exception as e:
        LogService.log('FunDCT_RemoveUnnecessaryChar Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_RemoveUnnecessaryChar Error while parsing file due to ' + str(e))

def FunDCT_MoveToMethod(oPostParameters):
    try:
        bIsValid = True
        bValidate = False
        iIndexMove = 0
        aValidations = defaultdict(dict)
        for cValidationString in oPostParameters['cValidationValObj']:
            oPostParameters['cValidationString'] = cValidationString
            if cValidationString.startswith('DEFAULT_'):
                # oPostParameters['oCellVal'] = FunDCT_RemoveUnnecessaryChar(str(oPostParameters['oCellVal']))
                bValidate = CommonValidations.FunDCT_DefaultValidations(oPostParameters)
            # if oPostParameters['_cUploadType'] != 'DISTRIBUTE':
            #     bValidate = True
            elif cValidationString.startswith('SERVICEINFO_HEADER'):
                bValidate = CommonValidations.FunDCT_Default_Scac(oPostParameters)
            elif cValidationString.startswith('MIXED_') and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                bValidate = CommonValidations.FunDCT_MixedData(oPostParameters)
            elif cValidationString.startswith('DEPENDENT(') and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                bValidate = CommonValidations.FunDCT_DependentData(oPostParameters)
            elif cValidationString.startswith('DATE_RANGE(') and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                bValidate = CommonValidations.FunDCT_DateRangeValidations(oPostParameters)
            elif cValidationString.startswith('RESTRICTED_TEXT(') and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                bValidate = CommonValidations.FunDCT_RestrictedTextValidations(oPostParameters)
            elif cValidationString.startswith('DATA_VALIDATION(') and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                bValidate = CommonValidations.FunDCT_DATAValidations(oPostParameters)
            elif (cValidationString.startswith('TEXT_RANGE(') or cValidationString.startswith('NUMERIC_RANGE(')) and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                
                bValidate = CommonValidations.FunDCT_RangeValidations(oPostParameters)
            elif cValidationString == 'NO_CHANGE' and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                bValidate = CommonValidations.FunDCT_NoChangeValidations(oPostParameters, aSavedSearchedRow, oSheetTemplateUniqueSeq)
                # else:
                #     bValidate = True if oPostParameters['oCellVal'] else False
            elif cValidationString == 'CHECKED_RATE_INCREASE' and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                
                if oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                    bValidate = CommonValidations.FunDCT_CheckedRateIncreaseValidations(oPostParameters, aSavedSearchedRow, oSheetTemplateUniqueSeq)
            elif cValidationString.startswith('NO_DUPLICATE') and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                bValidate = True
                if oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                    bValidate = CommonValidations.FunDCT_NoDuplicate    (oPostParameters)
            elif cValidationString.startswith('NOT_ACCEPTED') and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                bValidate = CommonValidations.FunDCT_NOT_ACCEPTEDIF(oPostParameters)
            elif cValidationString.upper() == 'NUMERIC' and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                    # bValidate = True if isinstance(oPostParameters['oCellVal'], int) else f"{oPostParameters['oCellVal']} is Not A Number"
                    try:
                        abc = int(oPostParameters['oCellVal'])
                        bValidate = True
                    except:
                        bValidate = f"Invalid Number"
            elif cValidationString.upper() == 'TEXT' and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                bValidate = True if  str(oPostParameters['oCellVal']) != '' else f"Invalid Text"
            elif cValidationString.upper() =='DECIMAL' and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                    try:
                        abc = float(oPostParameters['oCellVal'])
                        bValidate = True
                    except:
                        bValidate = f"Invalid Decimal Number"
                # else:
                #     bValidate = CommonValidations.FunDCT_CommonValidations(oPostParameters) 
            else:
                bValidate = CommonValidations.FunDCT_CommonValidations(oPostParameters)
            aValidations[iIndexMove] = {
                'bValidate':bValidate,
                'cCellVal':oPostParameters['oCellVal']
            }
            iIndexMove += 1
        for cKey in aValidations:
            bIsValid = aValidations[cKey]['bValidate']
            if bIsValid != True:
                break
        return bIsValid
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_MoveToMethod Error while parsing file due to ' + str(e))

def FunDCT_ValidateRegex(oPostParameters):
    try:
        bValidate = False
        bNoCond = True
        dValidations = oPostParameters['oValidationsDetails'][oPostParameters['iForwardValidator']]
        if dValidations.get('cValidationsCondition'):
            bNoCond = False
            oPostParameters['cValidationValObj'] = dValidations.get('cValidationsCondition')
            bValidate = FunDCT_MoveToMethod(oPostParameters)
        # for cValidationKeyObj, cValidationValObj in oPostParameters['oValidationsDetails'][oPostParameters['iForwardValidator']].items():
        #     if cValidationKeyObj == 'cValidationsCondition':
        #         bNoCond = False
        #         oPostParameters['cValidationValObj'] = cValidationValObj
        #         bValidate = FunDCT_MoveToMethod(oPostParameters)
        # if bNoCond:
        #     for cValidationKeyObj, cValidationValObjSingle in oPostParameters['oValidationsDetails'][oPostParameters['iForwardValidator']].items():
        #         if cValidationKeyObj == 'cValidations':
        #             oPostParameters['cValidationValObj'] = cValidationValObjSingle
        #             bValidate = FunDCT_MoveToMethod(oPostParameters)
        return bValidate
    except Exception as e:
        LogService.log('Error FunDCT_ValidateRegex Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_ValidateRegex Error while parsing file due to ' + str(e))

def FunDCT_GetNextColumn(jColumn):
    try:
        iSkipRows = 0
        for iLowerRange, iUpperRange in zip(aRangeStart, aRangeEnd):
            if iLowerRange <= jColumn <= iUpperRange:
                iSkipRows += aUnMappedColumnRangeDetails[str(iLowerRange)+"-"+str(iUpperRange)]['iTotalLen']
                jColumn += iSkipRows
                if jColumn <= iMaxColumnUploaded+1:
                    FunDCT_GetNextColumn(jColumn)
        return iSkipRows
    except Exception as e:
        LogService.log('Error FunDCT_GetNextColumn Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetNextColumn Error while parsing file due to ' + str(e))
def FunDCT_GetNextColumn1(jColumn):
    try:
        iSkipRows = 0
        for iLowerRange, iUpperRange in zip(aRangeStart, aRangeEnd):
            if iLowerRange <= jColumn <= iUpperRange:
                iSkipRows += aUnMappedColumnRangeDetails[str(iLowerRange)+"-"+str(iUpperRange)]['iTotalLen']
                jColumn += 1
                if jColumn <= iMaxColumnUploaded+1:
                    FunDCT_GetNextColumn(jColumn)
        return iSkipRows
    except Exception as e:
        LogService.log('Error FunDCT_GetNextColumn Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_GetNextColumn Error while parsing file due to ' + str(e))


# Adde by Asif for member upload/not 
def FunDCT_SplitFiles(oPostParameters,cS3DirStatusFile):
    try:
        bk = load_workbook(str(oPostParameters['cStatusFileFullFilePath']))
        exl = pandas.ExcelFile(bk,engine='openpyxl')
        # oDataFrame = pandas.read_excel(str(oPostParameters['cStatusFileFullFilePath']),engine='openpyxl')
        oDataFrame = exl.parse()
        iIndexDataFrame = iter(range(1, len(oDataFrame.columns) + 1))
        successFilePathL = oPostParameters['cFileDir'] + 'Success_' + oPostParameters['cFileName'] + '.xlsx'
        errorFilePathL = oPostParameters['cFileDir'] + 'Error_' + oPostParameters['cFileName']+ '.xlsx'
        oDataFrame.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in oDataFrame.columns]
        secondRowDF = oDataFrame[:1]
        # 
        oDataFrameError = oDataFrame[oDataFrame.stack().str.contains('Error -').any(level=0) == True]
        oDataFrameError = pandas.concat([secondRowDF,oDataFrameError]) 
        oDataFrameSuccess = oDataFrame[oDataFrame.stack().str.contains('Error -').any(level=0) == False]
        # oDataFrameSuccess = secondRowDF + oDataFrameSuccess
        oDataFrameSuccess.columns = oDataFrameSuccess.columns.str.replace('#DCT#.*', '')
        oDataFrameError.columns = oDataFrameSuccess.columns.str.replace('#DCT#.*', '')

        oDataFrameSuccess.to_excel(str(successFilePathL), index=False)
        oDataFrameError.to_excel(str(errorFilePathL), index=False)
        CommonValidations.FunDCT_ApplyHeaderFormating(str(successFilePathL), iMaxDepthHeaders=iMaxDepthHeaders)
        try:
            CommonValidations.FunDCT_ApplyHeaderFormating(str(errorFilePathL), iMaxDepthHeaders=iMaxDepthHeaders)
        except:
            pass

        oWorkStatus = openpyxl.load_workbook(str(errorFilePathL))
        oSheetStatus = oWorkStatus.active
        oSheetStatus.protection.sheet = True
        oSheetStatus.protection.password = loadConfig.EXL_PASS
        iMaxR = oSheetStatus.max_row
        iMaxC = oSheetStatus.max_column
        for oRow in  oSheetStatus.iter_rows(min_row=iMaxDepthHeaders+1,max_row=iMaxR):
            for oCell in oRow:
                # indexCol = 1
                if oCell.value:
                    if 'Error' in  str(oCell.value) :
                        oCell.font = Font(color="FF0000", italic=True)
                        oCell.protection = Protection(locked=False)

                
                # bStatus = oPostParameters['oSheetUploaded'].cell(iiRwo, iColx).protection.locked
                # oSheetStatus.cell(iRr,iColx).protection = Protection(locked=bStatus)
                
                
        # for iiRwo in range(3,oPostParameters['oSheetUploaded'].max_row + 1):
        #     for iRr in range(3,oSheetStatus.max_row+1):
        #         if str(oPostParameters['oSheetUploaded'].cell(iiRwo, 1).value) == str(oSheetStatus.cell(iRr,1).value):
        #             for iColx in range(1, oSheetStatus.max_column):
        #                 bStatus = oPostParameters['oSheetUploaded'].cell(iiRwo, iColx).protection.locked
        #                 oSheetStatus.cell(iRr,iColx).protection = Protection(locked=bStatus)
        #                 if oSheetStatus.cell(iRr,iColx).value:
        #                     if 'Error' in  oSheetStatus.cell(iRr,iColx).value :
        #                         oSheetStatus.cell(iRr,iColx).font = Font(color="FF0000", italic=True) 

        oWorkStatus.save(str(errorFilePathL))
        oWorkStatus.close()

        cS3UrlFileError = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirStatusFile), str(errorFilePathL))
        cS3UrlFileSuccess = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirStatusFile), str(successFilePathL))
        model_common.FunDCT_UpdateSuccessErrorFilePath(oPostParameters['_iTemplateUploadLogID'],cS3UrlFileSuccess,cS3UrlFileError)
        os.remove(errorFilePathL)
        os.remove(successFilePathL)
    except Exception as Err:
        LogService.log(f'FunDCT_SplitFiles {Err}')





# Validate Content
def FunDCT_ValidateContent(oPostParameters):
    try:
        
        iIndexValidator = 0
        iIndexValidatornext = 0
        bColumnHeaderNotInRange = False
        bInUnmappedRange = False
        aKeys = []
        if oPostParameters['_cUploadType'] != 'DISTRIBUTE':
            aUnMappedColumnRangeList = [(aRangeStart[iRangeIndex], aRangeEnd[iRangeIndex]) for iRangeIndex in range(0, len(aRangeStart))]
            aMappedColumnRangeList  = [(aRangeStartmapped[iRangeIndex], aRangeEndmapped[iRangeIndex]) for iRangeIndex in range(0, len(aRangeStartmapped))]
            bInUnmappedRange = True
            bColumnHeaderNotInRange = True
            aKeys = aUnMappedColumnLabels[oPostParameters['_cCompanyname']]['_cCompanyname'][0].keys() if len(aUnMappedColumnLabels[oPostParameters['_cCompanyname']]) > 0 else []
            
        aGetMemberColumns = model_common.FunDCT_GetMemberScacValidationCol(oPostParameters['iTemplateID'])
        aMemberScacColIndex = []
        aMemberScacColLabel = defaultdict(dict)
        mappingSCAC =  CommonValidations.getSCAC_columnMapping(oPostParameters['oValidationsDetails'])
        for cMemberColumnKey, aMemberColumnDetails in enumerate(aGetMemberColumns):
            # if aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel'] == 'SERVICE INFO':
            if aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel'] == 'SERVICE INFO':
                continue
            iColumnIndex = CommonValidations.FunDCT_GetColumnIndex(oPostParameters['oSheetUploaded'], aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel'])
            aMemberScacColIndex.append(iColumnIndex)
            aMemberScacColLabel[iColumnIndex]['cHeaderLabel'] = str(aGetMemberColumns[cMemberColumnKey]['aTemplateHeader']['cHeaderLabel']).strip()
        for iRow in range(1, iMaxRowUploaded+1):
            if iRow <= iMaxDepthHeaders:
                continue
            iSubsctractCols = 0
            oPostParameters['NS'] = False
            tRowList = []
            bRowAdd = False
            for ix, eachCell in enumerate(oSheetUploaded[f'A{iRow}:{get_column_letter(iMaxColumnUploaded)}{iRow}']):
                for eachCll in eachCell:
                    if str(eachCll.value).strip().lower() == 'ns':
                        oPostParameters['NS'] = True
                        break
             
            iMergedCellLen = 0
            for jColumn in range(1, iMaxColumnUploaded+1):
                oCell = oSheetUploaded.cell(row=iRow, column=jColumn)
                if oCell.number_format in ['0%', '0.00%']:
                    oCell.value = f"{round(oCell.value * 100, 2)}%"  # Convert 0.3 to '30%'
                oPostParameters['oCellVal'] = oCell.value
                cCellVal = oCell.value
                cStartRange = oCell.coordinate
                bValidContent = False
                
                if len(aKeys) > 0:
                    
                    bColumnHeaderNotInRange = True if str(oPostParameters['oCellVal']) in aKeys else False
                    bInUnmappedRange = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aUnMappedColumnRangeList)
                    binMappedRange  = any(iLowerRange <= jColumn <= iUpperRange for (iLowerRange, iUpperRange) in aMappedColumnRangeList)
                    if bInUnmappedRange == True:
                        
                            
                            # if FunDCT_GetNextColumn(jColumn-iMergedCellLen) > iMergedCellLen:
                            #     iMergedCellLen = FunDCT_GetNextColumn(jColumn-iMergedCellLen)
                            iMergedCellLen = 0
                            if iMaxDepthHeaders == 1:
                                    if FunDCT_GetNextColumn1(jColumn-iSubsctractCols) >iSubsctractCols:
                                        iSubsctractCols = FunDCT_GetNextColumn1(jColumn-iSubsctractCols)
                            else:                                
                                if FunDCT_GetNextColumn(jColumn-iSubsctractCols) >iSubsctractCols:
                                    iSubsctractCols = FunDCT_GetNextColumn(jColumn-iSubsctractCols)                                
                            # ccd = FunDCT_GetNextColumn(jColumn-iMergedCellLen)
                            # ccd1 = FunDCT_GetNextColumn(jColumn+iMergedCellLen)
                            # backclogskips = FunDCT_GetNextColumn(jColumn-1)
                            # if iMergedCellLen < backclogskips:
                            #     iMergedCellLen = backclogskips
                            # if ccd > iMergedCellLen:
                            #     iMergedCellLen = ccd
                            # if ccd1 > 0 and ccd > 0:
                            #     iMergedCellLen =ccd1 + ccd


                            
                    else:
                        if binMappedRange:
                            ccd1 = FunDCT_GetNextColumn(jColumn+iSubsctractCols)
                            ccd = FunDCT_GetNextColumn(jColumn-iSubsctractCols)
                            if ccd1 > 0 and ccd > 0:
                                iSubsctractCols = ccd1 + ccd
                   
                # if (oPostParameters['oCellVal'] is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True ) and iMergedCellLen > 0:
                # if (bColumnHeaderNotInRange is True or bInUnmappedRange is True ) and iMergedCellLen > 0:
                #     iSubsctractCols = iMergedCellLen
                # elif (oPostParameters['oCellVal'] is not None) and (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen <= 0:
                # elif (bColumnHeaderNotInRange is True or bInUnmappedRange is True) and iMergedCellLen <= 0:
                #     iSubsctractCols = iMergedCellLen
                
                oPostParameters['iForwardValidator'] = int(int(jColumn-1) + int(iSubsctractCols))
                oPostParameters['iBackwardValidator'] = int(jColumn-1)

                oUploadedCell = oSheetUploaded.cell(row=iRow, column=jColumn)
                oPostParameters['cBackColorUploadedCell'] = FunDCT_GetBackGroundColor(oUploadedCell)
                oPostParameters['cTextColorUploadedCell'] = FunDCT_GetFontColor(oUploadedCell)

                oPostParameters['iRow'] = iRow
                oPostParameters['iCol'] = jColumn
                ###############################################
                ###   NS LOGIC ##################
                ################################################
                dValidations = oPostParameters['oValidationsDetails'][oPostParameters['iForwardValidator']]
                conditinalValidation = dValidations.get('cValidationsCondition')
                sSCACname = ''
                if 'ORIGIN_COLUMN' in conditinalValidation:
                    sSCACname = 'DEFAULT_ORIGIN_SCAC' 
                elif 'OCEAN_COLUMN' in conditinalValidation:
                    sSCACname = 'DEFAULT_OCEAN_SCAC' 
                elif 'DESTINATION_COLUMN' in conditinalValidation:
                    sSCACname = 'DEFAULT_DESTINATION_SCAC' 
                elif 'SERVICEINFO_COLUMN' in conditinalValidation:
                    sSCACname ='SERVICEINFO_HEADER'
                
                bLiable = None
                if mappingSCAC.get(sSCACname) is not None:
                    try:
                        if mappingSCAC.get(sSCACname):
                            scacint = int(mappingSCAC.get(sSCACname))
                            liablitySCAC = oSheetUploaded.cell(row=iRow, column=scacint+1).value
                            bLiable = liablitySCAC ==_cCompanyname

                    except:
                        pass
                ######################### NS LOGIC #############################
                oPostParameters['iIndexValidatornext'] = iIndexValidatornext
                cCellVal = str(oPostParameters['oCellVal'])
                if oPostParameters['NS'] and bLiable == True and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                    bValidContent = CommonValidations.FunDCT_NS(oCell)
                elif bLiable == True and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                    bValidContent = FunDCT_ValidateRegex(oPostParameters)
                elif bLiable == False and oPostParameters['_cUploadType'] != 'DISTRIBUTE':
                    bValidContent = True
                else:
                    bValidContent = FunDCT_ValidateRegex(oPostParameters)
                
                
                if bValidContent is not True:
                    if str(cCellVal).lower() == 'none' : cCellVal = 'Empty'
                    aValidateContent[iIndexValidator] = {
                        'Message' : f"Error - {str(bValidContent)}, Cell Value is {str(cCellVal)}",
                        'Row': oPostParameters['iRow'],
                        'Column': get_column_letter(jColumn),
                        # 'iRow' : oPostParameters['iRow'],
                        # 'jColumn' : jColumn,
                        'cCellVal' : cCellVal ,
                        # 'cStartRange': cStartRange,                            
                        # 'bValidContent': bValidContent,
                        # 'iForwardValidator': oPostParameters['iForwardValidator'],
                        # 'iBackwardValidator': oPostParameters['iBackwardValidator'],
                        # 'iSubsctractCols': iSubsctractCols,
                    }
                    
                    # oSheetStatusFile.cell(iRow, jColumn).value = 'Error - ' + str(bValidContent) + ' ' + str(cCellVal)   
                    tRowList.append(f"Error - {str(bValidContent)} , Cell Value is {str(cCellVal)}")
                    # oSheetStatusFile.cell(iRow, jColumn).font = Font(color="FF0000", italic=True)
                    bRowAdd = True
                    # print(f'{MemoryService.getMemoryUsedStr()}- writing error file status')
                else:
                    cCellValUpdated = cCellVal.strip()
                    
                    if oPostParameters['_cUploadType'] == 'DISTRIBUTE':
                        if 'NO_CHANGE' not in oPostParameters['cValidationValObj']:
                            cCellValUpdated = ""
                    if cCellValUpdated.lower() == 'none':
                        cCellValUpdated = ''
                    
                    # oSheetStatusFile.cell(iRow, jColumn).value = cCellValUpdated
                    tRowList.append(cCellValUpdated)
                    # print(f'{MemoryService.getMemoryUsedStr()}- Writing status file memberoy')
                
                iIndexValidator += 1               
                iIndexValidatornext += 1
            
            if bRowAdd:
                oSheetStatusFileDATA.append(tRowList)
            else:
                oSheetStatusFileDATASuccess.append(tRowList)
            
        R = iMaxDepthHeaders+1
        for rows in  oSheetStatusFileDATA:
            C = 1
            for valec in rows:
                if str(valec).lower().startswith('='):
                    formula = valec
                    matches = re.findall(r"([A-Z]+[0-9]+)\b",formula)
                    for mtch in matches:
                        rownumber = re.search(r'([0-9]+)\b',str(mtch)).group(1)
                        newR = mtch.replace(rownumber,str(R))
                        valec = valec.replace(mtch,newR)

                oSheetStatusFile.cell(R,C).value = valec
                if str(valec).lower().startswith('error'):
                    oSheetStatusFile.cell(R,C).font = Font(color="FF0000", italic=True)
                C+=1
            R+=1
        R = iMaxDepthHeaders+1
        if len(oSheetStatusFileDATASuccess)> 0:
            for rows in  oSheetStatusFileDATASuccess:
                C = 1
                for valec in rows:
                    if str(valec).lower().startswith('='):
                        formula = valec
                        matches = re.findall(r"([A-Z]+[0-9]+)\b",formula)
                        for mtch in matches:
                            rownumber = re.search(r'([0-9]+)\b',str(mtch)).group(1)
                            newR = mtch.replace(rownumber,str(R))
                            valec = valec.replace(mtch,newR)
                    oSheetSuccessFile.cell(R,C).value = valec
                    C+=1
                R+=1
            
    except Exception as e:
        LogService.log('Error', 'FunDCT_ValidateContent Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_ValidateContent Error while parsing file due to ' + str(e))

def FunDCT_ExcelToDataFrame(cFilePath):
    try:
        oDataFrame = pandas.read_excel(cFilePath,engine='openpyxl')
        iIndexDataFrame = iter(range(1, len(oDataFrame.columns) + 1))
        # ASIF - change column as string
        oDataFrame.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in oDataFrame.columns]
        (r, c) = oDataFrame.shape
        # oDataFrame.head(2).apply(lambda x: str(x).strip())
        # for colindex  in range(0,c):
        #     oDataFrame.iloc[0,colindex] = str(oDataFrame.iloc[0,colindex]).strip()
        return oDataFrame.to_json()
    except Exception as e:
        LogService.log('Error FunDCT_ExcelToDataFrame Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_ExcelToDataFrame Error while parsing file due to ' + str(e))
# Validate Template File
def FunDCT_ValidateTemplate(oPostParameters):
    try:
        if str(oPostParameters['preFillData']).strip().lower() == 'no' and _cUploadType != 'DISTRIBUTE':
            
            cUploadedFileOrg = str(cFileDir) + str(cFileName) + str(cFileType)
            cS3UrlUploadedOrg = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirUploadFile), cUploadedFileOrg)
            aSelectedMembers = prefillNomemberupload.FunDCT_PreFillNoDataUpload(oPostParameters)
            
            cS3UrlStatusFile = cS3UrlUploadedOrg
            

        else:
            # Validate Headers
            LogService.log('Validating header started')
            FunDCT_ValidateHeader(oPostParameters)
            LogService.log('Validating header Ened')
            
            if len(aValidateHeader) > 0:
                oWorkbookStatusFile.save(cStatusFileFullFilePath)
                # DB Update Log Org/Status
                cTemplateStatusUploadedFileDB = FunDCT_ExcelToDataFrame(cStatusFileFullFilePath)
                cOrgUploadedFileDB = FunDCT_ExcelToDataFrame(oPostParameters['cFullFilePath'])
                cTable = 'tmpl_uploadlogs' if oPostParameters['_cUploadType'] == 'DISTRIBUTE' else 'mem_uploadlogs'
                model_common.FunDCT_InsertTemplateLog(cTable, _iTemplateUploadLogID, cTemplateStatusUploadedFileDB, 'cTemplateStatusUploadedFile')
                model_common.FunDCT_InsertTemplateLog(cTable, _iTemplateUploadLogID, cOrgUploadedFileDB, 'cUploadedFile',json.dumps(aValidateHeader))
                # S3 Upload Org/Status file
                cUploadedFileOrg = str(cFileDir) + str(cFileName) + str(cFileType)
                cS3UrlUploadedOrg = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirUploadFile), cUploadedFileOrg)
                cFullPath = str(cStatusFileFullFilePath)
                cS3UrlStatusFile = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirStatusFile), cFullPath)
                aValidateResponse = {
                    "cFileName": cFileName,
                    "cFileType": cFileType,
                    "cFileDir": cFileDir,
                    "cFullPath": cS3UrlUploadedOrg,
                    "cStatusFilePath": cS3UrlStatusFile,
                    "aValidateContent": aValidateHeader,
                    "iErrorLenth": len(aValidateHeader),
                    "cDistributionDetails": [],
                }

                with open(cDirValidateTemplateStatusFile + _cValidatedExcelToJSONSuccessFile, "w") as oFilePointer:
                    json.dump(aValidateResponse , oFilePointer)
                return MessageHandling.FunDCT_MessageHandling('Success', _cValidatedExcelToJSONSuccessFile)
            
            LogService.log('Validate Content Started')
            # Validate Content
            FunDCT_ValidateContent(oPostParameters)
            LogService.log('Validate Content Ended')
            # print('aValidateContent',aValidateContent)
            # exit
            
            oWorkbookStatusFile.save(oPostParameters['cStatusFileFullFilePath'])
            oWorkbookSuccessFile.save(succussFilePath)

           
            # DB Update Log Org/Status
            # cTemplateStatusUploadedFileDB = FunDCT_ExcelToDataFrame(oPostParameters['cStatusFileFullFilePath'])
            # cOrgUploadedFileDB = FunDCT_ExcelToDataFrame(oPostParameters['cFullFilePath'])
            cTable = 'tmpl_uploadlogs' if oPostParameters['_cUploadType'] == 'DISTRIBUTE' else 'mem_uploadlogs'
            # model_common.FunDCT_InsertTemplateLog(cTable, _iTemplateUploadLogID, cTemplateStatusUploadedFileDB, 'cTemplateStatusUploadedFile')
            # model_common.FunDCT_InsertTemplateLog(cTable, _iTemplateUploadLogID, cOrgUploadedFileDB, 'cUploadedFile')
            LogService.log('insert upload logs')
            # S3 Upload Org/Status file
            # cUploadedFileOrg = str(cFileDir) + str(cFileName) + str(cFileType)
            cUploadedFileOrg = str(cFileDir) + str(cFileName) + str(cFileType)
            # LogService.log(f'uploding file to == >{cUploadedFileOrg}')

            cS3UrlUploadedOrg = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirUploadFile), cUploadedFileOrg)
            model_common.FunDCT_InsertTemplateLog(cTable, _iTemplateUploadLogID, cS3UrlUploadedOrg, 'cUploadedFile', json.dumps(dict(list(aValidateContent.items())[:100])))
            cFullPath = str(oPostParameters['cStatusFileFullFilePath'])
            
            # Distribution and Consolidation
            LogService.log('Distribution and consolidation file Creating strated')
            aSelectedMembers = defaultdict(dict)
            oPostParameters['cFileToCompare'] = cDirValidateTemplateStatusFile + 'SUCCESS_ERROR_' + _iTemplateUploadLogID + oPostParameters['iTemplateID'] + str(oNow.strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
            copyfile(oPostParameters['cStatusFileFullFilePath'], oPostParameters['cFileToCompare'])
            if len(oSheetStatusFileDATA) <= 0 and _cUploadType == 'DISTRIBUTE':
                aSelectedMembers = Distribution.FunDCT_DistributionAndConsolidation(oPostParameters['_cCompanyname'], oPostParameters['cDirDistribution'], succussFilePath, oPostParameters['_cUploadType'], oSheetUploaded, _iTemplateUploadLogID, cDirValidateTemplateStatusFile, oPostParameters['iTemplateID'], cDirConsolidation, iMaxDepthHeaders, aValidateContent, cS3DirStatusFile, oPostParameters['cFullFilePath'], oPostParameters['cTemplateType'])
            elif _cUploadType != 'DISTRIBUTE':
                if len(oSheetStatusFileDATASuccess) > 0:
                       aSelectedMembers = Distribution.FunDCT_DistributionAndConsolidation(oPostParameters['_cCompanyname'], oPostParameters['cDirDistribution'], succussFilePath, oPostParameters['_cUploadType'], oSheetUploaded, _iTemplateUploadLogID, cDirValidateTemplateStatusFile, oPostParameters['iTemplateID'], cDirConsolidation, iMaxDepthHeaders, aValidateContent, cS3DirStatusFile, oPostParameters['cFullFilePath'], oPostParameters['cTemplateType'])
            if len(aValidateContent) > 0 and _cUploadType == 'DISTRIBUTE':
                model_common.FunDCT_ResetMappingColumnDetails(iTemplateID)
            LogService.log('Distribution and consolidation file Creating Ended')
            
            # if _cUploadType != 'DISTRIBUTE':
            #     FunDCT_SplitFiles(oPostParameters,str(cS3DirStatusFile))
            
            # LogService.log('Applying Border and merging headers for status file')
            
            if len(oSheetStatusFileDATASuccess) > 0 and len(oSheetStatusFileDATA) > 0:
                cFilePath = cFullPath
            elif len(oSheetStatusFileDATASuccess) <= 0 and len(oSheetStatusFileDATA) > 0:
                cFilePath = cFullPath
            else:
                # cFilePath = succussFilePath
                copyfile(cUploadedFileOrg,cFullPath)
                cFilePath =cFullPath
            LogService.log(f'File UPloading to S3 {str(loadConfig.AWS_BUCKET)}, {str(cS3DirStatusFile)}, {cFilePath}')
            cS3UrlStatusFile = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cS3DirStatusFile), cFilePath)
            
            model_common.FunDCT_InsertTemplateLog(cTable, _iTemplateUploadLogID, cS3UrlStatusFile, 'cTemplateStatusUploadedFile')

            LogService.log(f'File uploaded to S3 {str(loadConfig.AWS_BUCKET)}, {str(cS3DirStatusFile)}, {cFullPath}')
            
        aValidateResponse = {
            "cFileName": cFileName,
            "cFileType": cFileType,
            "cFileDir": cFileDir,
            "cFullPath": cS3UrlUploadedOrg,
            "cStatusFilePath": cS3UrlStatusFile,
            "aValidateContent": '',
            "iErrorLenth": len(aValidateContent),
            "cDistributionDetails": aSelectedMembers,
        }
        LogService.log(f'Validation Response == {aValidateResponse}')
        LogService.log(f'Writing to file {cDirValidateTemplateStatusFile + _cValidatedExcelToJSONSuccessFile}')
        with open(cDirValidateTemplateStatusFile + _cValidatedExcelToJSONSuccessFile, "w") as oFilePointer:
            json.dump(aValidateResponse , oFilePointer)
        LogService.log('Job done successfully, now sending success response to Node')
        LogService.log('Marking Job done in DB')
        model_common.FunDCT_UpdateEndprocessingAndProcesslock(_iTemplateUploadLogID)
        LogService.log('Marked Job done in DB')
        return MessageHandling.FunDCT_MessageHandling('Success', _cValidatedExcelToJSONSuccessFile)
    except Exception as e:
        LogService.log('Error FunDCT_ValidateTemplate Error while parsing file due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'FunDCT_ValidateTemplate Error while parsing file due to ' + str(e))



FunDCT_ValidateTemplate(oPostParameters)
