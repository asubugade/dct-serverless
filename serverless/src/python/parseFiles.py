import openpyxl
import re
from io import BytesIO
import json
from fastapi import FastAPI, HTTPException
import os,sys
from pathlib import Path
from collections import defaultdict
from openpyxl.styles import colors, PatternFill


app = FastAPI()

from .MessageHandling import MessageHandling
from pyconfig.LogService import LogService
from pyconfig import loadConfig
MessageHandling = MessageHandling()

aValidationsColors = defaultdict(dict)
aValidationsColors = defaultdict(dict)
aPrimaryValidations = defaultdict(dict)
aSecondaryValidations = defaultdict(dict)
aRangeValidations = defaultdict(dict)
aParentHeaders = defaultdict(dict)





def parseTemplateFileToJSON(event: dict, context=None):
    try:
        data = event if isinstance(event, dict) else json.loads(event)
        exData = json.loads(data['body']) if isinstance(data['body'], str) else data['body']
        cFileContent = exData['cFileContent']['data']

        # Read Excel file
        file_stream = BytesIO(bytes(cFileContent))
        oWorkbook = openpyxl.load_workbook(file_stream)
        sheetNames = oWorkbook.sheetnames
        for sheetName in sheetNames:
            oSheet = oWorkbook[sheetName]
            print("Sheet Names: =============>", oSheet)
            iMaxRow = oSheet.max_row
            iMaxColumn = oSheet.max_column
            aValues = {}
            iIndex = 0
            LogService.log('Workbook Loaded')
            LogService.log('REX Validation STarted')
            FunDCT_GetRegexValidations(iMaxRow, iMaxColumn, oSheet, oWorkbook)
            LogService.log('RED Validation Ended')
            # cSuccessErrorFileName = cDirTemplateSampleFile + "CREATE_TEMPLATE-IMPORT-JSON-" + str(cUserID) + '-' + str(oNow.strftime("%Y-%m-%d-%H-%M-%S")) + '.json'

            for iRow in range(1, iMaxRow):
                for jColumn in range(1, iMaxColumn+1):
                    oCell = oSheet.cell(row=iRow, column=jColumn)
                    aValues[oCell.value] = oCell.value
                    cCellVal = oCell.value

                    if cCellVal is not None:
                        aCellCordinate = []
                        aMergedCellString = []
                        cStartRange = oCell.coordinate
                        cStartRangePrimaryVal = "".join(filter(lambda x: not x.isdigit(), cStartRange))
                        cEndRange = '-'
                        if FunDCT_CellMerged(oWorkbook, oSheet, iRow, jColumn):
                            cEndRange = '-'
                        
                        cTextColor = FunDCT_GetTextColor(oCell)
                        cBackColor = FunDCT_GetBackColor(oCell,cTextColor)

                        cComment = ''
                        if oCell.comment:
                            cComment = oCell.comment.text
                        if iRow <= 1:
                            if sheetName not in aParentHeaders:
                                aParentHeaders[sheetName] = {}
                            aParentHeaders[sheetName][iIndex] = {
                                'cHeaderLabel' : cCellVal,
                                'cTextColor' : '#'+ cTextColor,
                                'cBackColor' : '#'+ cBackColor,
                                'children' : [],
                                'cParentHeader' : '',
                                'aCellCordinate' : aCellCordinate,
                                'cComment' : cComment,
                                'cCellCordinate' : cStartRange,
                                'aMergedCellString' : aMergedCellString,
                                'cStartRange' : "".join(filter(lambda x: not x.isdigit(), cStartRange)),
                                'cEndRange' : cEndRange,
                                'iRow' : iRow,
                                'iMaxRow' : iMaxRow,
                                'jColumn' : jColumn,
                                'iIndex' : iIndex,
                                'aPrimaryValidations' :  aPrimaryValidations,
                                # 'aPrimaryValidations' :  dict(enumerate(aPrimaryValidations.get(cStartRangePrimaryVal))),
                                # 'aValidationsColors' : aValidationsColors[cStartRangePrimaryVal],
                                'aValidationsColors' : aValidationsColors,
                                # 'aValidationsColors' : dict(enumerate(aValidationsColors.get(cStartRangePrimaryVal))),
                            }
                        else:
                            aParentHeaders[sheetName][iIndex] = {
                                'cHeaderLabel' : cCellVal,
                                'cTextColor' : '#'+ cTextColor,
                                'cBackColor' : '#'+ cBackColor,
                                'children' : [],
                                'cParentHeader' : '',
                                'aCellCordinate' : aCellCordinate,
                                'cComment' : cComment,
                                'cCellCordinate' : cStartRange,
                                'aMergedCellString' : aMergedCellString,
                                'cStartRange' : "".join(filter(lambda x: not x.isdigit(), cStartRange)),
                                'cEndRange' : cEndRange,
                                'iRow' : iRow,
                                'iMaxRow' : iMaxRow,
                                'jColumn' : jColumn,
                                'iIndex' : iIndex,
                                'aPrimaryValidations' :  aPrimaryValidations,
                                # 'aPrimaryValidations' :  dict(enumerate(aPrimaryValidations.get(cStartRangePrimaryVal))),
                                'aValidationsColors' : aValidationsColors,
                                # 'aValidationsColors' : aValidationsColors[cStartRangePrimaryVal],
                                # 'aValidationsColors' : dict(enumerate(aValidationsColors.get(cStartRangePrimaryVal))),
                            }
                    
                    else: continue
                    
                    iIndex += 1
        return MessageHandling.FunDCT_MessageHandling('Success', aParentHeaders)
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file due to ' + str(e))

def FunDCT_ExplodePrimaryValidations(jColumn, cCellVal, cStartRange, oWorkbook, oSheet, iRow):
    try:
       
        if loadConfig.PY_VALIDATION_SEPARATOR in cCellVal:
            
            aPrimaryValidations[str(cStartRange)] = cCellVal.split(loadConfig.PY_VALIDATION_SEPARATOR)
        else:
          
            aPrimaryValidations[str(cStartRange)] = cCellVal
        
                
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_ExplodePrimaryValidations - due to ' + str(e) + ' Line no = ' + str(line) + ' CellValue = ' + str(cCellVal))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_ExplodePrimaryValidations - due to ' + str(e) + ' Line no = ' + str(line) + ' CellValue = ' + str(cCellVal))

def FunDCT_ExplodeSecondaryValidations(jColumn, iIndexSecondaryVal, cCellVal):
    try:
        cSecondaryExploder = '('
        if cSecondaryExploder in cCellVal:
            aExplSecondaryValidations = cCellVal.split(cSecondaryExploder)
            iIndexSecondaryVal = 0
            for cValidations in aExplSecondaryValidations:
                aSecondaryValidations[jColumn] = cValidations
                aRangeValidations[jColumn] = cValidations
                iIndexSecondaryVal += 1
    except Exception as e:
        LogService.log('Error : Error while parsing file - FunDCT_ExplodeSecondaryValidations - due to ' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_ExplodeSecondaryValidations - due to ' + str(e))

def FunDCT_GetTextColor(oCell):
    try:
        cTextColor = '000000'
        if oCell.font.color:
            if 'rgb' in oCell.font.color.__dict__:
                cTextColor = oCell.font.color.rgb[2:]          

        return cTextColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetTextColor - due to ' + str(e) + ' Line = ' + str(line))

def FunDCT_GetBackColor(oCell, cTextColor):
    try:
        cBackColor = 'ffffff'
        
        if oCell.fill.start_color != '' and type(oCell.fill.start_color.index) != int:                    
            cBackColor = oCell.fill.start_color.index[2:]
            if cTextColor == '000000' and cBackColor == '000000':
                cBackColor = 'ffffff'
        elif oCell.fill.start_color != '' and type(oCell.fill.start_color.index) == int:
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index][2:])
        elif isinstance(oCell.fill, PatternFill) and oCell.fill.patternType == 'solid':
            if oCell.fill.fgColor.type =='rgb':
                cBackColor = oCell.fill.fgColor.rgb[2:]
            elif oCell.fill.fgColor.type =='indexed':
                cBackColor = str(colors.COLOR_INDEX[oCell.fill.fgColor.index -1][2:])
        
        return cBackColor
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetBackColor - due to ' + str(e) + ' Line = ' + str(line) + ' TYPE = ' + str(type(oCell.fill.start_color)) + ' Value = ' + str(type(oCell.fill.start_color)) + ' INDEX = ' + str(oCell.fill.start_color.index))

def FunDCT_GetRegexValidations(iMaxRow, iMaxColumn, oSheet, oWorkbook):
    try:
        iRegexIndex = 0
        for iRow in range(iMaxRow, iMaxRow+1):
            for jColumn in range(1, iMaxColumn+1):
                if iRow != iMaxRow:
                    continue
                oCell = oSheet.cell(row=iRow, column=jColumn)
                if type(oCell.value) == str:
                    cCellVal = oCell.value.strip('\n')
                else:
                    cCellVal = oCell.value
                cStartRange = "".join(filter(lambda x: not x.isdigit(), oCell.coordinate))
                cTextColor = FunDCT_GetTextColor(oCell)
                cBackColor = FunDCT_GetBackColor(oCell,cTextColor)

                aValidationsColors[str(cStartRange)] = {
                    'cTextColor' : '#'+ str(cTextColor),
                    'cBackColor' : '#'+ str(cBackColor),
                    'cStartRange' : cStartRange,
                }
                
                if type(cCellVal) == str:
                    

                    FunDCT_ExplodePrimaryValidations(jColumn, cCellVal, cStartRange, oWorkbook, oSheet, iRow)
                    iIndexSecondaryVal = 0
                    for cValidations in aPrimaryValidations[str(cStartRange)]:
                        FunDCT_ExplodeSecondaryValidations(jColumn, iIndexSecondaryVal, cValidations)
                        iIndexSecondaryVal += 1
                    iRegexIndex += 1
    except Exception as e:
        trace_back = sys.exc_info()[2]
        line = trace_back.tb_lineno
        LogService.log('Error : Error while parsing file - FunDCT_GetRegexValidations - due to ' + str(e) + ' Line = ' + str(line))

        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file - FunDCT_GetRegexValidations - due to ' + str(e) + ' Line = ' + str(line))

def FunDCT_CellMerged(oWorkbook, oSheet, iRow, jColumn):
    try:
        oCell = oSheet.cell(iRow, jColumn)
        for mergedCell in oSheet.merged_cells.ranges:
            if (oCell.coordinate in mergedCell):
                return True
        return False
    except Exception as e:
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while parsing file due to ' + str(e))
