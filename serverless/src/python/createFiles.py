import xlsxwriter
import json
import pandas
import os
from fastapi import FastAPI, HTTPException
from .MessageHandling import MessageHandling
from pyconfig.LogService import LogService
from pyconfig.model_common import model_common_Cls
from anytree.importer import DictImporter
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode
from openpyxl.utils import get_column_letter
from pyconfig import loadConfig
from pyconfig import awsCloudFileManager

app = FastAPI()
MessageHandling = MessageHandling()
model_common = model_common_Cls()
importer = DictImporter()

iMaxDepthHeaders = 1

@app.post("/createfiles")
def handler(event: dict, context=None):
    try:
        data = event if isinstance(event, dict) else json.loads(event)
        exData = json.loads(data['body']) if isinstance(data['body'], str) else data['body']
 
        cDirTemplateSampleFile = exData['cDirTemplateSampleFile']
        cTemplateSampleFile = exData['cTemplateSampleFile']
        iTemplateID = exData['iTemplateID']
        oTemplateMetaData = exData['oTemplateMetaData']
        aStructuredHeaderData = exData['aAppendDatatoRoot']
        cFileType = exData['cFileType']
 
        LogService.log('Creating sample files')
        oWorkbook = xlsxwriter.Workbook(cDirTemplateSampleFile + cTemplateSampleFile + cFileType)
 
        is_multi_sheet = isinstance(aStructuredHeaderData.get("children"), dict)
        roots = []
 
        if is_multi_sheet:
            for sheet_name, sheet_data in aStructuredHeaderData["children"].items():
                root = importer.import_({"cHeaderLabel": "root", "children": sheet_data})
                roots.append((sheet_name, root))
        else:
            root = importer.import_(aStructuredHeaderData)
            roots.append(("Sheet1", root))
 
        for sheet_name, root in roots:
            oWorksheet = oWorkbook.add_worksheet(sheet_name)
 
            aTemplateHeaderSheet = oTemplateMetaData['aTemplateHeader'].get(sheet_name, [])
            aFilteredParentHeader = [aHeader for aHeader in aTemplateHeaderSheet if aHeader['cParentHeader'] == '']
 
            iRow = 0
            iColStart = 0
            iColEnd = 0
            aParentHeaderList = []
            iGetTotalLeavesOfAttr = 0
            iIternationOfLeaves = 0
            for aHeaderDetails in aFilteredParentHeader:
                aParentHeaderList.append(aHeaderDetails['cHeaderLabel'])
                iSubHeaders = FunDCT_DefineHeaderLevels(root, aHeaderDetails['cHeaderLabel'], 'ParentChild Header Length')
                iColEnd = iColStart + iSubHeaders - 1
 
                oCellFormat = oWorkbook.add_format()
                oCellFormat.set_font_color('#' + aHeaderDetails['cTextColor']['hex'])
                oCellFormat.set_bg_color('#' + aHeaderDetails['cBackColor']['hex'])
                oCellFormat.set_border(1)
                oWorksheet.protect()
 
                if iSubHeaders <= 1:
                    oWorksheet.write(iRow, iColStart, aHeaderDetails['cHeaderLabel'].strip(), oCellFormat)
                    if aHeaderDetails['cComment'] != '':
                        for iIndexRangeColumn in range(iColStart, 0):
                            oWorksheet.write_comment(iRow, iIndexRangeColumn, aHeaderDetails['cComment'])
 
                    iGetTotalLeavesOfAttr = FunDCT_GetLeftMaxNodes(root,aHeaderDetails['cHeaderLabel'])
                    iIternationOfLeaves = 0
                    iHeaderLevel = 0
                    iChildColStart = iColStart
                    iChildColEnd = 0
                    if len(aHeaderDetails['children']) > 0:
                        iChildColStart = iColStart
                        FunDCT_TraverseThroughChildHeaders(root,aHeaderDetails['children'], iHeaderLevel, iChildColStart, iChildColEnd, iGetTotalLeavesOfAttr, iIternationOfLeaves, oWorksheet, oWorkbook)
                else:
                    oCellFormat.set_align('center')
                    oCellFormat.set_valign('vcenter')
                    oCellFormat.set_text_wrap(True)
                    oWorksheet.merge_range(iRow, iColStart, iRow, iColEnd, aHeaderDetails['cHeaderLabel'], oCellFormat)
                    if aHeaderDetails['cComment'] != '':
                        oWorksheet.write_comment(
                        iRow, iColStart, aHeaderDetails['cComment'])
                   
                    iGetTotalLeavesOfAttr = FunDCT_GetLeftMaxNodes(root,aHeaderDetails['cHeaderLabel'])
                    iIternationOfLeaves = 0
                    iHeaderLevel = 0
                    iChildColStart = iColStart
                    iChildColEnd = 0
 
                    if len(aHeaderDetails['children']) > 0:
                        iChildColStart = iColStart
                        FunDCT_TraverseThroughChildHeaders(root,aHeaderDetails['children'], iHeaderLevel, iChildColStart, iChildColEnd, iGetTotalLeavesOfAttr, iIternationOfLeaves, oWorksheet, oWorkbook)

 
                if 'NO_CHANGE' not in aHeaderDetails['cValidations']:
                    cColumnLetter = get_column_letter(iColStart + 1) + '1:' + get_column_letter(iColStart + 1) + '100000'
                    oWorksheet.unprotect_range(cColumnLetter)
 
                iColStart += iSubHeaders
 
        oWorkbook.close()
 
        cUploadFileTo = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_SAMPLEFILE) + '/'
        cFullPath = str(cDirTemplateSampleFile) + str(cTemplateSampleFile) + str(cFileType)
        cS3Url = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cUploadFileTo), cFullPath)
        oDataFrame = pandas.read_excel(cFullPath,engine='openpyxl')
        iIndexDataFrame = iter(range(1, len(oDataFrame.columns) + 1))
        # oDataFrame.columns = [cColumns if not cColumns.startswith('Unnamed') else next(iIndexDataFrame) for cColumns in oDataFrame.columns]
        oDataFrame.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in oDataFrame.columns]
        cUploadedFile = oDataFrame.to_json()
        model_common.FunDCT_InsertTemplateLog('gen_templates',iTemplateID, cUploadedFile, '')
        os.remove(cFullPath) if os.path.exists(cFullPath) else None
 
        oResponse = {
            "cFileName": cTemplateSampleFile,
            "cFileType": cFileType,
            "cDirTemplateSampleFile": cDirTemplateSampleFile,
            "cFullPath": cDirTemplateSampleFile + cTemplateSampleFile + cFileType,
            "iMaxDepthHeaders": iMaxDepthHeaders,
            "cS3Url": cS3Url,
        }
 
        return MessageHandling.FunDCT_MessageHandling('Success' , oResponse)
    except Exception as e:
        LogService.log('Error while creating file' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while creating file' + str(e))


def FunDCT_DefineHeaderLevels(root, cAttribute, cCase, cParrt=''):
    try:
        if cCase == 'ParentChild Header Length':
            oAttrDetails = find_by_attr(root, name="cHeaderLabel", value=cAttribute)
            iLeavesLen = len(oAttrDetails.leaves)
            return iLeavesLen
        elif cParrt!='':
            oAttrDetails = find_by_attr(root, name="cHeaderLabel", value=cParrt,)
            oAttrDetails = find_by_attr(oAttrDetails, name="cHeaderLabel", value=cAttribute,)
            iLeavesLen = len(oAttrDetails.leaves)
            return iLeavesLen
 
        else:
            return 0
    except Exception as e:
        print(f"Error in FunDCT_DefineHeaderLevels: {e}")
        return 0
    


def FunDCT_GetLeftMaxNodes(root,cAttribute):
    try:
        oAttrDetails = find_by_attr(root, name="cHeaderLabel", value=cAttribute)
        return len(oAttrDetails.leaves)
    except Exception as e:
        print(f"Error in FunDCT_GetLeftMaxNodes: {e}")
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while fetching parent headers')
        


def FunDCT_TraverseThroughChildHeaders(root, aChildrenHeadersTraverse, iHeaderLevel, iChildColStart, iChildColEnd, iGetTotalLeavesOfAttr, iIternationOfLeaves, oWorksheet, oWorkbook):
    try:
        iHeaderLevel += 1
        for aChildHeaderDetails in aChildrenHeadersTraverse:
            iChildSubHeaders = FunDCT_DefineHeaderLevels(root, aChildHeaderDetails['cHeaderLabel'], 'Child Header Length', aChildHeaderDetails['cParentHeader'])
            iChildColEnd = iChildColStart + iChildSubHeaders - 1
    
            oChildCellFormat = oWorkbook.add_format()
            oChildCellFormat.set_font_color('#' + aChildHeaderDetails['cTextColor']['hex'])
            oChildCellFormat.set_bg_color('#' + aChildHeaderDetails['cBackColor']['hex'])
            oChildCellFormat.set_border(1)
            oChildCellFormat.set_text_wrap(True)
    
            if iChildSubHeaders <= 1:
                oWorksheet.write(iHeaderLevel, iChildColStart, aChildHeaderDetails['cHeaderLabel'], oChildCellFormat)
                if aChildHeaderDetails['cComment'] != '':
                        oWorksheet.write_comment(
                        iHeaderLevel, iChildColStart, aChildHeaderDetails['cComment'])
                if 'NO_CHANGE' not in aChildHeaderDetails['cValidations']:
                    cChildColumnLetter = get_column_letter(iChildColStart + 1) + '1:' + get_column_letter(iChildColStart + 1) + '100000'
                    oWorksheet.unprotect_range(cChildColumnLetter)
            else:
                oChildCellFormat.set_align('center')
                oChildCellFormat.set_valign('vcenter')
                oWorksheet.merge_range(iHeaderLevel, iChildColStart, iHeaderLevel, iChildColEnd, aChildHeaderDetails['cHeaderLabel'], oChildCellFormat)
                if aChildHeaderDetails['cComment']:
                    for iChildIndexRangeColumn in range(iChildColStart, iChildColEnd):
                        oWorksheet.write_comment(iHeaderLevel, iChildIndexRangeColumn, aChildHeaderDetails['cComment'])
    
            iChildColStart += iChildSubHeaders
            iIternationOfLeaves += 1
    
            if len(aChildHeaderDetails['children']) > 0:
                iIternationOfLeaves = 0
                iGetTotalLeavesOfAttr = FunDCT_GetLeftMaxNodes(root ,aChildHeaderDetails['cHeaderLabel'])
                iNewChildrenColumnsStartall = iChildColStart - iChildSubHeaders
                FunDCT_TraverseThroughChildHeaders(root, aChildHeaderDetails['children'], iHeaderLevel, iNewChildrenColumnsStartall, iChildColEnd, iGetTotalLeavesOfAttr, iIternationOfLeaves, oWorksheet, oWorkbook)
    except Exception as e:
        print(f"Error in FunDCT_TraverseThroughChildHeaders: {e}")