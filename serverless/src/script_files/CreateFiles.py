import sys
import json
import pandas
from openpyxl import load_workbook
import openpyxl
import os
import xlrd as xl
import xlwt
from xlwt import Workbook
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
import xlsxwriter
import pydash
from anytree import Node, RenderTree, AsciiStyle, PreOrderIter, find, LevelOrderIter, find_by_attr, util, AnyNode
from anytree.importer import JsonImporter
from anytree.importer import DictImporter
from anytree.exporter import DotExporter
from anytree import AnyNode, RenderTree
from MessageHandling import MessageHandling
from collections import defaultdict
from shutil import copyfile

from pathlib import Path

# from .. import loadConfig
# from .. import awsCloudFileManager


# import loadConfig

from pyconfig import loadConfig
from pyconfig import awsCloudFileManager
from pyconfig.model_common import model_common_Cls
from pyconfig.LogService import LogService

# Explore Parameteres
oPostParameters = json.loads(sys.argv[1])
# with open(r'c:\vartemp\tempcreatet.txt','w') as f:
#     f.write(str(sys.argv[1]))
# with open(r'c:\vartemp\tempcreatet.txt','r') as f:
#     oPostParameters = json.loads(f.read())
cDirTemplateSampleFile = oPostParameters['cDirTemplateSampleFile']
cTemplateSampleFile = oPostParameters['cTemplateSampleFile']
iTemplateID = oPostParameters['iTemplateID']
cDirConsolidationTemplate = oPostParameters['cDirConsolidationTemplate']

with open(cDirTemplateSampleFile + cTemplateSampleFile + '_oTemplateMetaData.json', "r") as cTemplateMetaData:
    oTemplateMetaData = json.load(cTemplateMetaData)
with open(cDirTemplateSampleFile + cTemplateSampleFile + '_aStructuredHeaderData.json', "r") as cStructuredHeaderData:
    aStructuredHeaderData = json.load(cStructuredHeaderData)

cFileType = oPostParameters['cFileType']
iMaxDepthHeaders = 1
iGetMaxNodesFromHeaderTree = 1

importer = DictImporter()
root = importer.import_(aStructuredHeaderData)
model_common = model_common_Cls()
MessageHandling = MessageHandling()
# Max Siblings
def FunDCT_GetLeftMaxNodes(cAttribute):
    try:
        oAttrDetails = find_by_attr(root, name="cHeaderLabel", value=cAttribute)
        return len(oAttrDetails.leaves)
    except:
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while fetching parent headers')

def FunDCT_DefineHeaderLevels(cAttribute, cCase,cParrt=''):
    try:
        importer = DictImporter()
        root = importer.import_(aStructuredHeaderData)    
        
        if cCase == 'ParentChild Header Length':
            oAttrDetails = find_by_attr(root, name="cHeaderLabel", value=cAttribute,)
            iLeavesLen = len(oAttrDetails.leaves)
            return iLeavesLen
        elif cParrt!='':
            oAttrDetails = find_by_attr(root, name="cHeaderLabel", value=cParrt,)
            oAttrDetails = find_by_attr(oAttrDetails, name="cHeaderLabel", value=cAttribute,)
            iLeavesLen = len(oAttrDetails.leaves)
            return iLeavesLen

        else:
            return 0
    except:
        return 0

# Traverse though children headers Vertically
def FunDCT_TraverseThroughChildHeaders(aChildrenHeadersTraverse, iHeaderLevel, iChildColStart, iChildColEnd, iGetTotalLeavesOfAttr, iIternationOfLeaves, oWorksheet, oWorkbook):
    iHeaderLevel += 1
    for aChildHeaderDetails in aChildrenHeadersTraverse:
        iChildSubHeaders = FunDCT_DefineHeaderLevels(aChildHeaderDetails['cHeaderLabel'], 'Child Header Length',aChildHeaderDetails['cParentHeader'])
        iChildColEnd = iChildColStart + iChildSubHeaders-1

        oChildCellFormat = oWorkbook.add_format()
        oChildCellFormat.set_font_color('#' + aChildHeaderDetails['cTextColor']['hex'])
        oChildCellFormat.set_bg_color('#' + aChildHeaderDetails['cBackColor']['hex'])
        oChildCellFormat.set_border(1)
        oChildCellFormat.set_text_wrap(True)
        
        if iChildSubHeaders <= 1:
            oWorksheet.write(
                iHeaderLevel, iChildColStart, aChildHeaderDetails['cHeaderLabel'], oChildCellFormat)
            if aChildHeaderDetails['cComment'] != '':
                    oWorksheet.write_comment(
                    iHeaderLevel, iChildColStart, aChildHeaderDetails['cComment'])
            if 'NO_CHANGE' not in aChildHeaderDetails['cValidations']:
                cChildColumnLetter = get_column_letter(iChildColStart+1) + '1:' + get_column_letter(iChildColStart+1) + '100000'
                oWorksheet.unprotect_range(cChildColumnLetter)
        else:
            oChildCellFormat.set_align('center')
            oChildCellFormat.set_valign('vcenter')
            oWorksheet.merge_range(
                iHeaderLevel, iChildColStart, iHeaderLevel, iChildColEnd, aChildHeaderDetails['cHeaderLabel'], oChildCellFormat)
            if aChildHeaderDetails['cComment'] != '':
                for iChildIndexRangeColumn in range(iChildColStart, iChildColEnd):
                    oWorksheet.write_comment(iHeaderLevel, iChildIndexRangeColumn, aChildHeaderDetails['cComment'])
        
        iChildColStart += iChildSubHeaders
        iIternationOfLeaves += 1
                            
        if len(aChildHeaderDetails['children']) > 0:
            iIternationOfLeaves = 0
            iGetTotalLeavesOfAttr = FunDCT_GetLeftMaxNodes(aChildHeaderDetails['cHeaderLabel'])
            iNewChildrenColumnsStartall = iChildColStart - iChildSubHeaders
            FunDCT_TraverseThroughChildHeaders(aChildHeaderDetails['children'], iHeaderLevel, iNewChildrenColumnsStartall, iChildColEnd, iGetTotalLeavesOfAttr, iIternationOfLeaves, oWorksheet, oWorkbook)

# Create Sample Template File
def FunDCT_CreateSampleTemplateFile():
    try:
        LogService.log('createing sample files')
        oWorkbook = xlsxwriter.Workbook(
        cDirTemplateSampleFile + cTemplateSampleFile + cFileType)
        oWorksheet = oWorkbook.add_worksheet()

        aFilteredParentHeader = [
            aFilteredDetails for aFilteredDetails in oTemplateMetaData['aTemplateHeader'] if aFilteredDetails['cParentHeader'] == '']
        
        iRow = 0
        iColStart = 0
        iColEnd = 0
        aParentHeaderList = []
        iGetTotalLeavesOfAttr = 0
        iIternationOfLeaves = 0
        for aHeaderDetails in aFilteredParentHeader:
            aParentHeaderList.append(aHeaderDetails['cHeaderLabel'])
            iSubHeaders = FunDCT_DefineHeaderLevels(aHeaderDetails['cHeaderLabel'], 'ParentChild Header Length')
            iColEnd = iColStart + iSubHeaders-1

            oCellFormat = oWorkbook.add_format()
            oCellFormat.set_font_color('#' + aHeaderDetails['cTextColor']['hex'])
            oCellFormat.set_bg_color('#' + aHeaderDetails['cBackColor']['hex'])
            oCellFormat.set_border(1)
            oWorksheet.protect()
             
            if iSubHeaders <= 1:
                oWorksheet.write(
                    iRow, iColStart, aHeaderDetails['cHeaderLabel'].strip(), oCellFormat)
                if aHeaderDetails['cComment'] != '':
                    for iIndexRangeColumn in range(iColStart, 0):
                        oWorksheet.write_comment(iRow, iIndexRangeColumn, aHeaderDetails['cComment'])
                                    
                iGetTotalLeavesOfAttr = FunDCT_GetLeftMaxNodes(aHeaderDetails['cHeaderLabel'])
                iIternationOfLeaves = 0
                
                iHeaderLevel = 0
                iChildColStart = iColStart
                iChildColEnd = 0
                if len(aHeaderDetails['children']) > 0:
                    iChildColStart = iColStart
                    FunDCT_TraverseThroughChildHeaders(aHeaderDetails['children'], iHeaderLevel, iChildColStart, iChildColEnd, iGetTotalLeavesOfAttr, iIternationOfLeaves, oWorksheet, oWorkbook)
            else:
                oCellFormat.set_align('center')
                oCellFormat.set_valign('vcenter')
                oCellFormat.set_text_wrap(True)
                oWorksheet.merge_range(
                    iRow, iColStart, iRow, iColEnd, aHeaderDetails['cHeaderLabel'], oCellFormat)
                if aHeaderDetails['cComment'] != '':
                    oWorksheet.write_comment(
                    iRow, iColStart, aHeaderDetails['cComment'])
                
                iGetTotalLeavesOfAttr = FunDCT_GetLeftMaxNodes(aHeaderDetails['cHeaderLabel'])
                iIternationOfLeaves = 0
                iHeaderLevel = 0
                iChildColStart = iColStart
                iChildColEnd = 0

                if len(aHeaderDetails['children']) > 0:
                    iChildColStart = iColStart
                    FunDCT_TraverseThroughChildHeaders(aHeaderDetails['children'], iHeaderLevel, iChildColStart, iChildColEnd, iGetTotalLeavesOfAttr, iIternationOfLeaves, oWorksheet, oWorkbook)

            if 'NO_CHANGE' not in aHeaderDetails['cValidations']:
                cColumnLetter = get_column_letter(iColStart+1) + '1:' + get_column_letter(iColStart+1) + '100000'
                oWorksheet.unprotect_range(cColumnLetter)

            iColStart += iSubHeaders
        
        oWorkbook.close()

        cUploadFileTo = str(loadConfig.AWS_BUCKET_ENV) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES) + '/' + str(loadConfig.AWS_BUCKET_TEMPLATES_SAMPLEFILE) + '/'
        cFullPath = str(cDirTemplateSampleFile) + str(cTemplateSampleFile) + str(cFileType)
        cS3Url = awsCloudFileManager.upload_file_to_bucket(str(loadConfig.AWS_BUCKET), str(cUploadFileTo), cFullPath)
        os.remove(cDirTemplateSampleFile + cTemplateSampleFile + '_oTemplateMetaData.json') if os.path.exists(cDirTemplateSampleFile + cTemplateSampleFile + '_oTemplateMetaData.json') else None
        os.remove(cDirTemplateSampleFile + cTemplateSampleFile + '_aStructuredHeaderData.json') if os.path.exists(cDirTemplateSampleFile + cTemplateSampleFile + '_aStructuredHeaderData.json') else None
        oDataFrame = pandas.read_excel(cFullPath,engine='openpyxl')
        iIndexDataFrame = iter(range(1, len(oDataFrame.columns) + 1))
        # oDataFrame.columns = [cColumns if not cColumns.startswith('Unnamed') else next(iIndexDataFrame) for cColumns in oDataFrame.columns]
        oDataFrame.columns = [cColumns if not cColumns.startswith('Unnamed') else '#DCT#' + str(next(iIndexDataFrame)) for cColumns in oDataFrame.columns]
        cUploadedFile = oDataFrame.to_json()
        model_common.FunDCT_InsertTemplateLog('gen_templates',iTemplateID, cUploadedFile, '')
        os.remove(cFullPath) if os.path.exists(cFullPath) else None

        oResponse = {
            "cFileName": cTemplateSampleFile,
            "cFileType": cFileType,
            "cDirTemplateSampleFile": cDirTemplateSampleFile,
            "cFullPath": cDirTemplateSampleFile + cTemplateSampleFile + cFileType,
            "iMaxDepthHeaders": iMaxDepthHeaders,
            "cS3Url": cS3Url,
        }

        return MessageHandling.FunDCT_MessageHandling('Success', oResponse)
    except Exception as e:
        LogService.log('Error while creating file' + str(e))
        return MessageHandling.FunDCT_MessageHandling('Error', 'Error while creating file' + str(e))

FunDCT_CreateSampleTemplateFile()
