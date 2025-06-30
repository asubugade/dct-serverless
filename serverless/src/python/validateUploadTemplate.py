import openpyxl
from  .parseFiles import  parseTemplateFileToJSON
import re
from io import BytesIO
import json
from fastapi import FastAPI, HTTPException
import os,sys
from pathlib import Path
app = FastAPI()

from .MessageHandling import MessageHandling
from pyconfig.LogService import LogService
from pyconfig import loadConfig
MessageHandling = MessageHandling()

PY_VALIDATION_SEPARATOR = loadConfig.PY_VALIDATION_SEPARATOR
PY_RANGE_VALUE_SEPARATOR = loadConfig.PY_RANGE_VALUE_SEPARATOR
PY_BRACKET_VALIDATION_SEPARATOR = loadConfig.PY_BRACKET_VALIDATION_SEPARATOR
LogService.log(f'pyvalidatate seperatetor ==> {PY_VALIDATION_SEPARATOR}')
LogService.log(f'PY_RANGE_VALUE_SEPARATOR seperatetor ==> {PY_RANGE_VALUE_SEPARATOR}')
LogService.log(f'PY_BRACKET_VALIDATION_SEPARATOR seperatetor ==> {PY_BRACKET_VALIDATION_SEPARATOR}')

@app.post("/validateuploadtemplate")
def handler(event: dict, context=None):
    try:
        data = event if isinstance(event, dict) else json.loads(event)
        exData = json.loads(data['body']) if isinstance(data['body'], str) else data['body']
        
        cFileContent = exData['cFileContent']['data']
        aValidateRules = exData['aValidateRules']
        aValidateRulesAll = exData['aValidateRulesAll']

        # Read Excel file
        file_stream = BytesIO(bytes(cFileContent))
        workbook = openpyxl.load_workbook(file_stream)
        oSheet = workbook.active
        if oSheet.max_row == 3 or oSheet.max_row ==2:
            bValidation = True
            for j in range(1, oSheet.max_column+1):
                cell_obj = oSheet.cell(row=oSheet.max_row, column=j)
                cColumnValue = str(cell_obj.value).strip()
                if re.findall('[^a-zA-Z_]', cColumnValue):
                    bValidation = FunDCT_CheckMultipleValidations(
                        cColumnValue, aValidateRules, aValidateRulesAll)
                    if bValidation == False:
                        break

                elif cell_obj.value not in aValidateRules:
                    bValidation = False
                    break
            # LogService.log(f'Validation istatus == {bValidation}')
            if bValidation:
                return parseTemplateFileToJSON(event)
            else:
                LogService.log(f'ErrorValidation ==> {cColumnValue}')
                return MessageHandling.FunDCT_MessageHandling('ErrorValidation', cColumnValue)

        else:
            return MessageHandling.FunDCT_MessageHandling('ErrorValidation', 'sheet should have only 3 rows ')

    except Exception as e:
        LogService.log(e)
        print(e)


def FunDCT_CheckMultipleValidations(cColumnValue, aValidateRules, aValidateRulesAll):
    aValidations = cColumnValue.replace(PY_VALIDATION_SEPARATOR, ',').replace(
        PY_RANGE_VALUE_SEPARATOR, ',').replace(PY_BRACKET_VALIDATION_SEPARATOR, ',')
    aValidationFile = re.sub('[^a-zA-Z_]', ' ', aValidations).split(" ")
    # to convert string validations to list and to Filter out ' ' .
    liFileValidationsRaw = list(filter(None, aValidationFile))
    liFileValidations = []

    ##
    # loop to remove NS,NA,OR from validation provided in file such as MIXED_DATA(DEFAULT_LOCATION%%NS&&NA&&OR) or
    # to remove USD from MIXED_DATA(DEFAULT_CURRENCY%%USD)
    # to remove SUN,MON,TUE,WED from MIXED_DATA(SUN&&MON&&TUE&&WED)
    # and to append the remaining strings in liFileValidations list as ['MIXED_DATA','DEFAULT_LOCATION'] ,so that it can be cheecked
    # against validation rules present in db (aValidateRules) agains a provided template type.
    ##
    for av in aValidateRulesAll:
        for v in liFileValidationsRaw:
            if v == av:
                liFileValidations.append(v)

    # loop to check validations provided in file against the one present in db for a provided template type
    ##
    for v in liFileValidations:
        if v not in aValidateRules:
            return False

    return True
